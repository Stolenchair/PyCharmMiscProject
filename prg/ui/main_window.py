"""Main window for PRG Pipeline Manager - Classic Design."""

import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from typing import Optional, Dict, List, Any
from pathlib import Path


class PRGPipelineManager:
    """
    Менеджер привязок ПРГ к потребителям
    Версия: 7.4 - Modular Architecture
    """

    def __init__(self, root, settings_manager, excel_loader,
                 validation_service, calculation_service,
                 binding_service, search_service, style_manager):
        """Initialize main window with injected services."""
        self.root = root
        self.settings_manager = settings_manager
        self.excel_loader = excel_loader
        self.validation_service = validation_service
        self.calculation_service = calculation_service
        self.binding_service = binding_service
        self.search_service = search_service
        self.style_manager = style_manager

        # Data storage
        self.excel_path: Optional[Path] = None
        self.prg_data: List[Dict[str, Any]] = []
        self.grs_data: List[Dict[str, Any]] = []
        self.consumer_data: List[Dict[str, Any]] = []
        self.changes: Dict[str, Dict[str, Any]] = {}

        # Selected elements
        self.selected_prg = None
        self.selected_consumer = None

        # Setup UI
        self.root.title("PRG Pipeline Manager v7.4 - Professional Edition")

        # Load saved geometry or use default
        saved_geometry = self.settings_manager.get_ui_preference('window_geometry', '1500x900')
        self.root.geometry(saved_geometry)

        # Apply theme
        self.style_manager.apply()
        self._apply_window_theme()

        self.create_menu()
        self.create_top_panel()
        self.create_main_area()
        self.create_status_panel()

        # Handle window close
        self.root.protocol("WM_DELETE_WINDOW", self.on_close_window)

        print(f"[OK] Main window created with {self.style_manager.get_theme()} theme")

    def _apply_window_theme(self):
        """Apply theme colors to root window."""
        colors = self.style_manager.colors
        self.root.configure(bg=colors['bg'])

    def create_menu(self):
        """Создание главного меню"""
        colors = self.style_manager.colors

        menubar = tk.Menu(self.root, bg=colors['bg_panel'], fg=colors['text'],
                         activebackground=colors['primary'], activeforeground='white')
        self.root.config(menu=menubar)

        # Файл
        file_menu = tk.Menu(menubar, tearoff=0, bg=colors['bg_panel'], fg=colors['text'],
                           activebackground=colors['primary'], activeforeground='white')
        menubar.add_cascade(label="Файл", menu=file_menu)
        file_menu.add_command(label="Открыть Excel...", command=self.open_excel_file)
        file_menu.add_separator()
        file_menu.add_command(label="Сохранить изменения", command=self.save_changes_to_excel)
        file_menu.add_separator()
        file_menu.add_command(label="Выход", command=self.on_close_window)

        # Вид
        view_menu = tk.Menu(menubar, tearoff=0, bg=colors['bg_panel'], fg=colors['text'],
                           activebackground=colors['primary'], activeforeground='white')
        menubar.add_cascade(label="Вид", menu=view_menu)
        theme_label = "Темная тема" if self.style_manager.get_theme() == 'light' else "Светлая тема"
        view_menu.add_command(label=theme_label, command=self.toggle_theme)

        # Настройки
        settings_menu = tk.Menu(menubar, tearoff=0, bg=colors['bg_panel'], fg=colors['text'],
                               activebackground=colors['primary'], activeforeground='white')
        menubar.add_cascade(label="Настройки", menu=settings_menu)
        settings_menu.add_command(label="Настройки столбцов", command=self.open_settings_dialog)

        # Инструменты
        tools_menu = tk.Menu(menubar, tearoff=0, bg=colors['bg_panel'], fg=colors['text'],
                            activebackground=colors['primary'], activeforeground='white')
        menubar.add_cascade(label="Инструменты", menu=tools_menu)
        tools_menu.add_command(label="Привязать по поиску", command=self.bind_by_search)
        tools_menu.add_command(label="Привязать вручную", command=self.bind_manually)
        tools_menu.add_separator()
        tools_menu.add_command(label="Автопривязка ПРГ", command=self.auto_bind_all_prg)
        tools_menu.add_separator()
        tools_menu.add_command(label="Редактировать доли", command=self.edit_consumer_shares)
        tools_menu.add_command(label="Проверить доли всех", command=self.check_all_consumer_shares)
        tools_menu.add_separator()
        tools_menu.add_command(label="Подсчитать нагрузку ПРГ", command=self.calculate_prg_load)
        tools_menu.add_separator()
        tools_menu.add_command(label="Показать непривязанные", command=self.show_unbound_analysis)
        tools_menu.add_command(label="Показать без расходов", command=self.show_no_expenses_analysis)

    def create_top_panel(self):
        """Создание верхней панели"""
        colors = self.style_manager.colors

        top_frame = tk.Frame(self.root, bg=colors['bg_secondary'], height=100)
        top_frame.pack(fill=tk.X, padx=12, pady=8)
        top_frame.pack_propagate(False)

        file_frame = tk.Frame(top_frame, bg=colors['bg_secondary'])
        file_frame.pack(fill=tk.X, padx=8, pady=12)

        # Open file button
        open_btn = self.style_manager.create_button(
            file_frame,
            text="Открыть Excel файл",
            command=self.open_excel_file,
            color='success'
        )
        open_btn.pack(side=tk.LEFT, padx=(0, 15))

        self.file_label = tk.Label(file_frame, text="Файл не выбран",
                                   bg=colors['bg_secondary'], fg=colors['text'],
                                   font=('Segoe UI', 10))
        self.file_label.pack(side=tk.LEFT)

        # Save button
        self.save_button = self.style_manager.create_button(
            file_frame,
            text="Сохранить изменения",
            command=self.save_changes_to_excel,
            color='warning',
            state=tk.DISABLED
        )
        self.save_button.pack(side=tk.RIGHT, padx=(15, 0))

        self.changes_label = tk.Label(file_frame, text="",
                                      bg=colors['bg_secondary'], fg=colors['danger'],
                                      font=('Segoe UI', 10, 'bold'))
        self.changes_label.pack(side=tk.RIGHT, padx=(15, 15))

    def create_main_area(self):
        """Создание основной рабочей области"""
        colors = self.style_manager.colors

        main_frame = tk.Frame(self.root, bg=colors['bg'])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)

        # Левая панель - ПРГ
        prg_frame = tk.LabelFrame(main_frame, text="ПРГ",
                                  bg=colors['bg'], fg=colors['text'],
                                  font=('Segoe UI', 11, 'bold'),
                                  borderwidth=1, relief='solid',
                                  highlightbackground=colors['border'],
                                  highlightthickness=1)
        prg_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 8))

        prg_tree_frame = tk.Frame(prg_frame, bg=colors['bg'])
        prg_tree_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        self.prg_tree = ttk.Treeview(prg_tree_frame, columns=('prg_id', 'grs_id'),
                                     height=30, style='Modern.Treeview')
        self.prg_tree.heading('#0', text='Структура ПРГ')
        self.prg_tree.heading('prg_id', text='ПРГ ID')
        self.prg_tree.heading('grs_id', text='ГРС ID')
        self.prg_tree.column('#0', width=280)
        self.prg_tree.column('prg_id', width=80)
        self.prg_tree.column('grs_id', width=80)

        prg_scroll = ttk.Scrollbar(prg_tree_frame, orient=tk.VERTICAL, command=self.prg_tree.yview,
                                   style='Modern.Vertical.TScrollbar')
        self.prg_tree.configure(yscrollcommand=prg_scroll.set)

        self.prg_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        prg_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.prg_tree.bind('<<TreeviewSelect>>', self.on_prg_tree_select)

        # Центральная панель с кнопками
        center_frame = tk.Frame(main_frame, bg=colors['bg'], width=180)
        center_frame.pack(side=tk.LEFT, fill=tk.Y, padx=16)
        center_frame.pack_propagate(False)

        tk.Frame(center_frame, bg=colors['bg'], height=10).pack()

        # Кнопки действий
        self.bind_button = self.style_manager.create_button(
            center_frame, text="Привязать ко\nвсему НП",
            command=self.bind_prg_to_settlement, color='success',
            state=tk.DISABLED, height=3, width=16
        )
        self.bind_button.pack(pady=5)

        self.search_bind_button = self.style_manager.create_button(
            center_frame, text="Привязать\nпо поиску",
            command=self.bind_by_search, color='secondary',
            state=tk.DISABLED, height=3, width=16
        )
        self.search_bind_button.pack(pady=5)

        self.manual_bind_button = self.style_manager.create_button(
            center_frame, text="Привязать\nвручную",
            command=self.bind_manually, color='primary',
            state=tk.DISABLED, height=3, width=16
        )
        self.manual_bind_button.pack(pady=5)

        self.unbind_settlement_button = self.style_manager.create_button(
            center_frame, text="Отвязать\nвесь НП",
            command=self.unbind_entire_settlement, color='warning',
            state=tk.DISABLED, height=3, width=16
        )
        self.unbind_settlement_button.pack(pady=5)

        self.auto_bind_button = self.style_manager.create_button(
            center_frame, text="Авто-\nпривязка",
            command=self.auto_bind_all_prg, color='purple',
            state=tk.DISABLED, height=3, width=16
        )
        self.auto_bind_button.pack(pady=5)

        self.edit_shares_button = self.style_manager.create_button(
            center_frame, text="Редактировать\nдоли",
            command=self.edit_consumer_shares, color='primary',
            state=tk.DISABLED, height=3, width=16
        )
        self.edit_shares_button.pack(pady=5)

        self.unbind_button = self.style_manager.create_button(
            center_frame, text="Отвязать\nпотребителя",
            command=self.unbind_single_consumer, color='danger',
            state=tk.DISABLED, height=3, width=16
        )
        self.unbind_button.pack(pady=5)

        self.calculate_load_button = self.style_manager.create_button(
            center_frame, text="Подсчитать\nнагрузку ПРГ",
            command=self.calculate_prg_load, color='purple',
            state=tk.DISABLED, height=3, width=16
        )
        self.calculate_load_button.pack(pady=5)

        # Правая панель - Потребители
        consumer_frame = tk.LabelFrame(main_frame, text="Потребители (🟡 - без ПРГ или доли>1, 🔵 - доли<1, 🚫 - без расходов)",
                                       bg=colors['bg'], fg=colors['text'],
                                       font=('Segoe UI', 11, 'bold'),
                                       borderwidth=1, relief='solid',
                                       highlightbackground=colors['border'],
                                       highlightthickness=1)
        consumer_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(8, 0))

        consumer_tree_frame = tk.Frame(consumer_frame, bg=colors['bg'])
        consumer_tree_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        self.consumer_tree = ttk.Treeview(consumer_tree_frame, columns=('type', 'binding', 'total_share'),
                                         height=30, style='Modern.Treeview')
        self.consumer_tree.heading('#0', text='Потребители')
        self.consumer_tree.heading('type', text='Тип')
        self.consumer_tree.heading('binding', text='Привязки к ПРГ')
        self.consumer_tree.heading('total_share', text='Сумма долей')
        self.consumer_tree.column('#0', width=220)
        self.consumer_tree.column('type', width=90)
        self.consumer_tree.column('binding', width=250)
        self.consumer_tree.column('total_share', width=110)

        consumer_scroll = ttk.Scrollbar(consumer_tree_frame, orient=tk.VERTICAL, command=self.consumer_tree.yview,
                                       style='Modern.Vertical.TScrollbar')
        self.consumer_tree.configure(yscrollcommand=consumer_scroll.set)

        self.consumer_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        consumer_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.consumer_tree.bind('<<TreeviewSelect>>', self.on_consumer_tree_select)

    def create_status_panel(self):
        """Создание нижней панели статуса"""
        colors = self.style_manager.colors

        status_frame = tk.Frame(self.root, bg=colors['bg_secondary'], height=140)
        status_frame.pack(fill=tk.X, padx=12, pady=8)
        status_frame.pack_propagate(False)

        # Верхняя часть - основная информация
        info_frame = tk.Frame(status_frame, bg=colors['bg_secondary'])
        info_frame.pack(fill=tk.X, padx=10, pady=8)

        self.info_label = tk.Label(info_frame,
                                   text="v7.4 Professional Edition - Модульная архитектура с внедрением зависимостей",
                                   bg=colors['bg_secondary'], fg=colors['text_secondary'],
                                   font=('Segoe UI', 10), anchor=tk.W)
        self.info_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.stats_label = tk.Label(info_frame, text="ПРГ: 0 | ГРС: 0 | Потребители: 0",
                                    bg=colors['bg_secondary'], fg=colors['text'],
                                    font=('Segoe UI', 10, 'bold'))
        self.stats_label.pack(side=tk.RIGHT)

        # Детальная информационная панель
        detail_frame = tk.LabelFrame(status_frame, text="Детальная информация (можно выделять и копировать)",
                                     bg=colors['bg_secondary'], fg=colors['text'],
                                     font=('Segoe UI', 10, 'bold'),
                                     borderwidth=1, relief='solid')
        detail_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Text виджет с возможностью выделения
        self.detail_text = tk.Text(detail_frame, height=5, wrap=tk.WORD,
                                   font=('Segoe UI', 10),
                                   bg=colors['card'], fg=colors['text'],
                                   state=tk.DISABLED, cursor="arrow",
                                   borderwidth=0, highlightthickness=0)
        detail_scroll = ttk.Scrollbar(detail_frame, orient=tk.VERTICAL, command=self.detail_text.yview,
                                     style='Modern.Vertical.TScrollbar')
        self.detail_text.configure(yscrollcommand=detail_scroll.set)

        self.detail_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        detail_scroll.pack(side=tk.RIGHT, fill=tk.Y, pady=5)

        # Возможность выделения текста
        self.detail_text.bind("<Button-1>", lambda e: self.detail_text.config(state=tk.NORMAL))
        self.detail_text.bind("<FocusOut>", lambda e: self.detail_text.config(state=tk.DISABLED))

        # Контекстное меню
        self.create_detail_context_menu()

    def create_detail_context_menu(self):
        """Создание контекстного меню для детальной панели"""
        colors = self.style_manager.colors
        self.detail_context_menu = tk.Menu(self.root, tearoff=0,
                                          bg=colors['bg_panel'], fg=colors['text'],
                                          activebackground=colors['primary'], activeforeground='white')
        self.detail_context_menu.add_command(label="Копировать", command=self.copy_selected_text)
        self.detail_context_menu.add_command(label="Выделить всё", command=self.select_all_text)

        self.detail_text.bind("<Button-3>", self.show_detail_context_menu)

    def show_detail_context_menu(self, event):
        """Показ контекстного меню"""
        try:
            self.detail_context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.detail_context_menu.grab_release()

    def copy_selected_text(self):
        """Копирование выделенного текста"""
        try:
            selected = self.detail_text.selection_get()
            self.root.clipboard_clear()
            self.root.clipboard_append(selected)
            print("✅ Текст скопирован в буфер обмена")
        except tk.TclError:
            try:
                all_text = self.detail_text.get(1.0, tk.END).strip()
                if all_text:
                    self.root.clipboard_clear()
                    self.root.clipboard_append(all_text)
                    print("✅ Весь текст скопирован в буфер обмена")
            except:
                print("⚠️ Нет текста для копирования")

    def select_all_text(self):
        """Выделение всего текста"""
        self.detail_text.config(state=tk.NORMAL)
        self.detail_text.tag_add(tk.SEL, "1.0", tk.END)
        self.detail_text.mark_set(tk.INSERT, "1.0")
        self.detail_text.see(tk.INSERT)

    # === FILE OPERATIONS ===

    def open_excel_file(self):
        """Открытие Excel файла"""
        try:
            file_path = filedialog.askopenfilename(
                title="Выберите Excel файл",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )

            if file_path:
                self.excel_path = Path(file_path)
                self.file_label.config(text=f"📄 {self.excel_path.name}")

                self.clear_all_changes()
                self.show_settings_dialog()

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть файл: {str(e)}")

    def show_settings_dialog(self):
        """Диалог настройки таблиц перед загрузкой"""
        # Ask user if they want to check settings first
        response = messagebox.askyesnocancel(
            "Загрузка данных",
            f"Файл: {self.excel_path.name}\n\n"
            f"Загрузить данные с текущими настройками?\n\n"
            f"Да - загрузить сейчас\n"
            f"Нет - открыть настройки столбцов\n"
            f"Отмена - отменить открытие файла"
        )

        if response is None:  # Cancel
            self.excel_path = None
            self.file_label.config(text="Файл не выбран")
            return
        elif response is False:  # No - show settings
            self.open_settings_dialog()
            return

        # Yes - load data
        try:
            self.load_all_data()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка загрузки данных: {str(e)}")

    def open_settings_dialog(self):
        """Открыть диалог настройки столбцов"""
        try:
            from prg.ui.dialogs import SettingsDialog

            dialog = SettingsDialog(
                self.root,
                self.settings_manager,
                self.style_manager
            )

            # If settings were saved and file is loaded, ask to reload
            if dialog.result and self.excel_path:
                response = messagebox.askyesno(
                    "Перезагрузка данных",
                    "Настройки сохранены!\n\n"
                    "Перезагрузить данные из текущего файла с новыми настройками?"
                )

                if response:
                    try:
                        self.load_all_data()
                    except Exception as e:
                        messagebox.showerror("Ошибка", f"Ошибка загрузки данных: {str(e)}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка открытия диалога настроек:\n\n{str(e)}")
            import traceback
            traceback.print_exc()

    def load_all_data(self):
        """Загрузка всех данных из Excel"""
        try:
            print(f"[INFO] Loading data from: {self.excel_path}")

            # Load all data using ExcelLoader service
            data = self.excel_loader.load_all_data(str(self.excel_path))

            self.prg_data = data.get('prg', [])
            self.grs_data = data.get('grs', [])
            self.consumer_data = data.get('consumers', [])

            print(f"[OK] Loaded: {len(self.prg_data)} PRG, {len(self.grs_data)} GRS, {len(self.consumer_data)} consumers")

            # Update UI
            self.populate_prg_tree()
            self.populate_consumer_tree()
            self.update_statistics()
            self.update_button_states()

            # Show statistics
            self.show_load_statistics()

        except Exception as e:
            raise Exception(f"Ошибка загрузки данных: {str(e)}")

    def show_load_statistics(self):
        """Показать статистику загруженных данных"""
        unbound_prg = self.validation_service.find_unbound_prg(self.prg_data, self.consumer_data)
        unbound_consumers = self.validation_service.find_unbound_consumers(self.consumer_data)
        no_expenses = self.validation_service.find_consumers_without_expenses(self.consumer_data)

        # Count population and organizations
        population_count = sum(1 for c in self.consumer_data if c.get('type') == 'Население')
        organization_count = sum(1 for c in self.consumer_data if c.get('type') == 'Организация')

        # Get sheet names from settings
        prg_sheet = self.settings_manager.get_table_settings('prg')['sheet']
        grs_sheet = self.settings_manager.get_table_settings('grs')['sheet']
        pop_sheet = self.settings_manager.get_table_settings('population')['sheet']
        org_sheet = self.settings_manager.get_table_settings('organizations')['sheet']

        message = f"""✅ ДАННЫЕ УСПЕШНО ЗАГРУЖЕНЫ v7.4!

📂 ИСТОЧНИКИ ДАННЫХ:
• Лист "{prg_sheet}": ПРГ ({len(self.prg_data)})
• Лист "{grs_sheet}": ГРС ({len(self.grs_data)})
• Лист "{pop_sheet}": Население ({population_count})
• Лист "{org_sheet}": Организации ({organization_count})

📊 ОСНОВНАЯ СТАТИСТИКА:
• ПРГ: {len(self.prg_data)}
• ГРС: {len(self.grs_data)}
• Потребители: {len(self.consumer_data)} (Население: {population_count}, Организации: {organization_count})

🔍 АНАЛИЗ ПРИВЯЗОК:
• ПРГ без потребителей: {len(unbound_prg)}
• Потребители без ПРГ: {len(unbound_consumers)}
• Потребители без расходов: {len(no_expenses)}

📊 ДОСТУПНЫЕ ФУНКЦИИ:
• 📊 Подсчет нагрузки ПРГ
• 🔗 Привязка/отвязка потребителей
• 💾 Сохранение изменений"""

        messagebox.showinfo("Данные загружены v7.4", message)

    def clear_all_changes(self):
        """Очистка всех изменений"""
        self.changes.clear()
        self.update_changes_display()

    def save_changes_to_excel(self):
        """Сохранение изменений в Excel"""
        if not self.excel_path:
            messagebox.showwarning("Предупреждение", "Нет загруженного файла")
            return

        if not self.changes:
            messagebox.showinfo("Информация", "Нет изменений для сохранения")
            return

        # Ask for confirmation
        response = messagebox.askyesno(
            "Сохранение изменений",
            f"Сохранить {len(self.changes)} изменений в файл:\n\n"
            f"{self.excel_path.name}\n\n"
            f"ВНИМАНИЕ: Изменения будут записаны в файл.\n"
            f"Рекомендуется сделать резервную копию.\n\n"
            f"Продолжить?"
        )

        if not response:
            return

        try:
            from openpyxl import load_workbook
            from prg.utils.excel_utils import index_to_col

            print(f"[INFO] Saving {len(self.changes)} changes to {self.excel_path}...")

            # Load workbook
            wb = load_workbook(str(self.excel_path))

            success_count = 0
            error_count = 0
            errors = []

            # Group changes by sheet for efficiency
            changes_by_sheet = {}
            for change_id, change in self.changes.items():
                sheet_name = change.get('sheet_name', '')
                if sheet_name not in changes_by_sheet:
                    changes_by_sheet[sheet_name] = []
                changes_by_sheet[sheet_name].append(change)

            # Apply changes
            for sheet_name, sheet_changes in changes_by_sheet.items():
                try:
                    if sheet_name not in wb.sheetnames:
                        print(f"[WARNING] Sheet '{sheet_name}' not found")
                        for change in sheet_changes:
                            errors.append(f"Sheet not found: {sheet_name}")
                            error_count += 1
                        continue

                    ws = wb[sheet_name]

                    for change in sheet_changes:
                        try:
                            row = change.get('row', 0) + 1  # Convert to 1-based Excel row
                            col = change.get('col', 0) + 1  # Convert to 1-based Excel column
                            new_value = change.get('new_value', '')

                            # Write to cell
                            ws.cell(row=row, column=col, value=new_value)
                            success_count += 1

                            print(f"  [OK] {sheet_name}!{index_to_col(col-1)}{row} = '{new_value[:30]}...' " if len(str(new_value)) > 30 else f"  [OK] {sheet_name}!{index_to_col(col-1)}{row} = '{new_value}'")

                        except Exception as e:
                            error_msg = f"{sheet_name} row {row}: {str(e)}"
                            errors.append(error_msg)
                            error_count += 1
                            print(f"  [ERROR] {error_msg}")

                except Exception as e:
                    error_msg = f"Sheet {sheet_name}: {str(e)}"
                    errors.append(error_msg)
                    error_count += len(sheet_changes)
                    print(f"[ERROR] {error_msg}")

            # Save workbook
            wb.save(str(self.excel_path))
            wb.close()

            # Clear saved changes
            if success_count > 0:
                self.changes.clear()
                self.update_changes_display()

            # Show result
            if error_count == 0:
                messagebox.showinfo(
                    "Успех",
                    f"Все изменения успешно сохранены!\n\n"
                    f"Сохранено изменений: {success_count}\n"
                    f"Файл: {self.excel_path.name}"
                )
                print(f"[OK] All {success_count} changes saved successfully")
            else:
                error_display = "\n".join(errors[:5])
                if len(errors) > 5:
                    error_display += f"\n... и еще {len(errors) - 5} ошибок"

                messagebox.showwarning(
                    "Частичное сохранение",
                    f"Сохранено изменений: {success_count}\n"
                    f"Ошибок: {error_count}\n\n"
                    f"Ошибки:\n{error_display}"
                )
                print(f"[WARNING] {success_count} saved, {error_count} errors")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка сохранения файла:\n\n{str(e)}")
            print(f"[ERROR] {e}")
            import traceback
            traceback.print_exc()

    # === TREE POPULATION ===

    def populate_prg_tree(self):
        """Заполнение дерева ПРГ"""
        self.prg_tree.delete(*self.prg_tree.get_children())

        if not self.prg_data:
            return

        # Group by district -> settlement
        structure = {}
        for prg in self.prg_data:
            mo = prg.get('mo', '')
            settlement = prg.get('settlement', '')

            if mo not in structure:
                structure[mo] = {}
            if settlement not in structure[mo]:
                structure[mo][settlement] = []

            structure[mo][settlement].append(prg)

        # Build tree
        for mo in sorted(structure.keys()):
            mo_node = self.prg_tree.insert('', 'end', text=f"📍 {mo}", values=('', ''))

            for settlement in sorted(structure[mo].keys()):
                settlement_node = self.prg_tree.insert(mo_node, 'end', text=f"🏘️ {settlement}", values=('', ''))

                for prg in structure[mo][settlement]:
                    prg_id = prg.get('prg_id', '')
                    grs_id = prg.get('grs_id', '')

                    display_text = f"🏭 {prg_id}"

                    self.prg_tree.insert(settlement_node, 'end', text=display_text,
                                       values=(prg_id, grs_id),
                                       tags=(prg['id'],))

        print(f"[OK] PRG tree populated with {len(self.prg_data)} items")

    def populate_consumer_tree(self):
        """Заполнение дерева потребителей"""
        self.consumer_tree.delete(*self.consumer_tree.get_children())

        if not self.consumer_data:
            return

        # Group by district -> settlement -> type
        structure = {}

        for consumer in self.consumer_data:
            c_type = consumer.get('type', '')
            mo = consumer.get('mo', '')
            settlement = consumer.get('settlement', '')

            if mo not in structure:
                structure[mo] = {}
            if settlement not in structure[mo]:
                structure[mo][settlement] = {'Население': [], 'Организация': []}

            if c_type in structure[mo][settlement]:
                structure[mo][settlement][c_type].append(consumer)

        # Build tree
        from prg.data.parsers import parse_prg_bindings, calculate_total_share

        for mo in sorted(structure.keys()):
            mo_node = self.consumer_tree.insert('', 'end', text=f"📍 {mo}", values=('', '', ''))

            for settlement in sorted(structure[mo].keys()):
                settlement_node = self.consumer_tree.insert(mo_node, 'end', text=f"🏘️ {settlement}",
                                                          values=('', '', ''))

                # Add population consumers
                for consumer in structure[mo][settlement]['Население']:
                    name = consumer.get('name', consumer.get('settlement', ''))
                    code = consumer.get('code', '')

                    # Parse bindings
                    bindings = parse_prg_bindings(code)
                    total_share = calculate_total_share(bindings)

                    # Check expenses
                    has_expenses = self.validation_service.has_expenses(consumer)

                    # Display icon based on state
                    icon = "👤"
                    if not has_expenses:
                        icon = "🚫"
                    elif not bindings:
                        icon = "🟡"
                    elif total_share > 1.01:  # Sum > 1 with tolerance
                        icon = "🟡"
                    elif total_share < 0.99 and bindings:  # Sum < 1 with tolerance
                        icon = "🔵"

                    display_text = f"{icon} {name}"
                    binding_display = f"{len(bindings)} привязок" if bindings else "Нет"
                    share_display = f"{total_share:.2f}" if bindings else ""

                    self.consumer_tree.insert(settlement_node, 'end', text=display_text,
                                            values=('Население', binding_display, share_display),
                                            tags=(consumer['id'],))

                # Add organization consumers
                for consumer in structure[mo][settlement]['Организация']:
                    name = consumer.get('name', '')
                    code = consumer.get('code', '')

                    # Parse bindings
                    bindings = parse_prg_bindings(code)
                    total_share = calculate_total_share(bindings)

                    # Check expenses
                    has_expenses = self.validation_service.has_expenses(consumer)

                    # Display icon based on state
                    icon = "🏢"
                    if not has_expenses:
                        icon = "🚫"
                    elif not bindings:
                        icon = "🟡"
                    elif total_share > 1.01:  # Sum > 1 with tolerance
                        icon = "🟡"
                    elif total_share < 0.99 and bindings:  # Sum < 1 with tolerance
                        icon = "🔵"

                    display_text = f"{icon} {name}"
                    binding_display = f"{len(bindings)} привязок" if bindings else "Нет"
                    share_display = f"{total_share:.2f}" if bindings else ""

                    self.consumer_tree.insert(settlement_node, 'end', text=display_text,
                                            values=('Организация', binding_display, share_display),
                                            tags=(consumer['id'],))

        print(f"[OK] Consumer tree populated with {len(self.consumer_data)} items")

    # === EVENT HANDLERS ===

    def on_prg_tree_select(self, event):
        """Обработка выбора ПРГ"""
        selection = self.prg_tree.selection()
        if not selection:
            self.selected_prg = None
            return

        item = selection[0]
        tags = self.prg_tree.item(item, 'tags')

        if tags:
            prg_id = tags[0]
            self.selected_prg = next((p for p in self.prg_data if p['id'] == prg_id), None)

            if self.selected_prg:
                self.update_detail_panel_prg(self.selected_prg)
                self.update_button_states()

    def on_consumer_tree_select(self, event):
        """Обработка выбора потребителя"""
        selection = self.consumer_tree.selection()
        if not selection:
            self.selected_consumer = None
            return

        item = selection[0]
        tags = self.consumer_tree.item(item, 'tags')

        if tags:
            consumer_id = tags[0]
            self.selected_consumer = next((c for c in self.consumer_data if c['id'] == consumer_id), None)

            if self.selected_consumer:
                self.update_detail_panel_consumer(self.selected_consumer)
                self.update_button_states()

    # === DETAIL PANEL UPDATES ===

    def update_detail_panel_prg(self, prg):
        """Обновление панели деталей для ПРГ"""
        self.detail_text.config(state=tk.NORMAL)
        self.detail_text.delete(1.0, tk.END)

        details = f"Выбран ПРГ:\n\n"
        details += f"ПРГ ID: {prg.get('prg_id', '')}\n"
        details += f"Район: {prg.get('mo', '')}\n"
        details += f"НП: {prg.get('settlement', '')}\n"
        details += f"ГРС ID: {prg.get('grs_id', '')}\n"
        details += f"\nЛист Excel: {prg.get('sheet_name', '')}\n"
        details += f"Строка Excel: {prg.get('excel_row', '')}\n"

        self.detail_text.insert(1.0, details)
        self.detail_text.config(state=tk.DISABLED)

    def update_detail_panel_consumer(self, consumer):
        """Обновление панели деталей для потребителя"""
        self.detail_text.config(state=tk.NORMAL)
        self.detail_text.delete(1.0, tk.END)

        name = consumer.get('name', consumer.get('settlement', ''))
        c_type = consumer.get('type', '')

        details = f"Выбран потребитель:\n\n"
        details += f"Тип: {c_type}\n"
        details += f"Название: {name}\n"
        details += f"Район: {consumer.get('mo', '')}\n"
        details += f"НП: {consumer.get('settlement', '')}\n"
        details += f"Привязки: {consumer.get('code', 'Нет')}\n"

        self.detail_text.insert(1.0, details)
        self.detail_text.config(state=tk.DISABLED)

    # === UI UPDATES ===

    def update_statistics(self):
        """Обновление статистики"""
        self.stats_label.config(
            text=f"ПРГ: {len(self.prg_data)} | ГРС: {len(self.grs_data)} | Потребители: {len(self.consumer_data)}"
        )

    def update_changes_display(self):
        """Обновление отображения изменений"""
        if self.changes:
            self.changes_label.config(text=f"⚠️ Несохраненных изменений: {len(self.changes)}")
            self.save_button.config(state=tk.NORMAL)
        else:
            self.changes_label.config(text="")

    def update_button_states(self):
        """Обновление состояния кнопок"""
        has_data = bool(self.prg_data or self.consumer_data)
        has_prg = self.selected_prg is not None
        has_consumer = self.selected_consumer is not None

        # PRG-related buttons (require PRG selection)
        state = tk.NORMAL if has_prg else tk.DISABLED
        self.bind_button.config(state=state)
        self.search_bind_button.config(state=state)
        self.unbind_settlement_button.config(state=state)

        # Manual binding requires both PRG and consumer
        state = tk.NORMAL if (has_prg and has_consumer) else tk.DISABLED
        self.manual_bind_button.config(state=state)

        # Data-related buttons (require loaded data)
        state = tk.NORMAL if has_data else tk.DISABLED
        self.auto_bind_button.config(state=state)
        self.calculate_load_button.config(state=state)
        self.save_button.config(state=state if self.changes else tk.DISABLED)

        # Consumer-related buttons (require consumer selection)
        state = tk.NORMAL if has_consumer else tk.DISABLED
        self.unbind_button.config(state=state)
        self.edit_shares_button.config(state=state)

    # === ACTIONS ===

    def bind_prg_to_settlement(self):
        """Привязать ПРГ ко всему населенному пункту"""
        if not self.selected_prg:
            messagebox.showwarning("Предупреждение", "Выберите ПРГ")
            return

        prg_id = self.selected_prg['prg_id']
        mo = self.selected_prg['mo']
        settlement = self.selected_prg['settlement']
        grs_id = self.selected_prg.get('grs_id', '')

        # Get GRS name from grs_data
        grs_name = self.validation_service.get_grs_name_by_id(self.grs_data, grs_id)

        # Find consumers in the same settlement
        consumers_in_settlement = self.search_service.find_consumers_by_location(
            self.consumer_data, mo, settlement
        )



        # Count by type
        population_count = sum(1 for c in consumers_in_settlement.matches if c.get('type') == 'Население')
        organization_count = sum(1 for c in consumers_in_settlement.matches if c.get('type') == 'Организация')

        # Show dialog to configure binding
        colors = self.style_manager.colors

        dialog = tk.Toplevel(self.root)
        dialog.title("Привязка ко всему населенному пункту")
        dialog.geometry("600x550")
        dialog.resizable(True, True)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg=colors['bg'])

        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() - dialog.winfo_width()) // 2
        y = (dialog.winfo_screenheight() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{x}+{y}")

        main_frame = tk.Frame(dialog, padx=20, pady=20, bg=colors['bg'])
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Title
        tk.Label(main_frame, text="ПРИВЯЗКА КО ВСЕМУ НП",
                 font=('Segoe UI', 14, 'bold'), fg=colors['primary'],
                 bg=colors['bg']).pack(pady=(0, 15))

        # PRG info
        prg_frame = tk.LabelFrame(main_frame, text="ПРГ", font=('Segoe UI', 10, 'bold'),
                                 bg=colors['bg'], fg=colors['text'], borderwidth=1, relief='solid')
        prg_frame.pack(fill=tk.X, pady=(0, 10))
        tk.Label(prg_frame, text=f"ПРГ ID: {prg_id}", font=('Segoe UI', 10),
                bg=colors['bg'], fg=colors['text']).pack(anchor=tk.W, padx=10, pady=5)
        tk.Label(prg_frame, text=f"Район: {mo}", font=('Segoe UI', 10),
                bg=colors['bg'], fg=colors['text']).pack(anchor=tk.W, padx=10)
        tk.Label(prg_frame, text=f"НП: {settlement}", font=('Segoe UI', 10),
                bg=colors['bg'], fg=colors['text']).pack(anchor=tk.W, padx=10, pady=(0, 5))

        # Settlement info
        settlement_frame = tk.LabelFrame(main_frame, text="Потребители в НП", font=('Segoe UI', 10, 'bold'),
                                        bg=colors['bg'], fg=colors['text'], borderwidth=1, relief='solid')
        settlement_frame.pack(fill=tk.X, pady=(0, 10))
        tk.Label(settlement_frame, text=f"Всего потребителей: {consumers_in_settlement.total_count}",
                font=('Segoe UI', 10), bg=colors['bg'], fg=colors['text']).pack(anchor=tk.W, padx=10, pady=5)
        tk.Label(settlement_frame, text=f"  • Население: {population_count}",
                font=('Segoe UI', 10), bg=colors['bg'], fg=colors['text']).pack(anchor=tk.W, padx=10)
        tk.Label(settlement_frame, text=f"  • Организации: {organization_count}",
                font=('Segoe UI', 10), bg=colors['bg'], fg=colors['text']).pack(anchor=tk.W, padx=10)
        tk.Label(settlement_frame, text=f"С расходами: {consumers_in_settlement.with_expenses_count}",
                font=('Segoe UI', 10), bg=colors['bg'], fg=colors['success']).pack(anchor=tk.W, padx=10)
        tk.Label(settlement_frame, text=f"Без расходов: {consumers_in_settlement.without_expenses_count}",
                font=('Segoe UI', 10), bg=colors['bg'], fg=colors['danger']).pack(anchor=tk.W, padx=10, pady=(0, 5))

        # Share input
        share_frame = tk.Frame(main_frame, bg=colors['bg'])
        share_frame.pack(fill=tk.X, pady=10)
        tk.Label(share_frame, text="Доля привязки:", font=('Segoe UI', 11, 'bold'),
                bg=colors['bg'], fg=colors['text']).pack(side=tk.LEFT)
        share_var = tk.StringVar(value="1.0")
        share_entry = tk.Entry(share_frame, textvariable=share_var, font=('Segoe UI', 11),
                              width=10, bg=colors['bg_panel'], fg=colors['text'])
        share_entry.pack(side=tk.LEFT, padx=10)
        tk.Label(share_frame, text="(от 0 до 1)", font=('Segoe UI', 9),
                bg=colors['bg'], fg=colors['text_secondary']).pack(side=tk.LEFT)

        # Consumer type selection
        type_frame = tk.LabelFrame(main_frame, text="Типы потребителей для привязки",
                                   font=('Segoe UI', 10, 'bold'),
                                   bg=colors['bg'], fg=colors['text'], borderwidth=1, relief='solid')
        type_frame.pack(fill=tk.X, pady=(0, 15))

        bind_population = tk.BooleanVar(value=True)
        bind_organizations = tk.BooleanVar(value=True)

        pop_check = tk.Checkbutton(type_frame, text=f"Население ({population_count})",
                                   variable=bind_population, font=('Segoe UI', 10),
                                   bg=colors['bg'], fg=colors['text'],
                                   selectcolor=colors['bg_panel'], activebackground=colors['bg'])
        pop_check.pack(anchor=tk.W, padx=10, pady=5)

        org_check = tk.Checkbutton(type_frame, text=f"Организации ({organization_count})",
                                   variable=bind_organizations, font=('Segoe UI', 10),
                                   bg=colors['bg'], fg=colors['text'],
                                   selectcolor=colors['bg_panel'], activebackground=colors['bg'])
        org_check.pack(anchor=tk.W, padx=10, pady=(0, 5))

        # Warning
        warning_text = "Потребители без расходов будут пропущены автоматически."
        tk.Label(main_frame, text=warning_text, font=('Segoe UI', 9, 'italic'),
                 fg=colors['text_secondary'], bg=colors['bg'],
                 wraplength=550).pack(pady=(0, 15))

        result_holder = {'success': False, 'result': None}

        def do_bind():
            try:
                share = float(share_var.get().replace(',', '.'))
                if share <= 0 or share > 1:
                    messagebox.showerror("Ошибка", "Доля должна быть от 0 до 1", parent=dialog)
                    return

                if not bind_population.get() and not bind_organizations.get():
                    messagebox.showerror("Ошибка", "Выберите хотя бы один тип потребителей", parent=dialog)
                    return

                # Filter consumers by selected types
                consumers_to_bind = []
                for consumer in consumers_in_settlement.matches:
                    c_type = consumer.get('type', '')
                    if c_type == 'Население' and bind_population.get():
                        consumers_to_bind.append(consumer)
                    elif c_type == 'Организация' and bind_organizations.get():
                        consumers_to_bind.append(consumer)

                if not consumers_to_bind:
                    messagebox.showwarning("Предупреждение", "Нет потребителей для привязки", parent=dialog)
                    return

                # Perform binding
                result_holder['success'] = True
                result_holder['share'] = share
                result_holder['consumers'] = consumers_to_bind
                dialog.destroy()

            except ValueError:
                messagebox.showerror("Ошибка", "Введите корректную долю (например: 0.5)", parent=dialog)
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка: {str(e)}", parent=dialog)

        def do_cancel():
            """Cancel and close dialog"""
            dialog.destroy()

        # Bind keyboard shortcuts
        dialog.bind('<Return>', lambda e: do_bind())
        dialog.bind('<Escape>', lambda e: do_cancel())
        share_entry.bind('<Return>', lambda e: do_bind())

        # Buttons
        button_frame = tk.Frame(main_frame, bg=colors['bg'])
        button_frame.pack(fill=tk.X)

        bind_btn = self.style_manager.create_button(
            button_frame, text="Привязать (Enter)",
            command=do_bind, color='success', width=18
        )
        bind_btn.pack(side=tk.RIGHT, padx=(10, 0))

        cancel_btn = self.style_manager.create_button(
            button_frame, text="Отмена (Esc)",
            command=do_cancel, color='text_secondary', width=15
        )
        cancel_btn.config(bg=colors['text_secondary'])
        self.style_manager.add_button_hover(cancel_btn, colors['text_secondary'], colors['text_muted'])
        cancel_btn.pack(side=tk.RIGHT)

        # Focus on share entry
        share_entry.focus_set()

        dialog.wait_window()

        # If user confirmed, perform binding
        if result_holder['success']:
            try:
                success_count = 0
                skipped_count = 0
                already_bound_count = 0
                errors = []

                for consumer in result_holder['consumers']:
                    result = self.binding_service.bind_single_consumer(
                        consumer,
                        self.selected_prg,
                        grs_name,
                        result_holder['share'],
                        force=False
                    )

                    if result.success_count > 0:
                        success_count += 1
                        for change in result.changes:
                            self.changes[change['change_id']] = change
                    elif result.already_bound_count > 0:
                        already_bound_count += 1
                    else:
                        skipped_count += 1
                        if result.errors:
                            errors.extend(result.errors)

                # Update UI
                self.populate_consumer_tree()
                self.update_changes_display()
                self.update_button_states()

                # Show result
                messagebox.showinfo(
                    "Результат привязки",
                    f"Привязка ПРГ {prg_id} завершена:\n\n"
                    f"Успешно привязано: {success_count}\n"
                    f"Пропущено: {skipped_count}\n"
                    f"Уже привязано: {already_bound_count}\n"
                    f"Ошибок: {len(errors)}\n\n"
                    f"Не забудьте сохранить изменения!"
                )

                print(f"[OK] Settlement binding: {success_count} success, {skipped_count} skipped")

            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка привязки:\n\n{str(e)}")
                print(f"[ERROR] {e}")
                import traceback
                traceback.print_exc()

    def unbind_entire_settlement(self):
        """Отвязать весь населенный пункт"""
        if not self.selected_prg:
            messagebox.showwarning("Предупреждение", "Выберите ПРГ")
            return

        mo = self.selected_prg['mo']
        settlement = self.selected_prg['settlement']

        # Find consumers in the same settlement
        consumers_in_settlement = self.search_service.find_consumers_by_location(
            self.consumer_data, mo, settlement
        )

        if consumers_in_settlement.total_count == 0:
            messagebox.showwarning(
                "Предупреждение",
                f"В населенном пункте '{settlement}' района '{mo}' нет потребителей."
            )
            return

        # Count consumers with bindings
        from prg.data.parsers import parse_prg_bindings
        bound_count = sum(1 for c in consumers_in_settlement.matches
                         if parse_prg_bindings(c.get('code', '')))

        if bound_count == 0:
            messagebox.showinfo(
                "Информация",
                f"В населенном пункте '{settlement}' нет привязанных потребителей."
            )
            return

        # Ask for confirmation
        response = messagebox.askyesno(
            "Подтверждение отвязки",
            f"Отвязать ВСЕ привязки потребителей в:\n\n"
            f"Район: {mo}\n"
            f"НП: {settlement}\n\n"
            f"Потребителей с привязками: {bound_count}\n\n"
            f"ВНИМАНИЕ: Это удалит ВСЕ привязки!\n"
            f"Продолжить?"
        )

        if not response:
            return

        try:
            # Create a dummy consumer for the binding service
            dummy_consumer = {
                'mo': mo,
                'settlement': settlement
            }

            result = self.binding_service.unbind_entire_settlement(
                dummy_consumer,
                self.consumer_data
            )

            # Add changes to tracking
            for change in result.changes:
                self.changes[change['change_id']] = change

            # Update UI
            self.populate_consumer_tree()
            self.update_changes_display()
            self.update_button_states()

            # Show result
            messagebox.showinfo(
                "Результат отвязки",
                f"Отвязка потребителей завершена:\n\n"
                f"Успешно отвязано: {result.success_count}\n"
                f"Пропущено (без привязок): {result.skipped_count}\n"
                f"Ошибок: {len(result.errors)}\n\n"
                f"Не забудьте сохранить изменения!"
            )

            print(f"[OK] Settlement unbinding: {result.success_count} success, {result.skipped_count} skipped")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка отвязки:\n\n{str(e)}")
            print(f"[ERROR] {e}")
            import traceback
            traceback.print_exc()

    def unbind_single_consumer(self):
        """Отвязать одного потребителя"""
        if not self.selected_consumer:
            messagebox.showwarning("Предупреждение", "Выберите потребителя")
            return

        from prg.data.parsers import parse_prg_bindings

        consumer_name = self.selected_consumer.get('name', self.selected_consumer.get('settlement', ''))
        bindings = parse_prg_bindings(self.selected_consumer.get('code', ''))

        if not bindings:
            messagebox.showinfo(
                "Информация",
                f"Потребитель '{consumer_name}' не имеет привязок к ПРГ."
            )
            return

        # Build bindings list for display
        bindings_display = "\n".join([f"  - {b['prg_id']} (доля: {b['share']:.2f})" for b in bindings])

        # Ask for confirmation
        response = messagebox.askyesno(
            "Подтверждение отвязки",
            f"Отвязать потребителя:\n\n"
            f"Название: {consumer_name}\n"
            f"Тип: {self.selected_consumer.get('type', '')}\n"
            f"Район: {self.selected_consumer.get('mo', '')}\n"
            f"НП: {self.selected_consumer.get('settlement', '')}\n\n"
            f"Текущие привязки ({len(bindings)}):\n{bindings_display}\n\n"
            f"Удалить ВСЕ привязки?"
        )

        if not response:
            return

        try:
            result = self.binding_service.unbind_single_consumer(self.selected_consumer)

            # Add changes to tracking
            for change in result.changes:
                self.changes[change['change_id']] = change

            # Update UI
            self.populate_consumer_tree()
            self.update_changes_display()
            self.update_button_states()
            self.update_detail_panel_consumer(self.selected_consumer)

            # Show result
            if result.success_count > 0:
                messagebox.showinfo(
                    "Результат отвязки",
                    f"Потребитель '{consumer_name}' успешно отвязан.\n\n"
                    f"Удалено привязок: {len(bindings)}\n\n"
                    f"Не забудьте сохранить изменения!"
                )
                print(f"[OK] Consumer unbinding: {consumer_name}")
            else:
                messagebox.showwarning(
                    "Предупреждение",
                    f"Не удалось отвязать потребителя.\n\n"
                    f"Причина: {result.details[0] if result.details else 'Неизвестно'}"
                )

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка отвязки:\n\n{str(e)}")
            print(f"[ERROR] {e}")
            import traceback
            traceback.print_exc()

    def calculate_prg_load(self):
        """Подсчитать нагрузку ПРГ"""
        if not self.prg_data or not self.consumer_data:
            messagebox.showwarning("Предупреждение", "Загрузите данные перед расчетом")
            return

        # Ask for confirmation
        response = messagebox.askyesno(
            "Расчет нагрузки ПРГ",
            f"Рассчитать нагрузку для всех ПРГ?\n\n"
            f"ПРГ: {len(self.prg_data)}\n"
            f"Потребители: {len(self.consumer_data)}\n\n"
            f"Будут обновлены столбцы нагрузки в таблице ПРГ.\n"
            f"Продолжить?"
        )

        if not response:
            return

        try:
            from datetime import datetime

            print("[INFO] Calculating PRG loads...")

            # Calculate loads
            result = self.calculation_service.calculate_prg_loads(
                self.prg_data,
                self.consumer_data
            )

            # Apply loads to PRG data and create change records
            updated_count = self.calculation_service.apply_loads_to_prg_data(
                self.prg_data,
                result.prg_loads
            )

            # Calculate totals for display
            total_yearly = sum(prg.get('Year_volume', 0) for prg in self.prg_data)
            total_hourly = sum(prg.get('Max_Hour', 0) for prg in self.prg_data)

            # Create change records for each PRG with updated loads
            for prg in self.prg_data:
                prg_id = prg['prg_id']

                # Only create changes for PRGs that were calculated
                if prg_id in result.prg_loads:
                    load = result.prg_loads[prg_id]

                    # Create change records for each load column
                    load_columns = [
                        ('qy_pop_col', 'QY_pop', load['QY_pop']),
                        ('qh_pop_col', 'QH_pop', load['QH_pop']),
                        ('qy_ind_col', 'QY_ind', load['QY_ind']),
                        ('qh_ind_col', 'QH_ind', load['QH_ind']),
                        ('year_volume_col', 'Year_volume', load['QY_pop'] + load['QY_ind']),
                        ('max_hour_col', 'Max_Hour', load['QH_pop'] + load['QH_ind']),
                    ]

                    for col_key, field_name, value in load_columns:
                        if col_key in prg:
                            change_id = f"prg_load_{prg['id']}_{field_name}_{datetime.now().timestamp()}"
                            change = {
                                'change_id': change_id,
                                'type': 'prg_load',
                                'prg_id': prg_id,
                                'sheet_name': prg['sheet_name'],
                                'row': prg['excel_row'],
                                'col': prg[col_key],
                                'new_value': round(value, 4),
                                'old_value': prg.get(field_name, 0),
                                'description': f"Нагрузка ПРГ {prg_id}: {field_name} = {value:.4f}"
                            }
                            self.changes[change_id] = change

            # Update UI
            self.populate_prg_tree()
            self.update_changes_display()
            self.update_button_states()

            print(f"[OK] Loads calculated for {updated_count} PRGs")

            messagebox.showinfo(
                "Расчет завершен",
                f"Расчет нагрузки ПРГ завершен:\n\n"
                f"Обработано потребителей: {result.processed_consumers}\n"
                f"Обработано привязок: {result.processed_bindings}\n"
                f"ПРГ с нагрузкой: {result.updated_prg_count}\n"
                f"Всего ПРГ обновлено: {updated_count}\n\n"
                f"Общая годовая нагрузка: {total_yearly:.2f}\n"
                f"Общая часовая нагрузка: {total_hourly:.4f}\n\n"
                f"Создано изменений: {len(self.changes)}\n\n"
                f"Не забудьте сохранить изменения!"
            )

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка расчета:\n\n{str(e)}")
            print(f"[ERROR] {e}")
            import traceback
            traceback.print_exc()

    def bind_by_search(self):
        """Привязать по умному поиску"""
        if not self.selected_prg:
            messagebox.showwarning("Предупреждение", "Сначала выберите ПРГ в дереве слева")
            return

        if not self.consumer_data:
            messagebox.showwarning("Предупреждение", "Нет данных потребителей")
            return

        try:
            # Get unique districts, settlements, PRG IDs for dropdowns
            districts = self.search_service.get_unique_districts(self.consumer_data)
            settlements = self.search_service.get_settlements_by_district(
                self.consumer_data, self.selected_prg['mo']
            )
            prg_ids = self.search_service.get_prg_ids_by_location(
                self.prg_data,
                self.selected_prg['mo'],
                self.selected_prg['settlement']
            )

            # Import and show smart search dialog
            from prg.ui.dialogs import SmartSearchDialog
            dialog = SmartSearchDialog(
                self.root,
                districts,
                settlements,
                prg_ids,
                self.selected_prg,
                self.style_manager
            )

            if dialog.result:
                # Perform smart search
                search_result = self.search_service.smart_search_organizations(
                    self.consumer_data,
                    dialog.result['mo_district'],
                    dialog.result['settlement'],
                    dialog.result['street'],
                    require_expenses=True
                )

                if search_result.total_count == 0:
                    messagebox.showwarning(
                        "Результат поиска",
                        f"Организации не найдены по параметрам:\n\n"
                        f"Район: {dialog.result['mo_district']}\n"
                        f"НП: {dialog.result['settlement']}\n"
                        f"Улица в названии: {dialog.result['street']}\n\n"
                        f"Проверьте параметры поиска."
                    )
                    return

                # Show found organizations and ask for confirmation
                found_names = "\n".join([f"  - {c['name']}" for c in search_result.matches[:10]])
                if search_result.total_count > 10:
                    found_names += f"\n  ... и еще {search_result.total_count - 10}"

                response = messagebox.askyesno(
                    "Найденные организации",
                    f"Найдено организаций: {search_result.total_count}\n\n"
                    f"{found_names}\n\n"
                    f"Привязать к ПРГ {dialog.result['prg_id']} с долей {dialog.result['share']}?"
                )

                if not response:
                    return

                # Find the PRG to bind to
                prg_to_bind = self.search_service.find_prg_by_id(self.prg_data, dialog.result['prg_id'])
                if not prg_to_bind:
                    messagebox.showerror("Ошибка", f"ПРГ {dialog.result['prg_id']} не найден")
                    return

                # Get GRS name
                grs_id = prg_to_bind.get('grs_id', '')
                grs_name = self.validation_service.get_grs_name_by_id(self.grs_data, grs_id)

                # Bind each found consumer
                success_count = 0
                skipped_count = 0
                errors = []

                for consumer in search_result.matches:
                    result = self.binding_service.bind_single_consumer(
                        consumer,
                        prg_to_bind,
                        grs_name,
                        dialog.result['share'],
                        force=False
                    )

                    if result.success_count > 0:
                        success_count += 1
                        for change in result.changes:
                            self.changes[change['change_id']] = change
                    else:
                        skipped_count += 1
                        if result.errors:
                            errors.extend(result.errors)

                # Update UI
                self.populate_consumer_tree()
                self.update_changes_display()
                self.update_button_states()

                # Show result
                messagebox.showinfo(
                    "Результат привязки по поиску",
                    f"Привязка организаций к ПРГ {dialog.result['prg_id']}:\n\n"
                    f"Успешно привязано: {success_count}\n"
                    f"Пропущено: {skipped_count}\n"
                    f"Ошибок: {len(errors)}\n\n"
                    f"Не забудьте сохранить изменения!"
                )

                print(f"[OK] Search binding: {success_count} success, {skipped_count} skipped")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при поиске:\n\n{str(e)}")
            print(f"[ERROR] {e}")
            import traceback
            traceback.print_exc()

    def bind_manually(self):
        """Привязать вручную (принудительно)"""
        if not self.selected_prg:
            messagebox.showwarning("Предупреждение", "Сначала выберите ПРГ")
            return

        if not self.selected_consumer:
            messagebox.showwarning("Предупреждение", "Выберите потребителя для ручной привязки")
            return

        from prg.data.parsers import parse_prg_bindings, calculate_total_share

        prg_id = self.selected_prg['prg_id']
        consumer_name = self.selected_consumer.get('name', self.selected_consumer.get('settlement', ''))

        # Get current bindings
        bindings = parse_prg_bindings(self.selected_consumer.get('code', ''))
        current_total = calculate_total_share(bindings)
        has_expenses = self.validation_service.has_expenses(self.selected_consumer)

        # Check if already bound to this PRG
        already_bound = any(b['prg_id'] == prg_id for b in bindings)

        # Show manual binding dialog
        colors = self.style_manager.colors

        dialog = tk.Toplevel(self.root)
        dialog.title("Ручная привязка")
        dialog.geometry("600x500")
        dialog.resizable(True, True)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg=colors['bg'])

        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() - dialog.winfo_width()) // 2
        y = (dialog.winfo_screenheight() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{x}+{y}")

        main_frame = tk.Frame(dialog, padx=20, pady=20, bg=colors['bg'])
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Title
        tk.Label(main_frame, text="РУЧНАЯ ПРИВЯЗКА (FORCE)",
                 font=('Segoe UI', 14, 'bold'), fg=colors['danger'],
                 bg=colors['bg']).pack(pady=(0, 15))

        # Warning
        warning_text = "ВНИМАНИЕ: Ручная привязка обходит все проверки!\n" \
                       "Используйте только в крайних случаях."
        tk.Label(main_frame, text=warning_text, font=('Segoe UI', 10),
                 fg=colors['danger'], bg=colors['bg'],
                 wraplength=450).pack(pady=(0, 15))

        # PRG info
        prg_frame = tk.LabelFrame(main_frame, text="ПРГ", font=('Segoe UI', 10, 'bold'),
                                 bg=colors['bg'], fg=colors['text'], borderwidth=1, relief='solid')
        prg_frame.pack(fill=tk.X, pady=(0, 10))
        tk.Label(prg_frame, text=f"ПРГ ID: {prg_id}", font=('Segoe UI', 10),
                bg=colors['bg'], fg=colors['text']).pack(anchor=tk.W, padx=10, pady=5)
        tk.Label(prg_frame, text=f"Район: {self.selected_prg['mo']}", font=('Segoe UI', 10),
                bg=colors['bg'], fg=colors['text']).pack(anchor=tk.W, padx=10)
        tk.Label(prg_frame, text=f"НП: {self.selected_prg['settlement']}", font=('Segoe UI', 10),
                bg=colors['bg'], fg=colors['text']).pack(anchor=tk.W, padx=10, pady=(0, 5))

        # Consumer info
        consumer_frame = tk.LabelFrame(main_frame, text="Потребитель", font=('Segoe UI', 10, 'bold'),
                                       bg=colors['bg'], fg=colors['text'], borderwidth=1, relief='solid')
        consumer_frame.pack(fill=tk.X, pady=(0, 10))
        tk.Label(consumer_frame, text=f"Название: {consumer_name}", font=('Segoe UI', 10),
                bg=colors['bg'], fg=colors['text']).pack(anchor=tk.W, padx=10, pady=5)
        tk.Label(consumer_frame, text=f"Тип: {self.selected_consumer.get('type', '')}", font=('Segoe UI', 10),
                bg=colors['bg'], fg=colors['text']).pack(anchor=tk.W, padx=10)

        expenses_text = "Есть" if has_expenses else "НЕТ (будет привязан принудительно)"
        expenses_color = colors['success'] if has_expenses else colors['danger']
        tk.Label(consumer_frame, text=f"Расходы: {expenses_text}", font=('Segoe UI', 10),
                fg=expenses_color, bg=colors['bg']).pack(anchor=tk.W, padx=10)
        tk.Label(consumer_frame, text=f"Текущая сумма долей: {current_total:.2f}", font=('Segoe UI', 10),
                bg=colors['bg'], fg=colors['text']).pack(anchor=tk.W, padx=10, pady=(0, 5))

        if already_bound:
            tk.Label(consumer_frame, text=f"УЖЕ ПРИВЯЗАН к {prg_id}!", font=('Segoe UI', 10, 'bold'),
                    fg=colors['warning'], bg=colors['bg']).pack(anchor=tk.W, padx=10, pady=(0, 5))

        # Share input
        share_frame = tk.Frame(main_frame, bg=colors['bg'])
        share_frame.pack(fill=tk.X, pady=10)
        tk.Label(share_frame, text="Доля:", font=('Segoe UI', 11, 'bold'),
                bg=colors['bg'], fg=colors['text']).pack(side=tk.LEFT)
        share_var = tk.StringVar(value="1.0")
        share_entry = tk.Entry(share_frame, textvariable=share_var, font=('Segoe UI', 11),
                              width=10, bg=colors['bg_panel'], fg=colors['text'])
        share_entry.pack(side=tk.LEFT, padx=10)

        result_holder = {'success': False}

        def do_bind():
            try:
                share = float(share_var.get().replace(',', '.'))
                if share <= 0 or share > 1:
                    messagebox.showerror("Ошибка", "Доля должна быть от 0 до 1", parent=dialog)
                    return

                # Get GRS name
                grs_id = self.selected_prg.get('grs_id', '')
                grs_name = self.validation_service.get_grs_name_by_id(self.grs_data, grs_id)

                # Perform forced binding
                result = self.binding_service.bind_single_consumer(
                    self.selected_consumer,
                    self.selected_prg,
                    grs_name,
                    share,
                    force=True  # Force binding - bypasses validation
                )

                if result.success_count > 0:
                    # Add changes to tracking
                    for change in result.changes:
                        self.changes[change['change_id']] = change

                    result_holder['success'] = True
                    dialog.destroy()
                else:
                    messagebox.showerror(
                        "Ошибка",
                        f"Не удалось привязать.\n\n"
                        f"Причина: {result.errors[0] if result.errors else 'Неизвестно'}",
                        parent=dialog
                    )

            except ValueError:
                messagebox.showerror("Ошибка", "Введите корректную долю (например: 0.5)", parent=dialog)
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка привязки: {str(e)}", parent=dialog)

        def do_cancel():
            """Cancel and close dialog"""
            dialog.destroy()

        # Bind keyboard shortcuts
        dialog.bind('<Return>', lambda e: do_bind())
        dialog.bind('<Escape>', lambda e: do_cancel())
        share_entry.bind('<Return>', lambda e: do_bind())

        # Buttons
        button_frame = tk.Frame(main_frame, bg=colors['bg'])
        button_frame.pack(fill=tk.X, pady=(15, 0))

        bind_btn = self.style_manager.create_button(
            button_frame, text="Подтвердить (Enter)",
            command=do_bind, color='danger', width=20
        )
        bind_btn.pack(side=tk.RIGHT, padx=(10, 0))

        cancel_btn = self.style_manager.create_button(
            button_frame, text="Отмена (Esc)",
            command=do_cancel, color='text_secondary', width=15
        )
        # Override color for cancel button
        cancel_btn.config(bg=colors['text_secondary'])
        self.style_manager.add_button_hover(cancel_btn, colors['text_secondary'], colors['text_muted'])
        cancel_btn.pack(side=tk.RIGHT)

        # Focus on share entry
        share_entry.focus_set()

        dialog.wait_window()

        if result_holder['success']:
            # Update UI
            self.populate_consumer_tree()
            self.update_changes_display()
            self.update_button_states()
            self.update_detail_panel_consumer(self.selected_consumer)

            messagebox.showinfo(
                "Успех",
                f"Потребитель '{consumer_name}' принудительно привязан к ПРГ {prg_id}.\n\n"
                f"Не забудьте сохранить изменения!"
            )
            print(f"[OK] Manual binding: {consumer_name} -> {prg_id}")

    def auto_bind_all_prg(self):
        """Автоматическая привязка всех ПРГ"""
        if not self.prg_data or not self.consumer_data:
            messagebox.showwarning("Предупреждение", "Загрузите данные перед автопривязкой")
            return

        # Analyze what will be done
        from prg.data.parsers import parse_prg_bindings

        prg_to_process = []
        for prg in self.prg_data:
            mo = prg['mo']
            settlement = prg['settlement']

            # Find unbound consumers in this location
            consumers_in_location = [
                c for c in self.consumer_data
                if c['mo'].strip().lower() == mo.strip().lower()
                and c['settlement'].strip().lower() == settlement.strip().lower()
                and not parse_prg_bindings(c.get('code', ''))  # Not bound
                and self.validation_service.has_expenses(c)  # Has expenses
            ]

            if consumers_in_location:
                prg_to_process.append({
                    'prg': prg,
                    'consumers': consumers_in_location
                })

        if not prg_to_process:
            messagebox.showinfo(
                "Автопривязка",
                "Нет ПРГ с непривязанными потребителями (с расходами).\n\n"
                "Все потребители либо уже привязаны, либо не имеют расходов."
            )
            return

        total_consumers = sum(len(p['consumers']) for p in prg_to_process)

        response = messagebox.askyesno(
            "Автопривязка",
            f"Запустить автоматическую привязку?\n\n"
            f"ПРГ для обработки: {len(prg_to_process)}\n"
            f"Потребителей для привязки: {total_consumers}\n\n"
            f"Каждому потребителю будет назначена доля 1.0\n"
            f"к ПРГ в том же районе и НП.\n\n"
            f"Потребители без расходов будут пропущены.\n"
            f"Уже привязанные потребители будут пропущены.\n\n"
            f"Продолжить?"
        )

        if not response:
            return

        try:
            total_success = 0
            total_skipped = 0
            total_errors = 0

            # Process each PRG
            for item in prg_to_process:
                prg = item['prg']
                grs_id = prg.get('grs_id', '')
                grs_name = self.validation_service.get_grs_name_by_id(self.grs_data, grs_id)

                # Create dummy consumer for settlement binding
                dummy_consumer = {
                    'mo': prg['mo'],
                    'settlement': prg['settlement']
                }

                result = self.binding_service.bind_prg_to_settlement(
                    prg,
                    dummy_consumer,
                    self.consumer_data,
                    grs_name,
                    share=1.0
                )

                total_success += result.success_count
                total_skipped += result.skipped_count + result.already_bound_count
                total_errors += len(result.errors)

                # Add changes to tracking
                for change in result.changes:
                    self.changes[change['change_id']] = change

            # Update UI
            self.populate_consumer_tree()
            self.update_changes_display()
            self.update_button_states()

            # Show result
            messagebox.showinfo(
                "Результат автопривязки",
                f"Автоматическая привязка завершена:\n\n"
                f"ПРГ обработано: {len(prg_to_process)}\n"
                f"Успешно привязано: {total_success}\n"
                f"Пропущено: {total_skipped}\n"
                f"Ошибок: {total_errors}\n\n"
                f"Не забудьте сохранить изменения!"
            )

            print(f"[OK] Auto binding: {total_success} success, {total_skipped} skipped, {total_errors} errors")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка автопривязки:\n\n{str(e)}")
            print(f"[ERROR] {e}")
            import traceback
            traceback.print_exc()

    def edit_consumer_shares(self):
        """Редактировать доли потребителя"""
        if not self.selected_consumer:
            messagebox.showwarning("Предупреждение", "Сначала выберите потребителя")
            return

        # Get current bindings
        code = self.selected_consumer.get('code', '')
        from prg.data.parsers import parse_prg_bindings, calculate_total_share, format_prg_bindings
        from datetime import datetime

        bindings = parse_prg_bindings(code)
        total_share = calculate_total_share(bindings)

        consumer_name = self.selected_consumer.get('name', self.selected_consumer.get('settlement', ''))

        if not bindings:
            messagebox.showwarning(
                "Предупреждение",
                f"Потребитель '{consumer_name}' не имеет привязок к ПРГ.\n\n"
                f"Сначала привяжите потребителя к ПРГ."
            )
            return

        # Create edit dialog
        colors = self.style_manager.colors

        dialog = tk.Toplevel(self.root)
        dialog.title("Редактирование долей")
        dialog.geometry("700x600")
        dialog.resizable(True, True)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg=colors['bg'])

        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() - dialog.winfo_width()) // 2
        y = (dialog.winfo_screenheight() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{x}+{y}")

        main_frame = tk.Frame(dialog, padx=20, pady=20, bg=colors['bg'])
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Title
        tk.Label(main_frame, text="РЕДАКТИРОВАНИЕ ДОЛЕЙ",
                 font=('Segoe UI', 14, 'bold'), fg=colors['primary'],
                 bg=colors['bg']).pack(pady=(0, 10))

        # Consumer info
        tk.Label(main_frame, text=f"Потребитель: {consumer_name}",
                 font=('Segoe UI', 11), bg=colors['bg'], fg=colors['text']).pack(anchor=tk.W)
        tk.Label(main_frame, text=f"Тип: {self.selected_consumer.get('type', '')}",
                 font=('Segoe UI', 10), bg=colors['bg'], fg=colors['text_secondary']).pack(anchor=tk.W)

        # Current total
        total_label = tk.Label(main_frame, text=f"Текущая сумма долей: {total_share:.2f}",
                               font=('Segoe UI', 11, 'bold'), bg=colors['bg'],
                               fg=colors['success'] if abs(total_share - 1.0) < 0.01 else colors['danger'])
        total_label.pack(anchor=tk.W, pady=(5, 15))

        # Bindings frame with scrollbar
        bindings_frame = tk.LabelFrame(main_frame, text=f"Привязки ({len(bindings)})",
                                       font=('Segoe UI', 10, 'bold'),
                                       bg=colors['bg'], fg=colors['text'],
                                       borderwidth=1, relief='solid')
        bindings_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))

        # Canvas for scrolling
        canvas = tk.Canvas(bindings_frame, height=250, bg=colors['bg'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(bindings_frame, orient=tk.VERTICAL, command=canvas.yview,
                                 style='Modern.Vertical.TScrollbar')
        scrollable_frame = tk.Frame(canvas, bg=colors['bg'])

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, pady=5)

        # Create entry for each binding
        share_vars = []
        for i, binding in enumerate(bindings):
            row_frame = tk.Frame(scrollable_frame, bg=colors['bg'])
            row_frame.pack(fill=tk.X, pady=5, padx=5)

            tk.Label(row_frame, text=f"ПРГ: {binding['prg_id']}",
                     font=('Segoe UI', 10), width=20, anchor=tk.W,
                     bg=colors['bg'], fg=colors['text']).pack(side=tk.LEFT)
            tk.Label(row_frame, text=f"ГРС: {binding['grs_name']}",
                     font=('Segoe UI', 9), width=20, anchor=tk.W,
                     bg=colors['bg'], fg=colors['text_secondary']).pack(side=tk.LEFT)

            tk.Label(row_frame, text="Доля:", font=('Segoe UI', 10),
                    bg=colors['bg'], fg=colors['text']).pack(side=tk.LEFT, padx=(10, 5))

            share_var = tk.StringVar(value=f"{binding['share']:.3f}")
            share_vars.append((binding, share_var))

            entry = tk.Entry(row_frame, textvariable=share_var, font=('Segoe UI', 10),
                           width=8, bg=colors['bg_panel'], fg=colors['text'])
            entry.pack(side=tk.LEFT)

        def update_total():
            """Update total share display"""
            try:
                new_total = sum(float(sv.get().replace(',', '.')) for _, sv in share_vars)
                total_label.config(
                    text=f"Сумма долей: {new_total:.2f}",
                    fg=colors['success'] if abs(new_total - 1.0) < 0.01 else colors['danger']
                )
            except ValueError:
                total_label.config(text="Сумма долей: ОШИБКА", fg=colors['danger'])

        # Bind entry changes to update total
        for _, sv in share_vars:
            sv.trace_add('write', lambda *args: update_total())

        result_holder = {'success': False}

        def save_shares():
            try:
                # Validate and collect new shares
                new_bindings = []
                for binding, share_var in share_vars:
                    try:
                        new_share = float(share_var.get().replace(',', '.'))
                        if new_share < 0 or new_share > 1:
                            messagebox.showerror(
                                "Ошибка",
                                f"Доля для ПРГ {binding['prg_id']} должна быть от 0 до 1",
                                parent=dialog
                            )
                            return
                        new_bindings.append({
                            'prg_id': binding['prg_id'],
                            'share': new_share,
                            'grs_name': binding['grs_name']
                        })
                    except ValueError:
                        messagebox.showerror(
                            "Ошибка",
                            f"Некорректное значение доли для ПРГ {binding['prg_id']}",
                            parent=dialog
                        )
                        return

                # Calculate new total
                new_total = sum(b['share'] for b in new_bindings)

                # Warn if total is not 1.0
                if abs(new_total - 1.0) > 0.01:
                    response = messagebox.askyesno(
                        "Предупреждение",
                        f"Сумма долей ({new_total:.2f}) не равна 1.0\n\n"
                        f"Продолжить сохранение?",
                        parent=dialog
                    )
                    if not response:
                        return

                # Format and save new bindings
                old_code = self.selected_consumer.get('code', '')
                new_code = format_prg_bindings(new_bindings)
                self.selected_consumer['code'] = new_code

                # Create change record
                change_id = f"edit_shares_{self.selected_consumer['id']}_{datetime.now().timestamp()}"
                change = {
                    'change_id': change_id,
                    'type': 'edit_shares',
                    'consumer_id': self.selected_consumer['id'],
                    'sheet_name': self.selected_consumer['sheet_name'],
                    'row': self.selected_consumer['excel_row'],
                    'col': self.selected_consumer['code_col'],
                    'new_value': new_code,
                    'old_value': old_code,
                    'description': f"Редактирование долей: {consumer_name}"
                }

                self.changes[change_id] = change
                result_holder['success'] = True
                dialog.destroy()

            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка сохранения: {str(e)}", parent=dialog)

        def do_cancel():
            """Cancel and close dialog"""
            dialog.destroy()

        # Bind keyboard shortcuts
        dialog.bind('<Return>', lambda e: save_shares())
        dialog.bind('<Escape>', lambda e: do_cancel())

        # Buttons
        button_frame = tk.Frame(main_frame, bg=colors['bg'])
        button_frame.pack(fill=tk.X)

        save_btn = self.style_manager.create_button(
            button_frame, text="Сохранить (Enter)",
            command=save_shares, color='success', width=18
        )
        save_btn.pack(side=tk.RIGHT, padx=(10, 0))

        cancel_btn = self.style_manager.create_button(
            button_frame, text="Отмена (Esc)",
            command=do_cancel, color='text_secondary', width=15
        )
        cancel_btn.config(bg=colors['text_secondary'])
        self.style_manager.add_button_hover(cancel_btn, colors['text_secondary'], colors['text_muted'])
        cancel_btn.pack(side=tk.RIGHT)

        dialog.wait_window()

        if result_holder['success']:
            # Update UI
            self.populate_consumer_tree()
            self.update_changes_display()
            self.update_button_states()
            self.update_detail_panel_consumer(self.selected_consumer)

            messagebox.showinfo(
                "Успех",
                f"Доли потребителя '{consumer_name}' успешно изменены.\n\n"
                f"Не забудьте сохранить изменения!"
            )
            print(f"[OK] Shares edited: {consumer_name}")

    def check_all_consumer_shares(self):
        """Проверить доли всех потребителей"""
        if not self.consumer_data:
            messagebox.showwarning("Предупреждение", "Загрузите данные перед проверкой")
            return

        from prg.data.parsers import parse_prg_bindings, calculate_total_share

        issues = []
        for consumer in self.consumer_data:
            code = consumer.get('code', '')
            bindings = parse_prg_bindings(code)

            if bindings:
                total_share = calculate_total_share(bindings)
                if abs(total_share - 1.0) > 0.01:  # Not equal to 1.0 with tolerance
                    name = consumer.get('name', consumer.get('settlement', ''))
                    issues.append(f"{name}: {total_share:.2f}")

        if issues:
            message = f"Найдено потребителей с неправильными долями: {len(issues)}\n\n"
            message += "\n".join(issues[:20])  # Show first 20
            if len(issues) > 20:
                message += f"\n\n... и еще {len(issues) - 20}"
        else:
            message = "✅ Все потребители имеют корректные доли (сумма = 1.0)"

        messagebox.showinfo("Проверка долей", message)

    def show_unbound_analysis(self):
        """Показать анализ непривязанных элементов"""
        if not self.prg_data or not self.consumer_data:
            messagebox.showwarning("Предупреждение", "Загрузите данные перед анализом")
            return

        unbound_prg = self.validation_service.find_unbound_prg(self.prg_data, self.consumer_data)
        unbound_consumers = self.validation_service.find_unbound_consumers(self.consumer_data)

        message = f"""📊 АНАЛИЗ НЕПРИВЯЗАННЫХ ЭЛЕМЕНТОВ

🏭 ПРГ без потребителей: {len(unbound_prg)}
👥 Потребители без ПРГ: {len(unbound_consumers)}

Используйте кнопки привязки для добавления связей."""

        messagebox.showinfo("Анализ непривязанных", message)

    def show_no_expenses_analysis(self):
        """Показать анализ потребителей без расходов"""
        if not self.consumer_data:
            messagebox.showwarning("Предупреждение", "Загрузите данные перед анализом")
            return

        no_expenses = self.validation_service.find_consumers_without_expenses(self.consumer_data)

        if no_expenses:
            # Group by type
            population_count = sum(1 for c in no_expenses if c.get('type') == 'Население')
            organization_count = sum(1 for c in no_expenses if c.get('type') == 'Организация')

            message = f"""📊 АНАЛИЗ ПОТРЕБИТЕЛЕЙ БЕЗ РАСХОДОВ

Всего потребителей без расходов: {len(no_expenses)}

👥 Население: {population_count}
🏢 Организации: {organization_count}

Потребители без расходов не могут быть привязаны с валидацией.
Используйте ручную привязку для принудительной привязки."""
        else:
            message = "✅ Все потребители имеют данные о расходах"

        messagebox.showinfo("Анализ расходов", message)

    def on_close_window(self):
        """Обработка закрытия окна"""
        if self.changes:
            response = messagebox.askyesnocancel(
                "Несохраненные изменения",
                f"У вас есть {len(self.changes)} несохраненных изменений.\n\n"
                "Сохранить перед выходом?"
            )

            if response is None:  # Cancel
                return
            elif response:  # Yes
                self.save_changes_to_excel()

        # Save window geometry and theme preference
        geometry = self.root.geometry()
        self.settings_manager.set_ui_preference('window_geometry', geometry)
        self.settings_manager.set_ui_preference('theme', self.style_manager.get_theme())
        self.settings_manager.save()

        self.root.destroy()

    def toggle_theme(self):
        """Toggle between light and dark themes."""
        try:
            new_theme = self.style_manager.toggle_theme()
            self.settings_manager.set_ui_preference('theme', new_theme)
            self.settings_manager.save()

            # Show message and ask to restart
            messagebox.showinfo(
                "Тема изменена",
                f"Тема изменена на {'темную' if new_theme == 'dark' else 'светлую'}.\n\n"
                f"Перезапустите приложение, чтобы изменения вступили в силу полностью."
            )

            print(f"[OK] Theme changed to {new_theme}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при смене темы: {str(e)}")
            print(f"[ERROR] Theme toggle error: {e}")

    def run(self):
        """Start the application main loop."""
        self.root.mainloop()
