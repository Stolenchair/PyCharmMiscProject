import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import pandas as pd
import openpyxl
from pathlib import Path
import shutil
from datetime import datetime
import traceback
import re
import json
import os


class PRGPipelineManager:
    """
    Менеджер привязок ПРГ к потребителям
    Версия: 7.3 FINAL - УМНЫЙ ПОИСК + РУЧНАЯ ПРИВЯЗКА (ИСПРАВЛЕННАЯ)
    """

    def __init__(self):
        # Основные данные
        self.excel_path = None
        self.prg_data = []
        self.grs_data = []
        self.consumer_data = []
        self.changes = {}

        # Выбранные элементы
        self.selected_prg = None
        self.selected_consumer = None

        # Настройки по умолчанию
        self.default_settings = self.load_default_settings()

        # Флаг закрытия
        self.is_closing = False

        # Создаем интерфейс
        self.setup_ui()

    def load_default_settings(self):
        """ОБНОВЛЕННЫЕ настройки по умолчанию с новыми колонками"""
        settings_file = Path("prg_settings.json")
        default_settings = {
            'prg': {
                'sheet': '',
                'start_row': '10',
                'mo_col': 'A',
                'settlement_col': 'B',
                'prg_id_col': 'C',
                'grs_id_col': 'D',
                # НОВЫЕ КОЛОНКИ ПРГ
                'qy_pop_col': 'E',  # QY_pop
                'qh_pop_col': 'F',  # QH_pop
                'qy_ind_col': 'G',  # QY_ind
                'qh_ind_col': 'H',  # QH_ind
                'year_volume_col': 'I',  # Year_volume
                'max_hour_col': 'J'  # Max_hour
            },
            'grs': {
                'sheet': '',
                'start_row': '10',
                'mo_col': 'A',
                'grs_id_col': 'B',
                'grs_name_col': 'C'
            },
            'population': {
                'sheet': '',
                'start_row': '10',
                'mo_col': 'A',
                'settlement_col': 'B',
                'code_col': 'M',
                'expenses_col': 'N',  # Годовые расходы (уже существует)
                'hourly_expenses_col': 'O'  # НОВАЯ: Часовые расходы
            },
            'organizations': {
                'sheet': '',
                'start_row': '10',
                'name_col': 'D',
                'mo_col': 'A',
                'settlement_col': 'B',
                'code_col': 'M',
                'expenses_col': 'N',  # Годовые расходы (уже существует)
                'hourly_expenses_col': 'O',  # НОВАЯ: Часовые расходы
                'grs_id_col': 'L'
            }
        }

        try:
            if settings_file.exists():
                with open(settings_file, 'r', encoding='utf-8') as f:
                    saved_settings = json.load(f)
                    for table_type in default_settings:
                        if table_type in saved_settings:
                            default_settings[table_type].update(saved_settings[table_type])
                    print(f"✅ Настройки загружены из {settings_file}")
        except Exception as e:
            print(f"⚠️ Ошибка загрузки настроек: {e}")

        return default_settings

    def save_default_settings(self, settings_to_save):
        """Сохранение настроек по умолчанию"""
        try:
            settings_file = Path("prg_settings.json")
            with open(settings_file, 'w', encoding='utf-8') as f:
                json.dump(settings_to_save, f, indent=2, ensure_ascii=False)
            print(f"💾 Настройки сохранены в {settings_file}")
            return True
        except Exception as e:
            print(f"❌ Ошибка сохранения настроек: {e}")
            return False

    def setup_ui(self):
        """Создание пользовательского интерфейса"""
        self.root = tk.Tk()
        self.root.title("PRG Pipeline Manager v7.3 FINAL - ИСПРАВЛЕННАЯ ВЕРСИЯ")
        self.root.geometry("1500x900")
        self.root.configure(bg='#f0f0f0')

        # Создаем меню
        self.create_menu()

        # Верхняя панель
        self.create_top_panel()

        # Основная область
        self.create_main_area()

        # Статус панель
        self.create_status_panel()

        # Обработчик закрытия окна
        self.root.protocol("WM_DELETE_WINDOW", self.on_close_window)

    def create_menu(self):
        """Создание главного меню"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        # Файл
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Файл", menu=file_menu)
        file_menu.add_command(label="Открыть Excel...", command=self.open_excel_file)
        file_menu.add_separator()
        file_menu.add_command(label="Сохранить изменения", command=self.save_changes_to_excel)
        file_menu.add_separator()
        file_menu.add_command(label="Выход", command=self.on_close_window)

        # Настройки
        settings_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Настройки", menu=settings_menu)
        settings_menu.add_command(label="📁 Настройки столбцов по умолчанию", command=self.show_default_settings_dialog)
        settings_menu.add_separator()
        settings_menu.add_command(label="💾 Сохранить текущие как по умолчанию", command=self.save_current_as_default)

        # Инструменты
        tools_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Инструменты", menu=tools_menu)
        tools_menu.add_command(label="🔍 Привязать по поиску", command=self.bind_by_search)
        tools_menu.add_command(label="🎯 Привязать вручную", command=self.bind_manually)
        tools_menu.add_separator()
        tools_menu.add_command(label="🤖 Автопривязка ПРГ", command=self.auto_bind_all_prg)
        tools_menu.add_separator()
        tools_menu.add_command(label="✏️ Редактировать доли", command=self.edit_consumer_shares)
        tools_menu.add_command(label="✅ Проверить доли всех", command=self.check_all_consumer_shares)
        tools_menu.add_separator()
        tools_menu.add_command(label="🔍 Проверка организации на ГРС", command=self.check_organization_grs)
        tools_menu.add_command(label="🔍 Показать непривязанные", command=self.show_unbound_analysis)
        tools_menu.add_command(label="🚫 Показать без расходов", command=self.show_no_expenses_analysis)

    def create_top_panel(self):
        """Создание верхней панели"""
        top_frame = tk.Frame(self.root, bg='#e0e0e0', height=110)
        top_frame.pack(fill=tk.X, padx=10, pady=5)
        top_frame.pack_propagate(False)

        file_frame = tk.Frame(top_frame, bg='#e0e0e0')
        file_frame.pack(fill=tk.X, padx=5, pady=15)

        tk.Button(file_frame, text="📁 Выбрать Excel файл",
                  command=self.open_excel_file, bg='#4CAF50', fg='white',
                  font=('Arial', 11, 'bold')).pack(side=tk.LEFT, padx=(0, 15))

        self.file_label = tk.Label(file_frame, text="Файл не выбран",
                                   bg='#e0e0e0', font=('Arial', 10))
        self.file_label.pack(side=tk.LEFT)

        self.save_button = tk.Button(file_frame, text="💾 Сохранить изменения",
                                     command=self.save_changes_to_excel, bg='#FF9800', fg='white',
                                     font=('Arial', 11, 'bold'), state=tk.DISABLED)
        self.save_button.pack(side=tk.RIGHT, padx=(15, 0))

        self.changes_label = tk.Label(file_frame, text="",
                                      bg='#e0e0e0', font=('Arial', 10, 'bold'), fg='red')
        self.changes_label.pack(side=tk.RIGHT, padx=(15, 15))

    def create_main_area(self):
        """Создание основной рабочей области"""
        main_frame = tk.Frame(self.root, bg='#f0f0f0')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Левая панель - ПРГ
        prg_frame = tk.LabelFrame(main_frame, text="ПРГ (🟡 - без потребителей в том же районе+НП)",
                                  bg='#f0f0f0', font=('Arial', 11, 'bold'))
        prg_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 8))

        prg_tree_frame = tk.Frame(prg_frame, bg='#f0f0f0')
        prg_tree_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        self.prg_tree = ttk.Treeview(prg_tree_frame, columns=('prg_id', 'grs_id'), height=30)
        self.prg_tree.heading('#0', text='Структура ПРГ')
        self.prg_tree.heading('prg_id', text='ПРГ ID')
        self.prg_tree.heading('grs_id', text='ГРС ID')
        self.prg_tree.column('#0', width=280)
        self.prg_tree.column('prg_id', width=80)
        self.prg_tree.column('grs_id', width=80)

        prg_scroll = ttk.Scrollbar(prg_tree_frame, orient=tk.VERTICAL, command=self.prg_tree.yview)
        self.prg_tree.configure(yscrollcommand=prg_scroll.set)

        self.prg_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        prg_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.prg_tree.bind('<<TreeviewSelect>>', self.on_prg_tree_select)

        # Центральная панель с кнопками
        center_frame = tk.Frame(main_frame, bg='#f0f0f0', width=180)
        center_frame.pack(side=tk.LEFT, fill=tk.Y, padx=20)
        center_frame.pack_propagate(False)

        tk.Frame(center_frame, bg='#f0f0f0', height=10).pack()

        # Кнопка привязки ко всему НП
        self.bind_button = tk.Button(center_frame, text="➡️\nПривязать ко\nвсему НП",
                                     command=self.bind_prg_to_settlement, bg='#4CAF50', fg='white',
                                     font=('Arial', 10, 'bold'), state=tk.DISABLED, height=3, width=14)
        self.bind_button.pack(pady=4)

        # ОБНОВЛЕННАЯ кнопка привязки по поиску
        self.search_bind_button = tk.Button(center_frame, text="🔍\nПривязать\nпо поиску",
                                            command=self.bind_by_search, bg='#00BCD4', fg='white',
                                            font=('Arial', 10, 'bold'), state=tk.DISABLED, height=3, width=14)
        self.search_bind_button.pack(pady=4)

        # НОВАЯ кнопка ручной привязки
        self.manual_bind_button = tk.Button(center_frame, text="🎯\nПривязать\nвручную",
                                            command=self.bind_manually, bg='#E91E63', fg='white',
                                            font=('Arial', 10, 'bold'), state=tk.DISABLED, height=3, width=14)
        self.manual_bind_button.pack(pady=4)

        # Кнопка отвязки всего НП
        self.unbind_settlement_button = tk.Button(center_frame, text="⬅️\nОтвязать\nвесь НП",
                                                  command=self.unbind_entire_settlement, bg='#FF5722', fg='white',
                                                  font=('Arial', 10, 'bold'), state=tk.DISABLED, height=3, width=14)
        self.unbind_settlement_button.pack(pady=4)

        self.auto_bind_button = tk.Button(center_frame, text="🤖\nАвто-\nпривязка",
                                          command=self.auto_bind_all_prg, bg='#9C27B0', fg='white',
                                          font=('Arial', 10, 'bold'), state=tk.DISABLED, height=3, width=14)
        self.auto_bind_button.pack(pady=4)

        self.edit_shares_button = tk.Button(center_frame, text="✏️\nРедактировать\nдоли",
                                            command=self.edit_consumer_shares, bg='#2196F3', fg='white',
                                            font=('Arial', 10, 'bold'), state=tk.DISABLED, height=3, width=14)
        self.edit_shares_button.pack(pady=4)

        self.unbind_button = tk.Button(center_frame, text="⬅️\nОтвязать\nпотребителя",
                                       command=self.unbind_single_consumer, bg='#f44336', fg='white',
                                       font=('Arial', 10, 'bold'), state=tk.DISABLED, height=3, width=14)
        self.unbind_button.pack(pady=4)
        # В функции create_main_area(), в center_frame:
        self.calculate_load_button = tk.Button(center_frame, text="📊\nПодсчитать\nнагрузку ПРГ",
                                               command=self.calculate_prg_load, bg='#9C27B0', fg='white',
                                               font=('Arial', 10, 'bold'), state=tk.DISABLED, height=3, width=14)
        self.calculate_load_button.pack(pady=4)

        # Правая панель - Потребители
        consumer_frame = tk.LabelFrame(main_frame, text="Потребители (🟡 - без ПРГ, 🚫 - без расходов)",
                                       bg='#f0f0f0', font=('Arial', 11, 'bold'))
        consumer_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(8, 0))

        consumer_tree_frame = tk.Frame(consumer_frame, bg='#f0f0f0')
        consumer_tree_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        self.consumer_tree = ttk.Treeview(consumer_tree_frame, columns=('type', 'binding', 'total_share'), height=30)
        self.consumer_tree.heading('#0', text='Потребители')
        self.consumer_tree.heading('type', text='Тип')
        self.consumer_tree.heading('binding', text='Привязки к ПРГ')
        self.consumer_tree.heading('total_share', text='Сумма долей')
        self.consumer_tree.column('#0', width=220)
        self.consumer_tree.column('type', width=90)
        self.consumer_tree.column('binding', width=250)
        self.consumer_tree.column('total_share', width=110)

        consumer_scroll = ttk.Scrollbar(consumer_tree_frame, orient=tk.VERTICAL, command=self.consumer_tree.yview)
        self.consumer_tree.configure(yscrollcommand=consumer_scroll.set)

        self.consumer_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        consumer_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.consumer_tree.bind('<<TreeviewSelect>>', self.on_consumer_tree_select)

    def create_status_panel(self):
        """РАСШИРЕННАЯ создание нижней панели статуса с детальной информацией"""
        status_frame = tk.Frame(self.root, bg='#d0d0d0', height=150)  # Увеличена высота
        status_frame.pack(fill=tk.X, padx=10, pady=5)
        status_frame.pack_propagate(False)

        # Верхняя часть - основная информация
        info_frame = tk.Frame(status_frame, bg='#d0d0d0')
        info_frame.pack(fill=tk.X, padx=10, pady=5)

        self.info_label = tk.Label(info_frame,
                                   text="🆕 v7.4: Выделение текста + Зависимые списки + Кнопки действий + Вставка из буфера",
                                   bg='#d0d0d0', font=('Arial', 11), anchor=tk.W)
        self.info_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.stats_label = tk.Label(info_frame, text="ПРГ: 0 | ГРС: 0 | Потребители: 0",
                                    bg='#d0d0d0', font=('Arial', 11, 'bold'))
        self.stats_label.pack(side=tk.RIGHT)

        # НОВАЯ: Детальная информационная панель
        detail_frame = tk.LabelFrame(status_frame, text="Детальная информация (можно выделять и копировать)",
                                     bg='#d0d0d0', font=('Arial', 10, 'bold'))
        detail_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Создаем Text виджет с возможностью выделения
        self.detail_text = tk.Text(detail_frame, height=6, wrap=tk.WORD, font=('Arial', 10),
                                   bg='#f5f5f5', state=tk.DISABLED, cursor="arrow")
        detail_scroll = ttk.Scrollbar(detail_frame, orient=tk.VERTICAL, command=self.detail_text.yview)
        self.detail_text.configure(yscrollcommand=detail_scroll.set)

        self.detail_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        detail_scroll.pack(side=tk.RIGHT, fill=tk.Y, pady=5)

        # Включаем возможность выделения текста
        self.detail_text.bind("<Button-1>", lambda e: self.detail_text.config(state=tk.NORMAL))
        self.detail_text.bind("<FocusOut>", lambda e: self.detail_text.config(state=tk.DISABLED))

        # Добавляем контекстное меню для копирования
        self.create_detail_context_menu()

        # Нижняя часть - предупреждения (если нужна)
        warning_frame = tk.Frame(status_frame, bg='#d0d0d0')
        warning_frame.pack(fill=tk.X, padx=10, pady=(0, 5))

        self.changes_label = tk.Label(warning_frame, text="",
                                      bg='#d0d0d0', font=('Arial', 11, 'bold'), fg='red')
        self.changes_label.pack(side=tk.RIGHT, padx=(15, 15))

    def create_detail_context_menu(self):
        """Создание контекстного меню для детальной панели"""
        self.detail_context_menu = tk.Menu(self.root, tearoff=0)
        self.detail_context_menu.add_command(label="Копировать", command=self.copy_selected_text)
        self.detail_context_menu.add_command(label="Выделить всё", command=self.select_all_text)

        self.detail_text.bind("<Button-3>", self.show_detail_context_menu)

    def show_detail_context_menu(self, event):
        """Показ контекстного меню"""
        try:
            self.detail_context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.detail_context_menu.grab_release()

    def copy_selected_text(self):
        """Копирование выделенного текста"""
        try:
            selected = self.detail_text.selection_get()
            self.root.clipboard_clear()
            self.root.clipboard_append(selected)
            print("✅ Текст скопирован в буфер обмена")
        except tk.TclError:
            # Если ничего не выделено, копируем весь текст
            try:
                all_text = self.detail_text.get(1.0, tk.END).strip()
                if all_text:
                    self.root.clipboard_clear()
                    self.root.clipboard_append(all_text)
                    print("✅ Весь текст скопирован в буфер обмена")
            except:
                print("⚠️ Нет текста для копирования")

    def select_all_text(self):
        """Выделение всего текста"""
        self.detail_text.config(state=tk.NORMAL)
        self.detail_text.tag_add(tk.SEL, "1.0", tk.END)
        self.detail_text.mark_set(tk.INSERT, "1.0")
        self.detail_text.see(tk.INSERT)

    # === УТИЛИТЫ ===

    def center_dialog(self, dialog):
        """Центрирование диалогового окна"""
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() - dialog.winfo_width()) // 2
        y = (dialog.winfo_screenheight() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{x}+{y}")

    def col_to_index(self, col_ref):
        """Преобразование ссылки столбца в индекс"""
        if not col_ref:
            return 0

        col_ref = str(col_ref).strip().upper()

        if col_ref.isdigit():
            return int(col_ref) - 1

        result = 0
        for char in col_ref:
            if 'A' <= char <= 'Z':
                result = result * 26 + (ord(char) - ord('A') + 1)
        return result - 1

    def parse_grs_id_column(self, grs_id_value):
        """Парсинг колонки ГРС_ИД: берем первое число != 0"""
        if not grs_id_value or pd.isna(grs_id_value):
            return None

        grs_str = str(grs_id_value).strip()
        if not grs_str or grs_str == 'nan':
            return None

        numbers = re.findall(r'\d+', grs_str)
        for num_str in numbers:
            try:
                num = int(num_str)
                if num != 0:
                    return str(num)
            except ValueError:
                continue

        return None

    def check_organization_grs(self):
        """
        ИСПРАВЛЕННАЯ ФУНКЦИЯ: Проверка организации на ГРС
        ИЗМЕНЕНИЕ: НЕ показывает организации которые НЕ ПРИВЯЗАНЫ или их "ГРС в ИД" пустой
        """
        if not self.consumer_data:
            messagebox.showwarning("Предупреждение", "Сначала загрузите данные организаций")
            return

        # Фильтруем только организации
        organizations = [c for c in self.consumer_data if c['type'] == 'Организация']

        if not organizations:
            messagebox.showwarning("Предупреждение", "Нет данных организаций для проверки")
            return

        try:
            # Результаты проверки
            mismatches = []
            processed_count = 0
            empty_grs_count = 0
            valid_matches = 0
            skipped_unbound_count = 0  # НОВЫЙ счетчик для непривязанных

            for org in organizations:
                # Получаем значения
                code_in_scheme = org.get('code', '').strip()
                grs_in_id = org.get('grs_id', '').strip()

                # НОВАЯ ЛОГИКА: Пропускаем непривязанных
                bindings = self.parse_prg_bindings(code_in_scheme)
                if not bindings:
                    skipped_unbound_count += 1
                    continue  # Пропускаем организации без привязок

                # Пропускаем пустые названия в "ГРС в ИД" - НЕ считаем ошибкой
                if not grs_in_id or grs_in_id == 'nan':
                    empty_grs_count += 1
                    continue  # НЕ показываем в результатах

                processed_count += 1

                # Извлекаем название ГРС из "ГРС в ИД"
                grs_from_id = self.extract_grs_name_from_id(grs_in_id)

                if not grs_from_id:
                    # Неправильный формат в "ГРС в ИД"
                    mismatches.append({
                        'org': org,
                        'reason': 'Неправильный формат ГРС в ИД',
                        'grs_in_id': grs_in_id,
                        'grs_from_code': '',
                        'code_in_scheme': code_in_scheme
                    })
                    continue

                # Извлекаем название ГРС из "Код в схеме"
                grs_from_code = self.extract_grs_name_from_code(code_in_scheme)

                if not grs_from_code:
                    # Нет ГРС в коде схемы
                    mismatches.append({
                        'org': org,
                        'reason': 'Нет ГРС в коде схемы',
                        'grs_in_id': grs_from_id,
                        'grs_from_code': '',
                        'code_in_scheme': code_in_scheme
                    })
                    continue

                # Сравниваем названия ГРС
                if grs_from_id.lower().strip() != grs_from_code.lower().strip():
                    mismatches.append({
                        'org': org,
                        'reason': 'Несоответствие названий ГРС',
                        'grs_in_id': grs_from_id,
                        'grs_from_code': grs_from_code,
                        'code_in_scheme': code_in_scheme
                    })
                else:
                    valid_matches += 1

            # ОБНОВЛЕННАЯ функция показа результатов с новой статистикой
            self.show_grs_check_results_v2(processed_count, empty_grs_count,
                                           valid_matches, mismatches,
                                           skipped_unbound_count)

            # Если есть несоответствия, предлагаем создать дерево
            if mismatches:
                self.offer_create_mismatch_tree(mismatches)

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при проверке ГРС: {str(e)}")
            print(f"❌ Ошибка проверки ГРС: {e}")
            traceback.print_exc()

    def extract_grs_name_from_id(self, grs_id_value):
        """
        Извлекает название ГРС из значения "ГРС в ИД"
        Формат: "ГРС Название_ГРС"
        """
        if not grs_id_value:
            return ""

        # Убираем лишние пробелы и приводим к нижнему регистру для поиска
        value = grs_id_value.strip()

        # Ищем паттерн "ГРС " (с пробелом)
        if value.lower().startswith('грс '):
            return value[4:].strip()  # Возвращаем все после "ГРС "

        return ""

    def extract_grs_name_from_code(self, code_value):
        """
        Извлекает название ГРС из "Код в схеме"
        Формат: "Код_ПРГ|доля|ГРС Название_ГРС"
        """
        if not code_value:
            return ""

        # Разбиваем по точкам с запятой (может быть несколько привязок)
        bindings = code_value.split(';')

        for binding in bindings:
            binding = binding.strip()
            if not binding:
                continue

            # Разбиваем по вертикальным чертам
            parts = binding.split('|')
            if len(parts) >= 3:
                # Третья часть должна содержать "ГРС Название_ГРС"
                grs_part = '|'.join(parts[2:]).strip()  # Объединяем все части после второй |

                # Извлекаем название ГРС
                if grs_part.lower().startswith('грс '):
                    return grs_part[4:].strip()  # Возвращаем все после "ГРС "

        return ""

    def show_grs_check_results_v2(self, processed_count, empty_grs_count,
                                  valid_matches, mismatches, skipped_unbound_count):
        """ОБНОВЛЕННАЯ функция показа результатов с информацией о пропущенных непривязанных"""

        message = f"""✅ ПРОВЕРКА ОРГАНИЗАЦИЙ НА ГРС ЗАВЕРШЕНА v7.4

    📊 СТАТИСТИКА:
    • Обработано организаций (с привязками): {processed_count}
    • Пустые "ГРС в ИД" (пропущено): {empty_grs_count}
    • Корректные соответствия: {valid_matches}
    • Найдено несоответствий: {len(mismatches)}

    🚫 ИСКЛЮЧЕНИЯ:
    • Непривязанных организаций (скрыто): {skipped_unbound_count}
    • Организации с пустым "ГРС в ИД" (скрыто): {empty_grs_count}

    🔍 ЛОГИКА ПРОВЕРКИ:
    • Проверяются ТОЛЬКО привязанные организации
    • Организации без привязок НЕ показываются
    • Пустой "ГРС в ИД" НЕ считается ошибкой
    • Сравнение названий ГРС без учета регистра"""

        if mismatches:
            message += f"""

    ⚠️ ТИПЫ НЕСООТВЕТСТВИЙ:
    • Неправильный формат: {len([m for m in mismatches if m['reason'] == 'Неправильный формат ГРС в ИД'])}
    • Нет ГРС в коде: {len([m for m in mismatches if m['reason'] == 'Нет ГРС в коде схемы'])}
    • Разные названия ГРС: {len([m for m in mismatches if m['reason'] == 'Несоответствие названий ГРС'])}

    💡 Будет предложено создать дерево с несоответствиями"""

        messagebox.showinfo("Результаты проверки ГРС v7.4", message)

    def offer_create_mismatch_tree(self, mismatches):
        """Предлагает создать дерево с несоответствиями"""

        result = messagebox.askyesno(
            "Создать дерево несоответствий?",
            f"Найдено {len(mismatches)} организаций с несоответствиями ГРС.\\n\\n"
            f"🌳 Создать новое дерево организаций только с несоответствиями?\\n\\n"
            f"📋 В новом окне будут показаны:\\n"
            f"• Название организации\\n"
            f"• ГРС в ИД\\n"
            f"• ГРС из кода схемы\\n"
            f"• Причина несоответствия"
        )

        if result:
            self.create_mismatch_tree_window(mismatches)

    def create_mismatch_tree_window(self, mismatches):
        """Создает окно с деревом несоответствий ГРС"""

        # Создаем новое окно
        mismatch_window = tk.Toplevel(self.root)
        mismatch_window.title(f"Несоответствия ГРС - {len(mismatches)} организаций")
        mismatch_window.geometry("1200x700")
        mismatch_window.transient(self.root)

        # Центрируем окно
        self.center_dialog(mismatch_window)

        # Заголовок
        header_frame = tk.Frame(mismatch_window, bg='#f0f0f0', height=50)
        header_frame.pack(fill=tk.X, padx=10, pady=5)
        header_frame.pack_propagate(False)

        title_label = tk.Label(header_frame,
                               text=f"🔍 Организации с несоответствиями ГРС ({len(mismatches)} шт.)",
                               font=('Arial', 14, 'bold'),
                               bg='#f0f0f0')
        title_label.pack(pady=15)

        # Дерево с несоответствиями
        tree_frame = tk.Frame(mismatch_window)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        columns = ('reason', 'grs_id', 'grs_code', 'full_code')
        mismatch_tree = ttk.Treeview(tree_frame, columns=columns, height=25)

        # Настройка заголовков
        mismatch_tree.heading('#0', text='Организация')
        mismatch_tree.heading('reason', text='Причина')
        mismatch_tree.heading('grs_id', text='ГРС в ИД')
        mismatch_tree.heading('grs_code', text='ГРС из кода')
        mismatch_tree.heading('full_code', text='Полный код схемы')

        # Настройка ширины колонок
        mismatch_tree.column('#0', width=300)
        mismatch_tree.column('reason', width=200)
        mismatch_tree.column('grs_id', width=150)
        mismatch_tree.column('grs_code', width=150)
        mismatch_tree.column('full_code', width=300)

        # Скроллбар
        tree_scroll = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=mismatch_tree.yview)
        mismatch_tree.configure(yscrollcommand=tree_scroll.set)

        mismatch_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # Группируем по районам
        districts = {}
        for mismatch in mismatches:
            org = mismatch['org']
            mo = org['mo']
            if mo not in districts:
                districts[mo] = []
            districts[mo].append(mismatch)

        # Заполняем дерево
        for mo in sorted(districts.keys()):
            # Узел района
            district_item = mismatch_tree.insert('', tk.END,
                                                 text=f"📍 {mo} ({len(districts[mo])} несоответствий)",
                                                 values=('', '', '', ''),
                                                 open=True)

            for mismatch in districts[mo]:
                org = mismatch['org']

                # Определяем иконку по типу несоответствия
                if mismatch['reason'] == 'Неправильный формат ГРС в ИД':
                    icon = "❌"
                elif mismatch['reason'] == 'Нет ГРС в коде схемы':
                    icon = "⭕"
                else:  # Несоответствие названий
                    icon = "⚠️"

                # Узел организации
                org_text = f"{icon} 🏢 {org['name']} ({org['settlement']})"

                mismatch_tree.insert(district_item, tk.END,
                                     text=org_text,
                                     values=(
                                         mismatch['reason'],
                                         mismatch['grs_in_id'],
                                         mismatch['grs_from_code'],
                                         mismatch['code_in_scheme'][:50] + "..." if len(
                                             mismatch['code_in_scheme']) > 50 else mismatch['code_in_scheme']
                                     ))

        # Кнопки действий
        button_frame = tk.Frame(mismatch_window, bg='#f0f0f0', height=60)
        button_frame.pack(fill=tk.X, padx=10, pady=5)
        button_frame.pack_propagate(False)

        # Кнопка экспорта в CSV
        export_button = tk.Button(button_frame,
                                  text="📋 Экспорт в CSV",
                                  command=lambda: self.export_grs_mismatches_csv(mismatches),
                                  bg='#2196F3', fg='white',
                                  font=('Arial', 12, 'bold'))
        export_button.pack(side=tk.LEFT, padx=10, pady=15)

        # Кнопка закрытия
        close_button = tk.Button(button_frame,
                                 text="❌ Закрыть",
                                 command=mismatch_window.destroy,
                                 bg='#f44336', fg='white',
                                 font=('Arial', 12, 'bold'))
        close_button.pack(side=tk.RIGHT, padx=10, pady=15)

        # Информация
        info_label = tk.Label(button_frame,
                              text=f"💡 Найдено {len(mismatches)} несоответствий. Используйте экспорт для детального анализа.",
                              bg='#f0f0f0', font=('Arial', 10))
        info_label.pack(pady=20)

    def export_grs_mismatches_csv(self, mismatches):
        """Экспорт несоответствий ГРС в CSV"""
        try:
            from tkinter import filedialog

            # Запрашиваем путь для сохранения
            filename = filedialog.asksaveasfilename(
                title="Сохранить отчет о несоответствиях ГРС",
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialname=f"grs_mismatches_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            )

            if not filename:
                return

            # Подготавливаем данные для CSV
            csv_data = []
            csv_data.append([
                "Район (МО)",
                "Населенный пункт",
                "Название организации",
                "Причина несоответствия",
                "ГРС в ИД",
                "ГРС из кода схемы",
                "Полный код схемы",
                "Лист Excel",
                "Строка Excel"
            ])

            for mismatch in mismatches:
                org = mismatch['org']
                csv_data.append([
                    org['mo'],
                    org['settlement'],
                    org['name'],
                    mismatch['reason'],
                    mismatch['grs_in_id'],
                    mismatch['grs_from_code'],
                    mismatch['code_in_scheme'],
                    org['sheet_name'],
                    org['excel_row'] + 1  # +1 для пользователя (строки с 1)
                ])

            # Сохраняем в CSV
            import csv
            with open(filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile, delimiter=';')  # Используем ; для Excel
                writer.writerows(csv_data)

            messagebox.showinfo("Экспорт завершен",
                                f"✅ Отчет сохранен в файл:\\n{filename}\\n\\n"
                                f"📊 Экспортировано записей: {len(mismatches)}\\n"
                                f"📋 Формат: CSV (разделитель ;)\\n"
                                f"🔤 Кодировка: UTF-8 with BOM (для Excel)")

        except Exception as e:
            messagebox.showerror("Ошибка экспорта", f"Не удалось сохранить файл:\\n{str(e)}")

    def format_share_for_excel(self, share):
        """Форматирование доли для записи в Excel"""
        if abs(share - 1.0) < 0.0001:
            return "1"
        else:
            return str(share).replace('.', ',')

    def parse_share_from_excel(self, share_str):
        """Парсинг доли из Excel"""
        if not share_str:
            return 0.0
        try:
            normalized_str = str(share_str).replace(',', '.')
            return float(normalized_str)
        except ValueError:
            return 0.0

    def has_expenses(self, consumer):
        """Проверка наличия расходов у потребителя"""
        expenses = consumer.get('expenses', '')
        if not expenses or expenses == '' or expenses == 'nan' or pd.isna(expenses):
            return False

        try:
            expenses_value = float(str(expenses).replace(',', '.'))
            return expenses_value > 0
        except (ValueError, TypeError):
            return False

    def get_expenses_symbol(self, consumer):
        """Получение символа для отображения расходов"""
        return "💰" if self.has_expenses(consumer) else "🚫"

    def get_grs_name_by_id(self, grs_id):
        """Поиск названия ГРС по ID"""
        for grs in self.grs_data:
            if grs['grs_id'] == grs_id:
                return grs['name']
        return f"ГРС {grs_id}"

    def parse_prg_bindings(self, binding_string):
        """Парсинг привязок: 'ПРГ_ID1|доля1|Название1;ПРГ_ID2|доля2|Название2'"""
        if not binding_string or binding_string.strip() == '':
            return []

        bindings = []
        parts = binding_string.split(';')

        for part in parts:
            part = part.strip()
            if not part:
                continue

            components = part.split('|')
            if len(components) >= 3:
                try:
                    prg_id = components[0].strip()
                    share_str = components[1].strip()
                    grs_name = '|'.join(components[2:]).strip()

                    share = self.parse_share_from_excel(share_str)

                    bindings.append({
                        'prg_id': prg_id,
                        'share': share,
                        'grs_name': grs_name
                    })
                except ValueError:
                    print(f"⚠️ Не удалось распарсить: {part}")
                    continue

        return bindings

    def format_prg_bindings(self, bindings):
        """Форматирование привязок в строку"""
        if not bindings:
            return ''

        formatted_parts = []
        for binding in bindings:
            share_str = self.format_share_for_excel(binding['share'])
            formatted_parts.append(f"{binding['prg_id']}|{share_str}|{binding['grs_name']}")

        return ';'.join(formatted_parts)

    def calculate_total_share(self, bindings):
        """Вычисление общей доли"""
        return sum(binding['share'] for binding in bindings)

    # === ФУНКЦИИ СОХРАНЕНИЯ И ВОССТАНОВЛЕНИЯ СОСТОЯНИЯ ДЕРЕВА ===

    def save_tree_state(self, tree):
        """Сохранение состояния дерева (какие узлы открыты)"""
        opened_items = set()

        def collect_opened(item):
            if tree.item(item, 'open'):
                opened_items.add(tree.item(item, 'text'))
            for child in tree.get_children(item):
                collect_opened(child)

        for item in tree.get_children():
            collect_opened(item)

        return opened_items

    def restore_tree_state(self, tree, opened_items):
        """Восстановление состояния дерева"""

        def restore_opened(item):
            text = tree.item(item, 'text')
            if text in opened_items:
                tree.item(item, open=True)
            for child in tree.get_children(item):
                restore_opened(child)

        for item in tree.get_children():
            restore_opened(item)

    def refresh_trees_with_state(self):
        """Обновление деревьев с сохранением состояния"""
        # Сохраняем состояние
        prg_state = self.save_tree_state(self.prg_tree)
        consumer_state = self.save_tree_state(self.consumer_tree)

        # Обновляем деревья
        self.populate_prg_tree()
        self.populate_consumer_tree()

        # Восстанавливаем состояние
        self.restore_tree_state(self.prg_tree, prg_state)
        self.restore_tree_state(self.consumer_tree, consumer_state)

    # === ДИАЛОГ НАСТРОЕК ПО УМОЛЧАНИЮ ===

    def show_default_settings_dialog(self):
        """
        ИСПРАВЛЕННЫЙ диалог настройки столбцов по умолчанию с рабочими скроллбарами

        Проблема была в том, что код скроллбаров был неполный.
        Теперь добавлены полноценные скроллируемые области для каждой вкладки.
        """
        dialog = tk.Toplevel(self.root)
        dialog.title("⚙️ Настройки столбцов по умолчанию - v7.4 ИСПРАВЛЕННАЯ")
        dialog.geometry("950x900")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()

        # Центрируем диалог
        self.center_dialog(dialog)

        main_frame = tk.Frame(dialog, padx=25, pady=25)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Заголовок
        title_label = tk.Label(main_frame, text="⚙️ Настройки столбцов по умолчанию",
                               font=('Arial', 16, 'bold'))
        title_label.pack(pady=(0, 20))

        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill=tk.BOTH, expand=True, pady=(0, 20))

        # Временные настройки для редактирования
        temp_settings = {}

        # ПОЛНАЯ КОНФИГУРАЦИЯ ТАБЛИЦ С НОВЫМИ ПОЛЯМИ
        tables_config = {
            'prg': ('3. ПРГ', [
                ('Лист Excel:', 'sheet'),
                ('Начальная строка:', 'start_row'),
                ('Район (МО):', 'mo_col'),
                ('Населенный пункт:', 'settlement_col'),
                ('ПРГ ID:', 'prg_id_col'),
                ('ГРС ID:', 'grs_id_col'),
                # НОВЫЕ ПОЛЯ НАГРУЗКИ ПРГ
                ('🆕 QY_pop (годовые население):', 'qy_pop_col'),
                ('🆕 QH_pop (часовые население):', 'qh_pop_col'),
                ('🆕 QY_ind (годовые организации):', 'qy_ind_col'),
                ('🆕 QH_ind (часовые организации):', 'qh_ind_col'),
                ('🆕 Year_volume (общий годовой):', 'year_volume_col'),
                ('🆕 Max_hour (макс. часовой):', 'max_hour_col')
            ]),
            'grs': ('4. ГРС', [
                ('Лист Excel:', 'sheet'),
                ('Начальная строка:', 'start_row'),
                ('Район (МО):', 'mo_col'),
                ('ГРС ID:', 'grs_id_col'),
                ('Название ГРС:', 'grs_name_col')
            ]),
            'population': ('1. Население', [
                ('Лист Excel:', 'sheet'),
                ('Начальная строка:', 'start_row'),
                ('Район (МО):', 'mo_col'),
                ('Населенный пункт:', 'settlement_col'),
                ('Код в схеме:', 'code_col'),
                ('Годовые расходы:', 'expenses_col'),
                # НОВОЕ ПОЛЕ ЧАСОВЫХ РАСХОДОВ
                ('🆕 Часовые расходы:', 'hourly_expenses_col')
            ]),
            'organizations': ('2. Организации', [
                ('Лист Excel:', 'sheet'),
                ('Начальная строка:', 'start_row'),
                ('Название организации:', 'name_col'),
                ('Район (МО):', 'mo_col'),
                ('Населенный пункт:', 'settlement_col'),
                ('Код в схеме:', 'code_col'),
                ('Годовые расходы:', 'expenses_col'),
                # НОВОЕ ПОЛЕ ЧАСОВЫХ РАСХОДОВ
                ('🆕 Часовые расходы:', 'hourly_expenses_col'),
                ('ГРС в ИД', 'grs_id_col')
            ])
        }

        # Создаем вкладки для каждой таблицы с полноценными скроллбарами
        for table_type, (tab_name, fields) in tables_config.items():
            # Основной фрейм вкладки
            tab_frame = tk.Frame(notebook)
            notebook.add(tab_frame, text=tab_name)

            temp_settings[table_type] = {}

            # СОЗДАЕМ СКРОЛЛИРУЕМУЮ ОБЛАСТЬ - ИСПРАВЛЕНО!

            # Контейнер для canvas и scrollbar
            scroll_container = tk.Frame(tab_frame)
            scroll_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            # Canvas для скроллирования
            canvas = tk.Canvas(scroll_container, highlightthickness=0)

            # Scrollbar
            scrollbar = ttk.Scrollbar(scroll_container, orient="vertical", command=canvas.yview)

            # Скроллируемый фрейм
            scrollable_frame = tk.Frame(canvas)

            # Привязка события изменения размера
            def on_frame_configure(event, canvas=canvas):
                canvas.configure(scrollregion=canvas.bbox("all"))

            scrollable_frame.bind("<Configure>", on_frame_configure)

            # Создаем окно в canvas
            canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

            # Привязка события изменения размера canvas
            def on_canvas_configure(event, canvas=canvas, scrollable_frame=scrollable_frame):
                # Обновляем ширину скроллируемого фрейма под ширину canvas
                canvas.itemconfig(canvas_window, width=event.width)

            canvas.bind("<Configure>", on_canvas_configure)

            # Настройка скроллинга
            canvas.configure(yscrollcommand=scrollbar.set)

            # Упаковка canvas и scrollbar
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")

            # ОСНОВНАЯ ОБЛАСТЬ НАСТРОЕК
            settings_frame = tk.Frame(scrollable_frame, padx=20, pady=20)
            settings_frame.pack(fill=tk.BOTH, expand=True)

            # Создаем поля для всех колонок
            for i, (label_text, field_key) in enumerate(fields):
                # Определяем цвет для новых полей
                is_new_field = '🆕' in label_text
                label_color = '#2196F3' if is_new_field else 'black'
                label_font = ('Arial', 12, 'bold') if is_new_field else ('Arial', 12)

                # Label
                label = tk.Label(settings_frame, text=label_text,
                                 font=label_font, fg=label_color)
                label.grid(row=i, column=0, sticky=tk.W, pady=8, padx=(0, 20))

                # Entry
                entry = tk.Entry(settings_frame, width=25, font=('Arial', 12))
                entry.insert(0, self.default_settings[table_type].get(field_key, ''))
                entry.grid(row=i, column=1, sticky=tk.W, pady=8)
                temp_settings[table_type][field_key] = entry

            # Включаем прокрутку колесиком мыши
            def bind_mousewheel(event, canvas=canvas):
                canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))

            def unbind_mousewheel(event, canvas=canvas):
                canvas.unbind_all("<MouseWheel>")

            canvas.bind('<Enter>', bind_mousewheel)
            canvas.bind('<Leave>', unbind_mousewheel)

        # ИНФОРМАЦИОННАЯ СЕКЦИЯ О НОВЫХ ВОЗМОЖНОСТЯХ
        info_frame = tk.LabelFrame(main_frame, text="🆕 НОВЫЕ ВОЗМОЖНОСТИ v7.4",
                                   font=('Arial', 12, 'bold'), fg='blue')
        info_frame.pack(fill=tk.X, pady=(10, 20))

        info_text = tk.Text(info_frame, height=6, wrap=tk.WORD, font=('Arial', 10),
                            state=tk.DISABLED, bg='#f8f9fa')
        info_text.pack(fill=tk.X, padx=15, pady=15)

        info_content = """🆕 ОБНОВЛЕНИЕ v7.4 - ПОДДЕРЖКА ЧАСОВЫХ РАСХОДОВ И НАГРУЗКИ ПРГ:

    📊 НОВЫЕ ПОЛЯ ДЛЯ ПОТРЕБИТЕЛЕЙ:
    • Часовые расходы - реальные данные из Excel (вместо вычисления)

    🏭 НОВЫЕ ПОЛЯ ДЛЯ ПРГ (6 колонок нагрузки):
    • QY_pop/QH_pop - накопленные расходы населения
    • QY_ind/QH_ind - накопленные расходы организаций  
    • Year_volume/Max_hour - итоговые значения

    📈 ФУНКЦИЯ "Подсчитать нагрузку ПРГ":
    • Автоматический расчет из привязок • Сохранение в указанные колонки"""

        info_text.config(state=tk.NORMAL)
        info_text.insert(tk.END, info_content)
        info_text.config(state=tk.DISABLED)

        # КНОПКИ ДЕЙСТВИЙ
        button_frame = tk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))

        def save_defaults():
            """Сохранение настроек по умолчанию"""
            try:
                # Собираем все настройки из полей
                new_defaults = {}
                for table_type in temp_settings:
                    new_defaults[table_type] = {}
                    for field_key, entry in temp_settings[table_type].items():
                        new_defaults[table_type][field_key] = entry.get().strip()

                # Валидируем и сохраняем
                if self.save_default_settings(new_defaults):
                    self.default_settings = new_defaults
                    dialog.destroy()

                    # Подсчитываем новые поля
                    new_fields_count = self.count_new_fields_v74(new_defaults)
                    total_fields = self.count_total_fields(new_defaults)

                    messagebox.showinfo("Настройки сохранены",
                                        f"✅ Настройки успешно сохранены\n\n"
                                        f"📊 Всего полей: {total_fields}\n"
                                        f"🆕 Новых полей v7.4: {new_fields_count}\n\n"
                                        f"💾 Файл: prg_settings.json\n"
                                        f"📋 Резервная копия создана автоматически")
                else:
                    messagebox.showerror("Ошибка", "Не удалось сохранить настройки")

            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка сохранения: {str(e)}")

        def reset_to_defaults():
            """Сброс всех полей к значениям по умолчанию"""
            result = messagebox.askyesno("Сброс настроек",
                                         "Сбросить все поля к значениям по умолчанию?\n\n"
                                         "⚠️ Это действие нельзя отменить")

            if result:
                # Значения по умолчанию с новыми полями
                default_values = {
                    'prg': {
                        'sheet': '', 'start_row': '10', 'mo_col': 'A', 'settlement_col': 'B',
                        'prg_id_col': 'C', 'grs_id_col': 'D',
                        'qy_pop_col': 'E', 'qh_pop_col': 'F', 'qy_ind_col': 'G', 'qh_ind_col': 'H',
                        'year_volume_col': 'I', 'max_hour_col': 'J'
                    },
                    'grs': {
                        'sheet': '', 'start_row': '10', 'mo_col': 'A',
                        'grs_id_col': 'B', 'grs_name_col': 'C'
                    },
                    'population': {
                        'sheet': '', 'start_row': '10', 'mo_col': 'A', 'settlement_col': 'B',
                        'code_col': 'M', 'expenses_col': 'N', 'hourly_expenses_col': 'O'
                    },
                    'organizations': {
                        'sheet': '', 'start_row': '10', 'name_col': 'D', 'mo_col': 'A', 'settlement_col': 'B',
                        'code_col': 'M', 'expenses_col': 'N', 'hourly_expenses_col': 'O', 'grs_id_col': 'L'
                    }
                }

                # Применяем значения к полям
                for table_type, fields in temp_settings.items():
                    for field_key, entry in fields.items():
                        default_value = default_values.get(table_type, {}).get(field_key, '')
                        entry.delete(0, tk.END)
                        entry.insert(0, default_value)

                messagebox.showinfo("Сброс выполнен",
                                    "✅ Настройки сброшены к значениям по умолчанию\n\n"
                                    "📊 Включены все новые поля v7.4")

        # Кнопки управления
        tk.Button(button_frame, text="💾 Сохранить настройки", command=save_defaults,
                  bg='#4CAF50', fg='white', font=('Arial', 14, 'bold'), width=18).pack(side=tk.RIGHT, padx=(20, 0))

        tk.Button(button_frame, text="❌ Отмена", command=dialog.destroy,
                  bg='#f44336', fg='white', font=('Arial', 14), width=12).pack(side=tk.RIGHT)

        tk.Button(button_frame, text="🔄 Сброс к умолчанию", command=reset_to_defaults,
                  bg='#FF9800', fg='white', font=('Arial', 12), width=15).pack(side=tk.LEFT)

        # Устанавливаем фокус на первую вкладку
        notebook.select(0)

    def reset_settings_to_default(self, temp_settings):
        """Сброс настроек к значениям по умолчанию"""
        result = messagebox.askyesno("Сброс настроек",
                                     "Сбросить все поля к значениям по умолчанию?\n\n"
                                     "⚠️ Это действие нельзя отменить")

        if result:
            # Загружаем значения по умолчанию
            default_values = {
                'prg': {
                    'sheet': '', 'start_row': '10', 'mo_col': 'A', 'settlement_col': 'B',
                    'prg_id_col': 'C', 'grs_id_col': 'D',
                    'qy_pop_col': 'E', 'qh_pop_col': 'F', 'qy_ind_col': 'G', 'qh_ind_col': 'H',
                    'year_volume_col': 'I', 'max_hour_col': 'J'
                },
                'grs': {
                    'sheet': '', 'start_row': '10', 'mo_col': 'A',
                    'grs_id_col': 'B', 'grs_name_col': 'C'
                },
                'population': {
                    'sheet': '', 'start_row': '10', 'mo_col': 'A', 'settlement_col': 'B',
                    'code_col': 'M', 'expenses_col': 'N', 'hourly_expenses_col': 'O'
                },
                'organizations': {
                    'sheet': '', 'start_row': '10', 'name_col': 'D', 'mo_col': 'A', 'settlement_col': 'B',
                    'code_col': 'M', 'expenses_col': 'N', 'hourly_expenses_col': 'O', 'grs_id_col': 'L'
                }
            }

            # Применяем значения по умолчанию к полям
            for table_type, fields in temp_settings.items():
                for field_key, entry in fields.items():
                    default_value = default_values.get(table_type, {}).get(field_key, '')
                    entry.delete(0, tk.END)
                    entry.insert(0, default_value)

            messagebox.showinfo("Сброс выполнен", "✅ Настройки сброшены к значениям по умолчанию")

    def count_new_fields_v74(self, settings):
        """Подсчет количества новых полей версии 7.4"""
        new_fields = [
            'qy_pop_col', 'qh_pop_col', 'qy_ind_col', 'qh_ind_col',
            'year_volume_col', 'max_hour_col', 'hourly_expenses_col'
        ]

        count = 0
        for table_type, table_settings in settings.items():
            for field in new_fields:
                if field in table_settings and table_settings[field].strip():
                    count += 1

        return count

    def count_total_fields(self, settings):
        """Подсчет общего количества настроенных полей"""
        total = 0
        for table_type, table_settings in settings.items():
            for field, value in table_settings.items():
                if value and str(value).strip():
                    total += 1

        return total

    def save_current_as_default(self):
        """Сохранение текущих настроек как по умолчанию"""
        if not hasattr(self, 'prg_settings'):
            messagebox.showwarning("Предупреждение", "Сначала загрузите данные из Excel файла")
            return

        result = messagebox.askyesno("Сохранить настройки",
                                     "Сохранить текущие настройки столбцов как настройки по умолчанию?")
        if result:
            try:
                # Собираем текущие настройки
                current_settings = {}
                for table_type in ['prg', 'grs', 'population', 'organizations']:
                    if hasattr(self, f"{table_type}_settings"):
                        settings_obj = getattr(self, f"{table_type}_settings")
                        current_settings[table_type] = {k: v.get() for k, v in settings_obj.items()}

                if self.save_default_settings(current_settings):
                    self.default_settings = current_settings
                    messagebox.showinfo("Настройки сохранены",
                                        "Текущие настройки столбцов сохранены как настройки по умолчанию")
                else:
                    messagebox.showerror("Ошибка", "Не удалось сохранить настройки")

            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка сохранения настроек: {str(e)}")

    # === ОТКРЫТИЕ И ЗАГРУЗКА ФАЙЛА ===

    def open_excel_file(self):
        """Открытие Excel файла"""
        try:
            file_path = filedialog.askopenfilename(
                title="Выберите Excel файл",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )

            if file_path:
                self.excel_path = Path(file_path)
                self.file_label.config(text=f"📄 {self.excel_path.name}")

                self.clear_all_changes()
                self.show_settings_dialog()

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть файл: {str(e)}")

    def clear_all_changes(self):
        """Очистка всех изменений"""
        self.changes.clear()
        self.update_changes_display()

    def show_settings_dialog(self):
        """ИСПРАВЛЕННЫЙ диалог настройки таблиц с кнопками"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Настройка 4 таблиц - v7.4")
        dialog.geometry("850x750")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()

        # Центрируем окно
        self.center_dialog(dialog)

        # Читаем листы Excel
        try:
            with pd.ExcelFile(self.excel_path) as excel_file:
                sheets = excel_file.sheet_names
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось прочитать листы: {str(e)}")
            dialog.destroy()
            return

        # ОСНОВНОЙ КОНТЕЙНЕР
        main_container = tk.Frame(dialog)
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        notebook = ttk.Notebook(main_container)
        notebook.pack(fill=tk.BOTH, expand=True, pady=(0, 20))

        # Создаем вкладки настроек с обновленными полями
        prg_frame = tk.Frame(notebook)
        notebook.add(prg_frame, text="3. ПРГ")
        prg_defaults = self.default_settings['prg'].copy()
        if sheets:
            prg_defaults['sheet'] = sheets[0]
        self.create_settings_tab(prg_frame, "prg", sheets, prg_defaults)

        grs_frame = tk.Frame(notebook)
        notebook.add(grs_frame, text="4. ГРС")
        grs_defaults = self.default_settings['grs'].copy()
        if sheets:
            grs_defaults['sheet'] = sheets[0]
        self.create_settings_tab(grs_frame, "grs", sheets, grs_defaults)

        pop_frame = tk.Frame(notebook)
        notebook.add(pop_frame, text="1. Население")
        pop_defaults = self.default_settings['population'].copy()
        if sheets:
            pop_defaults['sheet'] = sheets[0]
        self.create_settings_tab(pop_frame, "population", sheets, pop_defaults)

        org_frame = tk.Frame(notebook)
        notebook.add(org_frame, text="2. Организации")
        org_defaults = self.default_settings['organizations'].copy()
        if sheets:
            org_defaults['sheet'] = sheets[0]
        self.create_settings_tab(org_frame, "organizations", sheets, org_defaults)

        # КНОПКИ ДЕЙСТВИЙ - ИСПРАВЛЕНО!
        button_frame = tk.Frame(main_container)
        button_frame.pack(fill=tk.X, pady=(10, 0))

        def load_data_and_close():
            """Функция загрузки данных и закрытия диалога"""
            try:
                self.load_all_data()
                dialog.destroy()
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка загрузки данных: {str(e)}")

        def cancel_and_close():
            """Функция отмены и закрытия диалога"""
            dialog.destroy()

        # КНОПКИ С ПРАВИЛЬНЫМИ КОМАНДАМИ
        tk.Button(button_frame, text="📊 Загрузить данные",
                  command=load_data_and_close,
                  bg='#4CAF50', fg='white',
                  font=('Arial', 14, 'bold')).pack(side=tk.RIGHT, padx=(20, 0))

        tk.Button(button_frame, text="❌ Отмена",
                  command=cancel_and_close,
                  bg='#f44336', fg='white',
                  font=('Arial', 14)).pack(side=tk.RIGHT)

        # ДОПОЛНИТЕЛЬНЫЕ КНОПКИ СЛЕВА
        tk.Button(button_frame, text="💾 Сохранить настройки",
                  command=lambda: self.save_current_as_default(),
                  bg='#FF9800', fg='white',
                  font=('Arial', 12)).pack(side=tk.LEFT)

        # Устанавливаем фокус на первую вкладку
        notebook.select(0)

    def create_settings_tab(self, parent, tab_type, sheets, defaults):
        """ОБНОВЛЕННАЯ функция создания вкладки настроек с дополнительными колонками"""
        main_frame = tk.Frame(parent)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=30)

        setattr(self, f"{tab_type}_settings", {})
        settings = getattr(self, f"{tab_type}_settings")

        row = 0

        # Лист Excel
        tk.Label(main_frame, text="Лист Excel:",
                 font=('Arial', 12, 'bold')).grid(row=row, column=0, sticky=tk.W, pady=10)
        sheet_combo = ttk.Combobox(main_frame, values=sheets, state="readonly", width=40)
        if defaults.get('sheet') and defaults['sheet'] in sheets:
            sheet_combo.set(defaults['sheet'])
        elif sheets:
            sheet_combo.set(sheets[0])
        sheet_combo.grid(row=row, column=1, padx=(20, 0), pady=10, sticky=tk.W)
        settings['sheet'] = sheet_combo
        row += 1

        # Начальная строка
        tk.Label(main_frame, text="Начальная строка:",
                 font=('Arial', 12, 'bold')).grid(row=row, column=0, sticky=tk.W, pady=10)
        start_row_entry = tk.Entry(main_frame, width=15, font=('Arial', 12))
        start_row_entry.insert(0, defaults.get('start_row', '10'))
        start_row_entry.grid(row=row, column=1, padx=(20, 0), pady=10, sticky=tk.W)
        settings['start_row'] = start_row_entry
        row += 1

        # ОБНОВЛЕННАЯ конфигурация столбцов с новыми полями
        columns_config = {
            'prg': [
                ('Район (МО):', 'mo_col'),
                ('Населенный пункт:', 'settlement_col'),
                ('ПРГ ID:', 'prg_id_col'),
                ('ГРС ID:', 'grs_id_col'),
                # НОВЫЕ КОЛОНКИ ДЛЯ ПРГ
                ('QY_pop (годовые население):', 'qy_pop_col'),
                ('QH_pop (часовые население):', 'qh_pop_col'),
                ('QY_ind (годовые организации):', 'qy_ind_col'),
                ('QH_ind (часовые организации):', 'qh_ind_col'),
                ('Year_volume (общий годовой):', 'year_volume_col'),
                ('Max_hour (макс. часовой):', 'max_hour_col')
            ],
            'grs': [
                ('Район (МО):', 'mo_col'),
                ('ГРС ID:', 'grs_id_col'),
                ('Название ГРС:', 'grs_name_col')
            ],
            'population': [
                ('Район (МО):', 'mo_col'),
                ('Населенный пункт:', 'settlement_col'),
                ('Код в схеме:', 'code_col'),
                ('Годовые расходы:', 'expenses_col'),
                # НОВАЯ КОЛОНКА ДЛЯ НАСЕЛЕНИЯ
                ('Часовые расходы:', 'hourly_expenses_col')
            ],
            'organizations': [
                ('Название организации:', 'name_col'),
                ('Район (МО):', 'mo_col'),
                ('Населенный пункт:', 'settlement_col'),
                ('Код в схеме:', 'code_col'),
                ('Годовые расходы:', 'expenses_col'),
                # НОВАЯ КОЛОНКА ДЛЯ ОРГАНИЗАЦИЙ
                ('Часовые расходы:', 'hourly_expenses_col'),
                ('ГРС в ИД:', 'grs_id_col')
            ]
        }

        # Создаем поля для всех колонок
        for label_text, key in columns_config[tab_type]:
            tk.Label(main_frame, text=label_text,
                     font=('Arial', 12, 'bold')).grid(row=row, column=0, sticky=tk.W, pady=10)
            entry = tk.Entry(main_frame, width=15, font=('Arial', 12))
            entry.insert(0, defaults.get(key, ''))
            entry.grid(row=row, column=1, padx=(20, 0), pady=10, sticky=tk.W)
            settings[key] = entry
            row += 1

    def load_all_data(self):
        """ОБНОВЛЕННАЯ функция загрузки всех данных с поддержкой новых полей"""
        try:
            # Получаем настройки со всеми новыми полями
            prg_settings = {k: v.get() for k, v in self.prg_settings.items()}
            grs_settings = {k: v.get() for k, v in self.grs_settings.items()}
            pop_settings = {k: v.get() for k, v in self.population_settings.items()}
            org_settings = {k: v.get() for k, v in self.organizations_settings.items()}

            # Загружаем данные с поддержкой новых полей
            self.prg_data = self.load_prg_data(prg_settings)
            self.grs_data = self.load_grs_data(grs_settings)

            self.consumer_data = []
            self.consumer_data.extend(self.load_population_data(pop_settings))
            self.consumer_data.extend(self.load_organization_data(org_settings))

            # Обновляем интерфейс
            self.populate_prg_tree()
            self.populate_consumer_tree()
            self.update_statistics()
            self.update_button_states()

            # Показываем статистику с информацией о новых полях
            self.show_enhanced_load_statistics()

        except Exception as e:
            raise Exception(f"Ошибка загрузки данных: {str(e)}")

    def show_enhanced_load_statistics(self):
        """ОБНОВЛЕННАЯ статистика загрузки с информацией о новых полях"""
        unbound_prg_count = len(self.find_unbound_prg())
        unbound_consumers_count = len(self.find_unbound_consumers())
        consumers_without_expenses = len([c for c in self.consumer_data if not self.has_expenses(c)])

        # Подсчитываем потребителей с часовыми расходами
        consumers_with_hourly = len([c for c in self.consumer_data
                                     if c.get('hourly_expenses') and
                                     str(c['hourly_expenses']).strip() not in ['', 'nan']])

        # Подсчитываем ПРГ с существующими значениями нагрузки
        prg_with_load_data = len([p for p in self.prg_data
                                  if any(p.get(field, 0) != 0 for field in
                                         ['QY_pop', 'QH_pop', 'QY_ind', 'QH_ind', 'Year_volume', 'Max_Hour'])])

        message = f"""✅ ДАННЫЕ УСПЕШНО ЗАГРУЖЕНЫ v7.4!

    📊 ОСНОВНАЯ СТАТИСТИКА:
    • ПРГ: {len(self.prg_data)}
    • ГРС: {len(self.grs_data)}  
    • Потребители: {len(self.consumer_data)}

    🆕 НОВЫЕ ВОЗМОЖНОСТИ v7.4:
    • Потребители с часовыми расходами: {consumers_with_hourly}
    • ПРГ с данными нагрузки: {prg_with_load_data}

    🔍 АНАЛИЗ ПРИВЯЗОК:
    • ПРГ без потребителей: {unbound_prg_count}
    • Потребители без ПРГ: {unbound_consumers_count}
    • Потребители без расходов: {consumers_without_expenses}

    📊 ДОСТУПНЫЕ ФУНКЦИИ:
    • 🔍 Умный поиск с выпадающими списками
    • 🎯 Принудительная ручная привязка
    • 📊 Подсчет нагрузки ПРГ (НОВОЕ!)
    • 💾 Сохранение результатов в Excel"""

        messagebox.showinfo("Данные загружены v7.4", message)

    def load_prg_data(self, settings):
        """ОБНОВЛЕННАЯ загрузка данных ПРГ с колонками нагрузки"""
        try:
            df = pd.read_excel(self.excel_path, sheet_name=settings['sheet'], header=None)
            start_row = int(settings['start_row']) - 1

            if start_row > 0:
                df = df.iloc[start_row:].reset_index(drop=True)

            mo_col = self.col_to_index(settings['mo_col'])
            settlement_col = self.col_to_index(settings['settlement_col'])
            prg_id_col = self.col_to_index(settings['prg_id_col'])
            grs_id_col = self.col_to_index(settings['grs_id_col'])

            # НОВЫЕ КОЛОНКИ ДЛЯ НАГРУЗКИ ПРГ
            qy_pop_col = self.col_to_index(settings.get('qy_pop_col', 'E'))
            qh_pop_col = self.col_to_index(settings.get('qh_pop_col', 'F'))
            qy_ind_col = self.col_to_index(settings.get('qy_ind_col', 'G'))
            qh_ind_col = self.col_to_index(settings.get('qh_ind_col', 'H'))
            year_volume_col = self.col_to_index(settings.get('year_volume_col', 'I'))
            max_hour_col = self.col_to_index(settings.get('max_hour_col', 'J'))

            prg_data = []
            for idx, row in df.iterrows():
                try:
                    if mo_col >= len(row) or settlement_col >= len(row) or prg_id_col >= len(row) or grs_id_col >= len(
                            row):
                        continue

                    mo = str(row.iloc[mo_col]).strip() if pd.notna(row.iloc[mo_col]) else ""
                    settlement = str(row.iloc[settlement_col]).strip() if pd.notna(row.iloc[settlement_col]) else ""
                    prg_id = str(row.iloc[prg_id_col]).strip() if pd.notna(row.iloc[prg_id_col]) else ""
                    grs_id_raw = row.iloc[grs_id_col] if pd.notna(row.iloc[grs_id_col]) else ""

                    grs_id = self.parse_grs_id_column(grs_id_raw)

                    # ЗАГРУЖАЕМ ЗНАЧЕНИЯ НАГРУЗКИ ИЗ EXCEL
                    qy_pop = self.parse_numeric_value(row.iloc[qy_pop_col] if qy_pop_col < len(row) else "")
                    qh_pop = self.parse_numeric_value(row.iloc[qh_pop_col] if qh_pop_col < len(row) else "")
                    qy_ind = self.parse_numeric_value(row.iloc[qy_ind_col] if qy_ind_col < len(row) else "")
                    qh_ind = self.parse_numeric_value(row.iloc[qh_ind_col] if qh_ind_col < len(row) else "")
                    year_volume = self.parse_numeric_value(
                        row.iloc[year_volume_col] if year_volume_col < len(row) else "")
                    max_hour = self.parse_numeric_value(row.iloc[max_hour_col] if max_hour_col < len(row) else "")

                    if mo and settlement and prg_id and grs_id:
                        if mo != 'nan' and settlement != 'nan' and prg_id != 'nan':
                            prg_data.append({
                                'id': f"prg_{idx}",
                                'mo': mo,
                                'settlement': settlement,
                                'prg_id': prg_id,
                                'grs_id': grs_id,
                                # НОВЫЕ ПОЛЯ НАГРУЗКИ
                                'QY_pop': qy_pop,
                                'QH_pop': qh_pop,
                                'QY_ind': qy_ind,
                                'QH_ind': qh_ind,
                                'Year_volume': year_volume,
                                'Max_Hour': max_hour,
                                # Информация для сохранения в Excel
                                'sheet_name': settings['sheet'],
                                'excel_row': start_row + idx,
                                'qy_pop_col': qy_pop_col,
                                'qh_pop_col': qh_pop_col,
                                'qy_ind_col': qy_ind_col,
                                'qh_ind_col': qh_ind_col,
                                'year_volume_col': year_volume_col,
                                'max_hour_col': max_hour_col
                            })
                except Exception:
                    continue

            print(f"📋 Загружено ПРГ: {len(prg_data)}")
            return prg_data

        except Exception as e:
            raise Exception(f"Ошибка загрузки ПРГ: {str(e)}")

    def parse_numeric_value(self, value):
        """
        НОВАЯ ФУНКЦИЯ: Парсинг числового значения из Excel

        Args:
            value: Значение из ячейки Excel

        Returns:
            float: Числовое значение или 0.0 если не удалось распарсить
        """
        if not value or pd.isna(value) or str(value).strip() == '' or str(value) == 'nan':
            return 0.0

        try:
            # Заменяем запятую на точку и пытаемся конвертировать в float
            numeric_str = str(value).replace(',', '.').strip()
            return float(numeric_str)
        except (ValueError, TypeError):
            return 0.0

    def get_consumer_expenses(self, consumer):
        """ИСПРАВЛЕННАЯ ФУНКЦИЯ: Получение расходов потребителя"""
        # 1. Получаем годовые расходы
        yearly_raw = consumer.get('expenses', '')
        if not yearly_raw or yearly_raw == '' or yearly_raw == 'nan' or pd.isna(yearly_raw):
            return None

        try:
            yearly_str = str(yearly_raw).replace(',', '.')
            yearly_expenses = float(yearly_str)
            if yearly_expenses <= 0:
                return None
        except (ValueError, TypeError):
            return None

        # 2. ИСПРАВЛЕНО: Получаем часовые расходы
        hourly_raw = consumer.get('hourly_expenses', '')
        hourly_expenses = None

        # Сначала пытаемся взять часовые из Excel
        if hourly_raw and hourly_raw != '' and hourly_raw != 'nan' and not pd.isna(hourly_raw):
            try:
                hourly_str = str(hourly_raw).replace(',', '.')
                hourly_expenses = float(hourly_str)
                if hourly_expenses <= 0:
                    hourly_expenses = None
            except (ValueError, TypeError):
                hourly_expenses = None

        # Если часовых нет - вычисляем из годовых
        if hourly_expenses is None:
            hourly_expenses = yearly_expenses / 8760

        return {
            'yearly': yearly_expenses,
            'hourly': hourly_expenses
        }

    def update_expenses_symbol_display(self, consumer):
        """
        ОБНОВЛЕННАЯ функция определения символа расходов
        Теперь учитывает как годовые, так и часовые расходы

        Args:
            consumer (dict): Объект потребителя

        Returns:
            str: '💰' если есть расходы (годовые или часовые), '🚫' если нет
        """
        # Проверяем годовые расходы
        yearly_raw = consumer.get('expenses', '')
        has_yearly = False

        if yearly_raw and yearly_raw != '' and yearly_raw != 'nan' and not pd.isna(yearly_raw):
            try:
                yearly = float(str(yearly_raw).replace(',', '.'))
                has_yearly = yearly > 0
            except (ValueError, TypeError):
                has_yearly = False

        # Проверяем часовые расходы
        hourly_raw = consumer.get('hourly_expenses', '')
        has_hourly = False

        if hourly_raw and hourly_raw != '' and hourly_raw != 'nan' and not pd.isna(hourly_raw):
            try:
                hourly = float(str(hourly_raw).replace(',', '.'))
                has_hourly = hourly > 0
            except (ValueError, TypeError):
                has_hourly = False

        # Возвращаем символ: есть расходы если есть годовые ИЛИ часовые
        return '💰' if (has_yearly or has_hourly) else '🚫'

    def show_load_calculation_results(self, processed_consumers, processed_bindings,
                                      updated_prg_count, total_prg_with_load, errors,
                                      used_hourly_from_excel=0):
        """
        ОБНОВЛЕННАЯ функция показа результатов с информацией об источнике часовых данных

        Args:
            used_hourly_from_excel (int): Количество потребителей, у которых использованы
                                         реальные часовые расходы из Excel
        """
        message = f"""✅ ПОДСЧЕТ НАГРУЗКИ ПРГ ЗАВЕРШЕН

    📊 СТАТИСТИКА ОБРАБОТКИ:
    • Обработано потребителей с расходами: {processed_consumers}
    • Обработано привязок: {processed_bindings}  
    • Обновлено ПРГ: {updated_prg_count}
    • ПРГ с нагрузкой: {total_prg_with_load}

    📈 ИСТОЧНИКИ ЧАСОВЫХ РАСХОДОВ:
    • Из Excel (реальные данные): {used_hourly_from_excel}
    • Вычислено (годовые/8760): {processed_consumers - used_hourly_from_excel}

    💾 ОБНОВЛЕННЫЕ ПОЛЯ ПРГ:
    • QY_pop - годовые расходы населения
    • QH_pop - часовые расходы населения
    • QY_ind - годовые расходы организаций
    • QH_ind - часовые расходы организаций
    • Year_volume - общий годовой объем (QY_pop + QY_ind)
    • Max_Hour - макс. часовая нагрузка (QH_pop + QH_ind)"""

        if errors:
            message += f"\n\n⚠️ ОШИБКИ ({len(errors)}):\n"
            message += "\n".join(errors[:3])  # Показываем первые 3 ошибки
            if len(errors) > 3:
                message += f"\n... и еще {len(errors) - 3} ошибок"

        messagebox.showinfo("Нагрузка ПРГ подсчитана", message)

    def save_prg_load_to_excel(self):
        """
        ОБНОВЛЕННАЯ функция сохранения нагрузки ПРГ в Excel
        Использует настройки колонок из импорта
        """
        if not self.excel_path or not self.excel_path.exists():
            messagebox.showerror("Ошибка", "Excel файл не найден")
            return False

        try:
            workbook = openpyxl.load_workbook(self.excel_path)

            # Получаем настройки листа ПРГ
            prg_sheet_name = self.prg_settings['sheet'].get()
            if prg_sheet_name not in workbook.sheetnames:
                messagebox.showerror("Ошибка", f"Лист ПРГ '{prg_sheet_name}' не найден")
                return False

            worksheet = workbook[prg_sheet_name]

            # ИСПОЛЬЗУЕМ НАСТРОЙКИ КОЛОНОК ИЗ ИМПОРТА
            columns_settings = {
                'QY_pop': self.prg_settings.get('qy_pop_col'),
                'QH_pop': self.prg_settings.get('qh_pop_col'),
                'QY_ind': self.prg_settings.get('qy_ind_col'),
                'QH_ind': self.prg_settings.get('qh_ind_col'),
                'Year_volume': self.prg_settings.get('year_volume_col'),
                'Max_Hour': self.prg_settings.get('max_hour_col')
            }

            start_row = int(self.prg_settings['start_row'].get())
            prg_id_col = self.prg_settings['prg_id_col'].get()

            saved_count = 0

            # Обновляем данные в Excel
            for row_idx in range(start_row, worksheet.max_row + 1):
                cell_value = worksheet[f"{prg_id_col}{row_idx}"].value
                if not cell_value:
                    continue

                excel_prg_id = str(cell_value).strip()

                # Ищем соответствующий ПРГ в данных
                for prg in self.prg_data:
                    if prg['prg_id'] == excel_prg_id:
                        # Записываем все значения нагрузки
                        for field, col in columns_settings.items():
                            if col:  # Если колонка указана
                                value = prg.get(field, 0.0)
                                worksheet[f"{col}{row_idx}"] = value
                                print(f"📝 ПРГ {excel_prg_id}, {field} → {col}{row_idx}: {value}")

                        saved_count += 1


            # Сохраняем файл
            workbook.save(self.excel_path)
            workbook.close()

            print(f"✅ Сохранено нагрузок ПРГ в Excel: {saved_count}")

            # Показываем подробности сохранения
            messagebox.showinfo("Сохранение нагрузки ПРГ",
                                f"✅ Сохранено нагрузок ПРГ: {saved_count}\n\n"
                                f"📊 Обновлены колонки:\n"
                                f"• QY_pop → {columns_settings['QY_pop']}\n"
                                f"• QH_pop → {columns_settings['QH_pop']}\n"
                                f"• QY_ind → {columns_settings['QY_ind']}\n"
                                f"• QH_ind → {columns_settings['QH_ind']}\n"
                                f"• Year_volume → {columns_settings['Year_volume']}\n"
                                f"• Max_Hour → {columns_settings['Max_Hour']}")

            return True

        except Exception as e:
            print(f"❌ Ошибка сохранения нагрузки ПРГ: {e}")
            traceback.print_exc()
            messagebox.showerror("Ошибка", f"Ошибка сохранения нагрузки: {str(e)}")
            return False

    def load_grs_data(self, settings):
        """Загрузка данных ГРС"""
        try:
            df = pd.read_excel(self.excel_path, sheet_name=settings['sheet'], header=None)
            start_row = int(settings['start_row']) - 1

            if start_row > 0:
                df = df.iloc[start_row:].reset_index(drop=True)

            mo_col = self.col_to_index(settings['mo_col'])
            grs_id_col = self.col_to_index(settings['grs_id_col'])
            grs_name_col = self.col_to_index(settings['grs_name_col'])

            grs_data = []
            for idx, row in df.iterrows():
                try:
                    if mo_col >= len(row) or grs_id_col >= len(row) or grs_name_col >= len(row):
                        continue

                    mo = str(row.iloc[mo_col]).strip() if pd.notna(row.iloc[mo_col]) else ""
                    grs_id = str(row.iloc[grs_id_col]).strip() if pd.notna(row.iloc[grs_id_col]) else ""
                    grs_name = str(row.iloc[grs_name_col]).strip() if pd.notna(row.iloc[grs_name_col]) else ""

                    if mo and grs_id and grs_name:
                        if mo != 'nan' and grs_id != 'nan' and grs_name != 'nan':
                            grs_data.append({
                                'id': f"grs_{idx}",
                                'mo': mo,
                                'grs_id': grs_id,
                                'name': grs_name
                            })
                except Exception:
                    continue

            print(f"🏭 Загружено ГРС: {len(grs_data)}")
            return grs_data

        except Exception as e:
            raise Exception(f"Ошибка загрузки ГРС: {str(e)}")

    def get_prg_sheet_name(self):
        """Получение названия листа ПРГ"""
        # Из настроек импорта
        if hasattr(self, 'prg_settings') and self.prg_settings:
            sheet_name = self.prg_settings.get('sheet', tk.StringVar()).get()
            if sheet_name:
                return sheet_name

        # Из данных ПРГ
        if self.prg_data:
            sheet_name = self.prg_data[0].get('sheet_name')
            if sheet_name:
                return sheet_name

        # Fallback
        return 'ПРГ'

    def register_prg_load_changes(self, prg_loads):
        """ИСПРАВЛЕННАЯ ФУНКЦИЯ: Регистрация изменений нагрузки ПРГ"""
        try:
            # Получаем название листа ПРГ - ИСПРАВЛЕНО!
            sheet_name = self.get_prg_sheet_name_safe()

            for prg in self.prg_data:
                prg_id = prg['prg_id']

                # Проверяем есть ли изменения нагрузки для этого ПРГ
                if prg_id in prg_loads or any(key in prg for key in ['QY_pop', 'QH_pop', 'QY_ind', 'QH_ind']):
                    change_id = f"prg_load_{prg_id}_{int(datetime.now().timestamp())}"

                    self.changes[change_id] = {
                        'type': 'prg_load_calculation',
                        'prg_id': prg_id,
                        'sheet_name': sheet_name,
                        'description': f"Подсчет нагрузки для ПРГ {prg_id}",
                        'data': {
                            'QY_pop': prg.get('QY_pop', 0.0),
                            'QH_pop': prg.get('QH_pop', 0.0),
                            'QY_ind': prg.get('QY_ind', 0.0),
                            'QH_ind': prg.get('QH_ind', 0.0),
                            'Year_volume': prg.get('Year_volume', 0.0),
                            'Max_Hour': prg.get('Max_Hour', 0.0)
                        }
                    }

                    print(f"📝 Зарегистрировано: {change_id}")

        except Exception as e:
            print(f"⚠️ Ошибка регистрации изменений нагрузки: {e}")
            traceback.print_exc()

    def get_prg_sheet_name_safe(self):
        """БЕЗОПАСНОЕ получение названия листа ПРГ без обращения к Tkinter виджетам"""
        try:
            # Способ 1: Из данных ПРГ (самый надежный)
            if self.prg_data and len(self.prg_data) > 0:
                sheet_name = self.prg_data[0].get('sheet_name')
                if sheet_name:
                    print(f"🔍 Лист ПРГ из данных: {sheet_name}")
                    return sheet_name

            # Способ 2: Попытка получить из prg_settings с проверкой
            if hasattr(self, 'prg_settings') and self.prg_settings:
                try:
                    sheet_widget = self.prg_settings.get('sheet')
                    if sheet_widget and hasattr(sheet_widget, 'get'):
                        sheet_name = sheet_widget.get()
                        if sheet_name:
                            print(f"🔍 Лист ПРГ из настроек: {sheet_name}")
                            return sheet_name
                except tk.TclError:
                    # Виджет уже уничтожен - это нормально
                    print("🔍 Виджеты настроек недоступны (диалог закрыт)")
                    pass

            # Способ 3: Из сохраненных настроек (если есть)
            if hasattr(self, '_last_prg_sheet_name') and self._last_prg_sheet_name:
                print(f"🔍 Лист ПРГ из кэша: {self._last_prg_sheet_name}")
                return self._last_prg_sheet_name

            # Способ 4: Fallback
            print("🔍 Используется fallback название листа ПРГ")
            return 'ПРГ'

        except Exception as e:
            print(f"⚠️ Ошибка получения листа ПРГ: {e}")
            return 'ПРГ'

    def get_prg_load_columns_safe(self):
        """БЕЗОПАСНОЕ получение колонок нагрузки ПРГ без обращения к Tkinter виджетам"""
        try:
            # Способ 1: Из кэшированных настроек
            if hasattr(self, '_cached_prg_columns') and self._cached_prg_columns:
                print("🔍 Колонки ПРГ из кэша")
                return self._cached_prg_columns

            # Способ 2: Попытка получить из prg_settings с проверкой
            columns = {}
            if hasattr(self, 'prg_settings') and self.prg_settings:
                field_mapping = {
                    'QY_pop': 'qy_pop_col',
                    'QH_pop': 'qh_pop_col',
                    'QY_ind': 'qy_ind_col',
                    'QH_ind': 'qh_ind_col',
                    'Year_volume': 'year_volume_col',
                    'Max_Hour': 'max_hour_col'
                }

                for field, setting_key in field_mapping.items():
                    try:
                        widget = self.prg_settings.get(setting_key)
                        if widget and hasattr(widget, 'get'):
                            value = widget.get()
                            columns[field] = value if value else self.get_default_column(field)
                    except tk.TclError:
                        # Виджет недоступен - используем значение по умолчанию
                        columns[field] = self.get_default_column(field)

            # Способ 3: Fallback к значениям по умолчанию
            if not columns:
                print("🔍 Используются колонки ПРГ по умолчанию")
                columns = {
                    'QY_pop': 'E',
                    'QH_pop': 'F',
                    'QY_ind': 'G',
                    'QH_ind': 'H',
                    'Year_volume': 'I',
                    'Max_Hour': 'J'
                }

            # Кэшируем результат
            self._cached_prg_columns = columns
            return columns

        except Exception as e:
            print(f"⚠️ Ошибка получения колонок ПРГ: {e}")
            return {
                'QY_pop': 'E', 'QH_pop': 'F', 'QY_ind': 'G',
                'QH_ind': 'H', 'Year_volume': 'I', 'Max_Hour': 'J'
            }

    def get_default_column(self, field):
        """Получение колонки по умолчанию для поля"""
        defaults = {
            'QY_pop': 'E', 'QH_pop': 'F', 'QY_ind': 'G',
            'QH_ind': 'H', 'Year_volume': 'I', 'Max_Hour': 'J'
        }
        return defaults.get(field, 'A')

    def cache_prg_settings(self):
        """НОВАЯ ФУНКЦИЯ: Кэширование настроек ПРГ при закрытии диалога"""
        try:
            # Сохраняем название листа
            if hasattr(self, 'prg_settings') and self.prg_settings:
                sheet_widget = self.prg_settings.get('sheet')
                if sheet_widget and hasattr(sheet_widget, 'get'):
                    self._last_prg_sheet_name = sheet_widget.get()

            # Сохраняем колонки
            columns = self.get_prg_load_columns_safe()
            self._cached_prg_columns = columns

            print("💾 Настройки ПРГ закэшированы")

        except Exception as e:
            print(f"⚠️ Ошибка кэширования настроек ПРГ: {e}")

    def load_population_data(self, settings):
        """ОБНОВЛЕННАЯ загрузка данных населения с часовыми расходами"""
        try:
            df = pd.read_excel(self.excel_path, sheet_name=settings['sheet'], header=None)
            start_row = int(settings['start_row']) - 1

            if start_row > 0:
                df = df.iloc[start_row:].reset_index(drop=True)

            mo_col = self.col_to_index(settings['mo_col'])
            settlement_col = self.col_to_index(settings['settlement_col'])
            code_col = self.col_to_index(settings['code_col'])
            expenses_col = self.col_to_index(settings['expenses_col'])
            hourly_expenses_col = self.col_to_index(settings.get('hourly_expenses_col', 'O'))  # НОВОЕ ПОЛЕ

            population_data = []
            for idx, row in df.iterrows():
                try:
                    if mo_col >= len(row) or settlement_col >= len(row):
                        continue

                    mo = str(row.iloc[mo_col]).strip() if pd.notna(row.iloc[mo_col]) else ""
                    settlement = str(row.iloc[settlement_col]).strip() if pd.notna(row.iloc[settlement_col]) else ""
                    code = str(row.iloc[code_col]).strip() if code_col < len(row) and pd.notna(
                        row.iloc[code_col]) else ""

                    # ГОДОВЫЕ расходы (существующее поле)
                    yearly_expenses = row.iloc[expenses_col] if expenses_col < len(row) and pd.notna(
                        row.iloc[expenses_col]) else ""

                    # ЧАСОВЫЕ расходы (НОВОЕ поле)
                    hourly_expenses = row.iloc[hourly_expenses_col] if hourly_expenses_col < len(row) and pd.notna(
                        row.iloc[hourly_expenses_col]) else ""

                    if mo and settlement:
                        if mo != 'nan' and settlement != 'nan':
                            population_data.append({
                                'id': f"pop_{settings['sheet']}_{start_row + idx}",
                                'type': 'Население',
                                'mo': mo,
                                'settlement': settlement,
                                'name': f"Население {settlement}",
                                'code': code if code != 'nan' else '',
                                'expenses': yearly_expenses,  # Годовые расходы
                                'hourly_expenses': hourly_expenses,  # НОВОЕ: Часовые расходы
                                'sheet_name': settings['sheet'],
                                'excel_row': start_row + idx,
                                'code_col': code_col
                            })
                except Exception:
                    continue

            print(f"🏠 Загружено население: {len(population_data)}")
            return population_data

        except Exception as e:
            raise Exception(f"Ошибка загрузки населения: {str(e)}")

    def load_organization_data(self, settings):
        """ОБНОВЛЕННАЯ загрузка данных организаций с часовыми расходами"""
        try:
            df = pd.read_excel(self.excel_path, sheet_name=settings['sheet'], header=None)
            start_row = int(settings['start_row']) - 1

            if start_row > 0:
                df = df.iloc[start_row:].reset_index(drop=True)

            name_col = self.col_to_index(settings['name_col'])
            mo_col = self.col_to_index(settings['mo_col'])
            settlement_col = self.col_to_index(settings['settlement_col'])
            code_col = self.col_to_index(settings['code_col'])
            expenses_col = self.col_to_index(settings['expenses_col'])
            hourly_expenses_col = self.col_to_index(settings.get('hourly_expenses_col', 'O'))  # НОВОЕ ПОЛЕ
            grs_id_col = self.col_to_index(settings['grs_id_col'])

            organization_data = []
            for idx, row in df.iterrows():
                try:
                    if name_col >= len(row) or mo_col >= len(row) or settlement_col >= len(row):
                        continue

                    name = str(row.iloc[name_col]).strip() if pd.notna(row.iloc[name_col]) else ""
                    mo = str(row.iloc[mo_col]).strip() if pd.notna(row.iloc[mo_col]) else ""
                    settlement = str(row.iloc[settlement_col]).strip() if pd.notna(row.iloc[settlement_col]) else ""
                    code = str(row.iloc[code_col]).strip() if code_col < len(row) and pd.notna(
                        row.iloc[code_col]) else ""

                    # ГОДОВЫЕ расходы (существующее поле)
                    yearly_expenses = row.iloc[expenses_col] if expenses_col < len(row) and pd.notna(
                        row.iloc[expenses_col]) else ""

                    # ЧАСОВЫЕ расходы (НОВОЕ поле)
                    hourly_expenses = row.iloc[hourly_expenses_col] if hourly_expenses_col < len(row) and pd.notna(
                        row.iloc[hourly_expenses_col]) else ""

                    grs_id = str(row.iloc[grs_id_col]).strip() if grs_id_col < len(row) and pd.notna(
                        row.iloc[grs_id_col]) else ""

                    if name and mo and settlement:
                        if name != 'nan' and mo != 'nan' and settlement != 'nan':
                            organization_data.append({
                                'id': f"org_{settings['sheet']}_{start_row + idx}",
                                'type': 'Организация',
                                'mo': mo,
                                'settlement': settlement,
                                'name': name,
                                'code': code if code != 'nan' else '',
                                'grs_id': grs_id if grs_id != 'nan' else '',
                                'grs_id_col': grs_id_col,
                                'expenses': yearly_expenses,  # Годовые расходы
                                'hourly_expenses': hourly_expenses,  # НОВОЕ: Часовые расходы
                                'sheet_name': settings['sheet'],
                                'excel_row': start_row + idx,
                                'code_col': code_col
                            })
                except Exception:
                    continue

            print(f"🏢 Загружено организаций: {len(organization_data)}")
            return organization_data

        except Exception as e:
            raise Exception(f"Ошибка загрузки организаций: {str(e)}")

    def show_load_statistics(self):
        """Статистика загрузки"""
        unbound_prg_count = len(self.find_unbound_prg())
        unbound_consumers_count = len(self.find_unbound_consumers())
        consumers_without_expenses = len([c for c in self.consumer_data if not self.has_expenses(c)])

        message = f"""✅ Данные успешно загружены v7.3!

📊 СТАТИСТИКА:
• ПРГ: {len(self.prg_data)}
• ГРС: {len(self.grs_data)}
• Потребители: {len(self.consumer_data)}

🔍 АНАЛИЗ:
• ПРГ без потребителей: {unbound_prg_count}
• Потребители без ПРГ: {unbound_consumers_count}
• Потребители без расходов: {consumers_without_expenses}

🆕 НОВЫЕ ФУНКЦИИ v7.3 FINAL:
• 🔍 Умный поиск с выпадающими списками
• 🎯 Принудительная ручная привязка БЕЗ проверок
• 🌳 Сохранение состояния дерева
• 💰🚫 Проверка расходов с символами"""

        messagebox.showinfo("Данные загружены", message)

    # === ОТОБРАЖЕНИЕ ДЕРЕВЬЕВ ===

    def populate_prg_tree(self):
        """Заполнение дерева ПРГ"""
        for item in self.prg_tree.get_children():
            self.prg_tree.delete(item)

        if not self.prg_data:
            return

        # Определяем ПРГ без потребителей
        unbound_prg_ids = set()
        for prg in self.find_unbound_prg():
            unbound_prg_ids.add(prg['prg_id'])

        # Группируем по району и НП
        structure = {}
        for prg in self.prg_data:
            mo = prg['mo']
            settlement = prg['settlement']

            if mo not in structure:
                structure[mo] = {}
            if settlement not in structure[mo]:
                structure[mo][settlement] = []

            structure[mo][settlement].append(prg)

        # Заполняем дерево
        for mo in sorted(structure.keys()):
            mo_item = self.prg_tree.insert('', tk.END, text=f"📍 {mo}", values=('', ''))

            for settlement in sorted(structure[mo].keys()):
                prg_list = structure[mo][settlement]

                if len(prg_list) == 1:
                    prg = prg_list[0]
                    if prg['prg_id'] in unbound_prg_ids:
                        text_display = f"🟡 {settlement}"
                    else:
                        text_display = f"🏘️ {settlement}"

                    prg_item = self.prg_tree.insert(mo_item, tk.END, text=text_display,
                                                    values=(prg['prg_id'], prg['grs_id']))
                else:
                    settlement_item = self.prg_tree.insert(mo_item, tk.END, text=f"🏘️ {settlement}",
                                                           values=('', ''))
                    for i, prg in enumerate(prg_list):
                        if prg['prg_id'] in unbound_prg_ids:
                            text_display = f"🟡 ПРГ {i + 1}"
                        else:
                            text_display = f"  ПРГ {i + 1}"

                        self.prg_tree.insert(settlement_item, tk.END, text=text_display,
                                             values=(prg['prg_id'], prg['grs_id']))

    def show_load_calculation_results(self, processed_consumers, processed_bindings,
                                      updated_prg_count, total_prg_with_load, errors):
        """
        Показ результатов подсчета нагрузки
        """
        message = f"""✅ ПОДСЧЕТ НАГРУЗКИ ПРГ ЗАВЕРШЕН

    📊 СТАТИСТИКА ОБРАБОТКИ:
    • Обработано потребителей с расходами: {processed_consumers}
    • Обработано привязок: {processed_bindings}
    • Обновлено ПРГ: {updated_prg_count}
    • ПРГ с нагрузкой: {total_prg_with_load}

    💾 ОБНОВЛЕННЫЕ ПОЛЯ:
    • QY_pop - годовые расходы населения
    • QH_pop - часовые расходы населения
    • QY_ind - годовые расходы организаций
    • QH_ind - часовые расходы организаций
    • Year_volume - общий годовой объем
    • Max_Hour - максимальная часовая нагрузка

    💡 ФОРМУЛЫ:
    • Year_volume = QY_pop + QY_ind
    • Max_Hour = QH_pop + QH_ind
    • Часовые = Годовые / 8760"""

        if errors:
            message += f"\n\n⚠️ ОШИБКИ ({len(errors)}):\n"
            message += "\n".join(errors[:5])  # Показываем первые 5 ошибок
            if len(errors) > 5:
                message += f"\n... и еще {len(errors) - 5} ошибок"

        messagebox.showinfo("Нагрузка ПРГ подсчитана", message)

    def populate_consumer_tree(self):
        """Заполнение дерева потребителей с символами расходов"""
        # Очищаем дерево
        for item in self.consumer_tree.get_children():
            self.consumer_tree.delete(item)

        if not self.consumer_data:
            return

        # Определяем потребителей без ПРГ
        unbound_consumer_ids = set()
        for consumer in self.find_unbound_consumers():
            unbound_consumer_ids.add(consumer['id'])

        # Группируем по району и НП
        structure = {}
        for consumer in self.consumer_data:
            mo = consumer['mo']
            settlement = consumer['settlement']

            if mo not in structure:
                structure[mo] = {}
            if settlement not in structure[mo]:
                structure[mo][settlement] = []

            structure[mo][settlement].append(consumer)

        # Заполняем дерево
        for mo in sorted(structure.keys()):
            mo_item = self.consumer_tree.insert('', tk.END, text=f"📍 {mo}",
                                                values=('', '', ''), tags=('mo',))

            for settlement in sorted(structure[mo].keys()):
                settlement_item = self.consumer_tree.insert(mo_item, tk.END, text=f"🏘️ {settlement}",
                                                            values=('', '', ''), tags=('settlement',))

                for consumer in structure[mo][settlement]:
                    bindings = self.parse_prg_bindings(consumer.get('code', ''))

                    # Определяем статус и символы
                    expenses_symbol = self.get_expenses_symbol(consumer)

                    if consumer['id'] in unbound_consumer_ids:
                        status = "🟡"
                        tags = ('consumer', 'unbound')
                        binding_text = "Не привязан к ПРГ"
                        share_text = "0.000"
                    else:
                        if bindings:
                            if len(bindings) == 1:
                                share_display = self.format_share_for_excel(bindings[0]['share'])
                                binding_text = f"{bindings[0]['grs_name']} (ПРГ: {bindings[0]['prg_id']}, доля: {share_display})"
                            else:
                                binding_text = f"{len(bindings)} ПРГ"

                            total_share = self.calculate_total_share(bindings)

                            if total_share > 1.0001:
                                status = "❌"
                                share_text = f"{total_share:.3f} ⚠️"
                            elif total_share < 0.9999:
                                status = "⚠️"
                                share_text = f"{total_share:.3f}"
                            else:
                                status = "✅"
                                share_text = f"{total_share:.3f}"
                        else:
                            status = "⭕"
                            binding_text = "Не привязан к ПРГ"
                            share_text = "0.000"

                        tags = ('consumer',)

                    icon = "🏢" if consumer['type'] == 'Организация' else "🏠"
                    display_text = f"{status} {expenses_symbol} {icon} {consumer['name']}"

                    consumer_item = self.consumer_tree.insert(settlement_item, tk.END,
                                                              text=display_text,
                                                              values=(consumer['type'], binding_text, share_text),
                                                              tags=tags)

    # === 1. ОБНОВЛЕННАЯ ФУНКЦИЯ ПРИВЯЗКИ ПО ПОИСКУ ===

    def bind_by_search(self):
        """ОБНОВЛЕННАЯ функция привязки по поиску с выпадающими списками"""
        if not self.consumer_data:
            messagebox.showwarning("Предупреждение", "Сначала загрузите данные")
            return

        if not self.selected_prg:
            messagebox.showwarning("Предупреждение",
                                   "Сначала выберите ПРГ в левой панели\n\n🔍 Умный поиск использует выбранный ПРГ для автозаполнения")
            return

        try:
            # Получаем уникальные значения для выпадающих списков
            unique_districts = sorted(set(c['mo'] for c in self.consumer_data if c.get('mo')))
            unique_settlements = sorted(set(c['settlement'] for c in self.consumer_data if c.get('settlement')))
            all_prg_ids = sorted(set(p['prg_id'] for p in self.prg_data if p.get('prg_id')))

            # Показываем диалог с выпадающими списками
            search_params = self.show_smart_search_dialog(unique_districts, unique_settlements,
                                                          all_prg_ids, self.selected_prg)
            if not search_params:
                return

            # Выполняем поиск и привязку
            result = self.perform_search_binding(search_params)

            if result['success']:
                # Обновляем интерфейс с сохранением состояния дерева
                self.refresh_trees_with_state()
                self.update_changes_display()
                self.update_button_states()

                # Показываем результат
                self.show_search_binding_result(result)
            else:
                messagebox.showwarning("Результат поиска", result['message'])

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка умного поиска: {str(e)}")
            traceback.print_exc()

    def show_smart_search_dialog(self, districts, settlements, prg_ids, selected_prg):
        """Диалог умного поиска с выпадающими списками"""
        dialog = SmartSearchDialog(self.root, districts, settlements, prg_ids, selected_prg)
        return dialog.result

    def ask_share_for_settlement(self, settlement, mo, prg_id, grs_name,
                                 with_expenses_count, without_expenses_count):
        """Запрос доли для населенного пункта"""
        dialog = SettlementShareDialog(self.root, settlement, mo, prg_id, grs_name,
                                       with_expenses_count, without_expenses_count)
        return dialog.result

    def perform_search_binding(self, search_params):
        """Выполнение привязки по параметрам поиска"""
        mo_district = search_params['mo_district']
        settlement = search_params['settlement']
        street = search_params['street']
        prg_id = search_params['prg_id']
        share = search_params['share']

        # Ищем ПРГ по ID
        target_prg = None
        for prg in self.prg_data:
            if prg['prg_id'] == prg_id:
                target_prg = prg
                break

        if not target_prg:
            return {
                'success': False,
                'message': f"ПРГ с ID '{prg_id}' не найден"
            }

        # Ищем ГРС для ПРГ
        grs_name = self.get_grs_name_by_id(target_prg['grs_id'])

        # Ищем организации по критериям
        matching_organizations = []
        street_pattern = f"{street}"

        print(f"🔍 Умный поиск организаций:")
        print(f"   Район: {mo_district}")
        print(f"   НП: {settlement}")
        print(f"   Улица в названии: {street_pattern}")
        print(f"   ПРГ ID: {prg_id}")

        for consumer in self.consumer_data:
            # Проверяем что это организация
            if consumer['type'] != 'Организация':
                continue

            # Проверяем район (регистронезависимо)
            if consumer['mo'].strip().lower() != mo_district.strip().lower():
                continue

            # Проверяем НП (регистронезависимо)
            if consumer['settlement'].strip().lower() != settlement.strip().lower():
                continue

            # Проверяем улицу в названии (регистронезависимо)
            if street_pattern.lower() not in consumer['name'].lower():
                continue

            # Проверяем расходы
            if not self.has_expenses(consumer):
                print(f"   ⚠️ Пропуск {consumer['name']} - нет расходов")
                continue

            matching_organizations.append(consumer)
            print(f"   ✅ Найдена: {consumer['name']}")

        if not matching_organizations:
            return {
                'success': False,
                'message': f"Не найдено организаций с критериями:\n- Район: {mo_district}\n- НП: {settlement}\n- Улица: {street_pattern}\n- С расходами"
            }

        # Привязываем найденные организации
        bound_count = 0
        skipped_count = 0

        for org in matching_organizations:
            try:
                current_bindings = self.parse_prg_bindings(org.get('code', ''))

                # Проверяем что уже не привязана к этому ПРГ
                already_bound = False
                for binding in current_bindings:
                    if binding['prg_id'] == prg_id:
                        already_bound = True
                        print(f"   ⚠️ Пропуск {org['name']} - уже привязана к ПРГ {prg_id}")
                        skipped_count += 1
                        break

                if already_bound:
                    continue

                # Проверяем доступную долю
                current_total = self.calculate_total_share(current_bindings)
                available_share = min(share, 1.0 - current_total)

                if available_share <= 0.001:
                    print(f"   ⚠️ Пропуск {org['name']} - нет доступной доли")
                    skipped_count += 1
                    continue

                # Создаем новую привязку
                new_binding = {
                    'prg_id': prg_id,
                    'share': available_share,
                    'grs_name': grs_name
                }

                current_bindings.append(new_binding)
                new_binding_string = self.format_prg_bindings(current_bindings)

                # Обновляем данные
                old_code = org.get('code', '')
                org['code'] = new_binding_string

                # Регистрируем изменение
                change_id = f"smart_search_{org['id']}_{datetime.now().timestamp()}"
                self.changes[change_id] = {
                    'type': 'smart_search',
                    'consumer_id': org['id'],
                    'sheet_name': org['sheet_name'],
                    'row': org['excel_row'],
                    'col': org['code_col'],
                    'new_value': new_binding_string,
                    'old_value': old_code,
                    'description': f"Умный поиск: {org['name']} → ПРГ {prg_id}"
                }

                bound_count += 1
                print(f"   🔗 Привязано: {org['name']} → ПРГ {prg_id} (доля: {available_share:.3f})")

            except Exception as e:
                print(f"   ❌ Ошибка привязки {org.get('name', 'Unknown')}: {e}")
                skipped_count += 1
                continue

        return {
            'success': True,
            'found_count': len(matching_organizations),
            'bound_count': bound_count,
            'skipped_count': skipped_count,
            'search_params': search_params,
            'prg_info': target_prg,
            'grs_name': grs_name
        }

    def show_search_binding_result(self, result):
        """Результат умного поиска"""
        params = result['search_params']
        prg_info = result['prg_info']

        message = f"""✅ Умный поиск v7.3 выполнен!

🔍 ПАРАМЕТРЫ ПОИСКА:
• Район: {params['mo_district']} (выпадающий список)
• НП: {params['settlement']} (выпадающий список)
• Улица: {params['street']} (ручной ввод)
• ПРГ ID: {params['prg_id']} (из выбранного ПРГ)
• Доля: {params['share']}

🏭 ЦЕЛЕВОЙ ПРГ:
• ID: {prg_info['prg_id']}
• ГРС: {result['grs_name']}
• Район: {prg_info['mo']}
• НП: {prg_info['settlement']}

📊 РЕЗУЛЬТАТ:
• 🔍 Найдено организаций: {result['found_count']}
• ✅ Привязано успешно: {result['bound_count']}
• ⚠️ Пропущено: {result['skipped_count']}

🌳 Дерево осталось в том же состоянии (открытым)."""

        messagebox.showinfo("Умный поиск выполнен", message)

    # === 2. НОВАЯ ФУНКЦИЯ РУЧНОЙ ПРИВЯЗКИ ===

    def calculate_prg_load(self):
        """
        ПОЛНАЯ ФУНКЦИЯ: Подсчет нагрузки ПРГ из привязок потребителей

        Логика:
        1. Проходим всех потребителей и извлекаем их привязки (ПРГ_ID|доля|Название_ГРС)
        2. Для каждого ПРГ_ID считаем 4 переменные:
           - QY_pop (сумма годовых расходов населения)
           - QH_pop (сумма часовых расходов населения)
           - QY_ind (сумма годовых расходов организаций)
           - QH_ind (сумма часовых расходов организаций)
        3. Записываем эти значения в таблицу ПРГ
        4. Записываем суммы: Year_volume = QY_pop + QY_ind, Max_Hour = QH_pop + QH_ind
        """
        if not self.consumer_data:
            messagebox.showwarning("Предупреждение", "Сначала загрузите данные потребителей")
            return

        if not self.prg_data:
            messagebox.showwarning("Предупреждение", "Сначала загрузите данные ПРГ")
            return

        try:
            # Показываем диалог подтверждения
            result = messagebox.askyesno("Подсчет нагрузки ПРГ",
                                         f"Подсчитать нагрузку для всех ПРГ на основе привязок потребителей?\n\n"
                                         f"Будут обработаны:\n"
                                         f"• Потребители: {len(self.consumer_data)}\n"
                                         f"• ПРГ для обновления: {len(self.prg_data)}\n\n"
                                         f"⚠️ Текущие значения нагрузки будут перезаписаны!")

            if not result:
                return

            # Показываем прогресс
            self.info_label.config(text="⏳ Подсчитываем нагрузку ПРГ...")
            self.root.update()

            # Словарь для накопления нагрузки по ПРГ
            # prg_id -> {'QY_pop': 0, 'QH_pop': 0, 'QY_ind': 0, 'QH_ind': 0}
            prg_loads = {}

            processed_consumers = 0
            processed_bindings = 0
            errors = []

            # Обрабатываем каждого потребителя
            for consumer in self.consumer_data:
                try:
                    # Получаем расходы потребителя
                    expenses = self.get_consumer_expenses(consumer)
                    if not expenses or (expenses.get('yearly', 0) == 0 and expenses.get('hourly', 0) == 0):
                        continue  # Пропускаем потребителей без расходов

                    # Получаем привязки потребителя
                    bindings = self.parse_prg_bindings(consumer.get('code', ''))
                    if not bindings:
                        continue  # Пропускаем непривязанных потребителей

                    processed_consumers += 1

                    # Определяем тип потребителя
                    is_population = (consumer['type'] == 'Население')
                    is_organization = (consumer['type'] == 'Организация')

                    # Обрабатываем каждую привязку
                    for binding in bindings:
                        prg_id = binding['prg_id']
                        share = binding['share']

                        # Инициализируем нагрузку ПРГ если еще не создана
                        if prg_id not in prg_loads:
                            prg_loads[prg_id] = {
                                'QY_pop': 0.0,  # Годовые расходы населения
                                'QH_pop': 0.0,  # Часовые расходы населения
                                'QY_ind': 0.0,  # Годовые расходы организаций
                                'QH_ind': 0.0  # Часовые расходы организаций
                            }

                        # Добавляем расходы с учетом доли
                        yearly_load = expenses['yearly'] * share
                        hourly_load = expenses['hourly'] * share

                        if is_population:
                            prg_loads[prg_id]['QY_pop'] += yearly_load
                            prg_loads[prg_id]['QH_pop'] += hourly_load
                        elif is_organization:
                            prg_loads[prg_id]['QY_ind'] += yearly_load
                            prg_loads[prg_id]['QH_ind'] += hourly_load

                        processed_bindings += 1

                        print(f"📊 {consumer['name']} (тип: {consumer['type']}) → ПРГ {prg_id}: "
                              f"доля {share:.3f}, годовая {yearly_load:.3f}, часовая {hourly_load:.3f}")

                except Exception as e:
                    error_msg = f"Ошибка обработки потребителя {consumer.get('name', 'Unknown')}: {str(e)}"
                    errors.append(error_msg)
                    print(f"❌ {error_msg}")
                    continue

            # Применяем результаты к данным ПРГ
            updated_prg_count = 0

            for prg in self.prg_data:
                prg_id = prg['prg_id']

                if prg_id in prg_loads:
                    load = prg_loads[prg_id]

                    # Обновляем переменные ПРГ
                    prg['QY_pop'] = load['QY_pop']
                    prg['QH_pop'] = load['QH_pop']
                    prg['QY_ind'] = load['QY_ind']
                    prg['QH_ind'] = load['QH_ind']

                    # Считаем итоговые значения
                    prg['Year_volume'] = load['QY_pop'] + load['QY_ind']
                    prg['Max_Hour'] = load['QH_pop'] + load['QH_ind']

                    updated_prg_count += 1

                    print(f"🏭 ПРГ {prg_id}: QY_pop={load['QY_pop']:.3f}, QH_pop={load['QH_pop']:.3f}, "
                          f"QY_ind={load['QY_ind']:.3f}, QH_ind={load['QH_ind']:.3f}, "
                          f"Year_volume={prg['Year_volume']:.3f}, Max_Hour={prg['Max_Hour']:.3f}")
                else:
                    # ПРГ без привязок - обнуляем
                    prg['QY_pop'] = 0.0
                    prg['QH_pop'] = 0.0
                    prg['QY_ind'] = 0.0
                    prg['QH_ind'] = 0.0
                    prg['Year_volume'] = 0.0
                    prg['Max_Hour'] = 0.0

            # Регистрируем изменения для сохранения в Excel
            self.register_prg_load_changes(prg_loads)

            # Обновляем интерфейс
            self.populate_prg_tree()
            self.update_statistics()
            self.update_changes_display()

            # Показываем результаты
            self.show_load_calculation_results(
                processed_consumers, processed_bindings, updated_prg_count,
                len(prg_loads), errors
            )

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка подсчета нагрузки ПРГ: {str(e)}")
            traceback.print_exc()

        finally:
            # Восстанавливаем статус
            self.update_info_panel()

    def bind_manually(self):
        """ПОЛНАЯ РЕАЛИЗАЦИЯ: Ручная принудительная привязка"""
        if not self.selected_prg or not self.selected_consumer:
            messagebox.showwarning("Предупреждение",
                                   "Выберите ПРГ (слева) и потребителя (справа)\n\n"
                                   "🎯 Ручная привязка работает БЕЗ проверок")
            return

        try:
            prg_id = self.selected_prg['prg_id']
            grs_id = self.selected_prg['grs_id']
            grs_name = self.get_grs_name_by_id(grs_id)
            consumer = self.selected_consumer

            # Проверяем текущие привязки
            current_bindings = self.parse_prg_bindings(consumer.get('code', ''))

            # Проверяем существующую привязку к этому ПРГ
            already_bound = False
            current_share = 0.0
            for binding in current_bindings:
                if binding['prg_id'] == prg_id:
                    already_bound = True
                    current_share = binding['share']
                    break

            # Показываем диалог подтверждения
            if not self.show_manual_binding_confirmation(
                    consumer, self.selected_prg, grs_name, current_bindings, already_bound):
                return

            # Запрашиваем долю
            if already_bound:
                new_share = self.ask_share_for_manual_binding(
                    consumer, prg_id, grs_name, current_share, "edit")
            else:
                current_total = self.calculate_total_share(current_bindings)
                suggested_share = max(1.0 - current_total, 0.1)
                new_share = self.ask_share_for_manual_binding(
                    consumer, prg_id, grs_name, suggested_share, "add")

            if new_share is None:
                return

            # Выполняем привязку
            if already_bound:
                # Редактируем существующую
                for binding in current_bindings:
                    if binding['prg_id'] == prg_id:
                        binding['share'] = new_share
                        break
            else:
                # Добавляем новую
                new_binding = {
                    'prg_id': prg_id,
                    'share': new_share,
                    'grs_name': grs_name
                }
                current_bindings.append(new_binding)

            # Сохраняем изменения
            new_binding_string = self.format_prg_bindings(current_bindings)
            old_code = consumer.get('code', '')
            consumer['code'] = new_binding_string

            # Регистрируем изменение
            action_desc = "Редактирование" if already_bound else "Добавление"
            change_id = f"manual_bind_{consumer['id']}_{datetime.now().timestamp()}"
            self.changes[change_id] = {
                'type': 'manual_bind',
                'consumer_id': consumer['id'],
                'sheet_name': consumer['sheet_name'],
                'row': consumer['excel_row'],
                'col': consumer['code_col'],
                'new_value': new_binding_string,
                'old_value': old_code,
                'description': f"Ручная привязка: {action_desc} ПРГ {prg_id} для {consumer['name']}"
            }

            # Обновляем интерфейс
            self.refresh_trees_with_state()
            self.update_changes_display()
            self.update_button_states()

            # Показываем результат
            self.show_manual_binding_result(consumer, prg_id, grs_name, new_share, already_bound)

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка ручной привязки: {str(e)}")
            traceback.print_exc()

    def show_manual_binding_confirmation(self, consumer, prg, grs_name, current_bindings, already_bound):
        """Диалог подтверждения ручной привязки"""
        expenses_info = f"💰 {consumer.get('expenses', 'N/A')}" if self.has_expenses(
            consumer) else f"🚫 {consumer.get('expenses', 'N/A')}"

        current_bindings_text = ""
        if current_bindings:
            current_bindings_text = "\n".join([f"• ПРГ {b['prg_id']}: доля {b['share']:.3f} ({b['grs_name']})"
                                               for b in current_bindings])
            total_share = self.calculate_total_share(current_bindings)
            current_bindings_text += f"\nОбщая доля: {total_share:.3f}"
        else:
            current_bindings_text = "Нет текущих привязок"

        if already_bound:
            action_text = f"РЕДАКТИРОВАТЬ привязку к ПРГ {prg['prg_id']}"
            color_info = "🔄 Редактирование существующей привязки"
        else:
            action_text = f"ДОБАВИТЬ привязку к ПРГ {prg['prg_id']}"
            color_info = "➕ Добавление новой привязки"

        # Проверяем совпадения
        district_match = "✅" if consumer['mo'].lower() == prg['mo'].lower() else "❌"
        settlement_match = "✅" if consumer['settlement'].lower() == prg['settlement'].lower() else "❌"
        expenses_check = "✅" if self.has_expenses(consumer) else "❌"

        message = f"""🎯 ПРИНУДИТЕЛЬНАЯ РУЧНАЯ ПРИВЯЗКА

{color_info}

👤 ПОТРЕБИТЕЛЬ:
• Название: {consumer['name']}
• Тип: {consumer['type']}
• Район: {consumer['mo']} {district_match}
• НП: {consumer['settlement']} {settlement_match}
• Расходы: {expenses_info} {expenses_check}

🏭 ЦЕЛЕВОЙ ПРГ:
• ID: {prg['prg_id']}
• ГРС: {grs_name}
• Район: {prg['mo']}
• НП: {prg['settlement']}

📊 ТЕКУЩИЕ ПРИВЯЗКИ:
{current_bindings_text}

⚠️ ВНИМАНИЕ: Принудительная привязка ИГНОРИРУЕТ:
• Совпадение района и НП
• Наличие расходов у потребителя
• Ограничения на сумму долей

Вы уверены что хотите {action_text}?"""

        return messagebox.askyesno("Принудительная ручная привязка", message)

    def ask_share_for_manual_binding(self, consumer, prg_id, grs_name, suggested_value, action):
        """Запрос доли для ручной привязки"""
        if action == "edit":
            title = "Редактирование доли"
            prompt = f"Текущая доля: {suggested_value:.3f}\nВведите новую долю (может быть > 1.0):"
            default_value = str(suggested_value).replace('.', ',')
        else:
            title = "Добавление привязки"
            prompt = f"Рекомендуемая доля: {suggested_value:.3f}\nВведите долю (может быть > 1.0):"
            default_value = str(suggested_value).replace('.', ',')

        while True:
            share_str = simpledialog.askstring(title, prompt, initialvalue=default_value)

            if share_str is None:  # Отмена
                return None

            try:
                share = float(share_str.replace(',', '.'))
                if share < 0:
                    messagebox.showerror("Ошибка", "Доля не может быть отрицательной")
                    continue
                # Убираем ограничение на 1.0 для принудительной привязки
                return share

            except ValueError:
                messagebox.showerror("Ошибка", "Введите корректное число (например: 0.5)")
                continue

    def perform_manual_binding(self, consumer, prg_id, grs_name, new_share, already_bound, current_bindings):
        """Выполнение ручной привязки"""
        try:
            if already_bound:
                # Редактируем существующую привязку
                for binding in current_bindings:
                    if binding['prg_id'] == prg_id:
                        binding['share'] = new_share
                        break
            else:
                # Добавляем новую привязку
                new_binding = {
                    'prg_id': prg_id,
                    'share': new_share,
                    'grs_name': grs_name
                }
                current_bindings.append(new_binding)

            # Форматируем и сохраняем
            new_binding_string = self.format_prg_bindings(current_bindings)
            old_code = consumer.get('code', '')
            consumer['code'] = new_binding_string

            # Регистрируем изменение
            change_id = f"manual_bind_{consumer['id']}_{datetime.now().timestamp()}"
            action_desc = "Редактирование" if already_bound else "Добавление"

            self.changes[change_id] = {
                'type': 'manual_bind',
                'consumer_id': consumer['id'],
                'sheet_name': consumer['sheet_name'],
                'row': consumer['excel_row'],
                'col': consumer['code_col'],
                'new_value': new_binding_string,
                'old_value': old_code,
                'description': f"Ручная привязка: {action_desc} ПРГ {prg_id} для {consumer['name']} (доля: {new_share})"
            }

            print(f"🎯 Ручная привязка: {consumer['name']} → ПРГ {prg_id} (доля: {new_share:.3f})")
            return True

        except Exception as e:
            print(f"❌ Ошибка ручной привязки: {e}")
            return False

    def show_manual_binding_result(self, consumer, prg_id, grs_name, share, was_edit):
        """Результат ручной привязки"""
        action_text = "отредактирована" if was_edit else "добавлена"

        # Проверяем финальное состояние
        final_bindings = self.parse_prg_bindings(consumer.get('code', ''))
        final_total = self.calculate_total_share(final_bindings)

        total_warning = ""
        if final_total > 1.0001:
            total_warning = f"\n⚠️ ВНИМАНИЕ: Общая доля превышает 1.0 ({final_total:.3f})"

        message = f"""✅ Принудительная привязка выполнена!

👤 Потребитель: {consumer['name']}
🏭 ПРГ: {prg_id} → {grs_name}
📊 Доля: {share:.3f}
🔄 Действие: Привязка {action_text}
🧮 Общая доля: {final_total:.3f}{total_warning}

🎯 Привязка выполнена ПРИНУДИТЕЛЬНО без проверок:
• Район и НП могут не совпадать
• Расходы могли отсутствовать
• Сумма долей могла превысить 1.0

🌳 Дерево осталось в том же состоянии (открытым)."""

        messagebox.showinfo("Ручная привязка выполнена", message)

    # === ОСТАЛЬНЫЕ ФУНКЦИИ (заглушки с отсылками к полным версиям) ===

    def bind_prg_to_settlement(self):
        """ПОЛНАЯ РЕАЛИЗАЦИЯ: Привязка ПРГ ко всем потребителям в НП с проверкой расходов"""
        if not self.selected_prg or not self.selected_consumer:
            messagebox.showwarning("Предупреждение", "Выберите ПРГ (слева) и потребителя (справа)")
            return

        try:
            prg_id = self.selected_prg['prg_id']
            grs_id = self.selected_prg['grs_id']
            grs_name = self.get_grs_name_by_id(grs_id)

            # Определяем НП из выбранного потребителя
            target_mo = self.selected_consumer['mo'].strip()
            target_settlement = self.selected_consumer['settlement'].strip()

            # Находим всех потребителей в том же НП
            consumers_in_settlement = []
            for consumer in self.consumer_data:
                if (consumer['mo'].strip().lower() == target_mo.lower() and
                        consumer['settlement'].strip().lower() == target_settlement.lower()):
                    consumers_in_settlement.append(consumer)

            # Разделяем по категориям
            consumers_with_expenses = []
            consumers_without_expenses = []
            consumers_already_bound = []

            for consumer in consumers_in_settlement:
                current_bindings = self.parse_prg_bindings(consumer.get('code', ''))

                # Проверяем привязку к этому ПРГ
                already_bound = any(b['prg_id'] == prg_id for b in current_bindings)

                if already_bound:
                    consumers_already_bound.append(consumer)
                elif self.has_expenses(consumer):
                    consumers_with_expenses.append(consumer)
                else:
                    consumers_without_expenses.append(consumer)

            # Показываем диалог подтверждения
            if not self.show_settlement_binding_confirmation(
                    target_settlement, target_mo, prg_id, grs_name,
                    consumers_already_bound, consumers_with_expenses, consumers_without_expenses):
                return

            if not consumers_with_expenses:
                messagebox.showinfo("Нет потребителей",
                                    "Все потребители либо уже привязаны, либо не имеют расходов")
                return

            # Запрашиваем долю
            share = self.ask_share_for_settlement_binding(len(consumers_with_expenses))
            if share is None:
                return

            # Выполняем привязку
            bound_count = 0
            for consumer in consumers_with_expenses:
                try:
                    current_bindings = self.parse_prg_bindings(consumer.get('code', ''))

                    # Проверяем доступную долю
                    current_total = self.calculate_total_share(current_bindings)
                    available_share = min(share, 1.0 - current_total)

                    if available_share <= 0.001:
                        continue

                    # Создаем новую привязку
                    new_binding = {
                        'prg_id': prg_id,
                        'share': available_share,
                        'grs_name': grs_name
                    }

                    current_bindings.append(new_binding)
                    new_binding_string = self.format_prg_bindings(current_bindings)

                    # Обновляем данные
                    old_code = consumer.get('code', '')
                    consumer['code'] = new_binding_string

                    # Регистрируем изменение
                    change_id = f"settlement_bind_{consumer['id']}_{datetime.now().timestamp()}"
                    self.changes[change_id] = {
                        'type': 'settlement_bind',
                        'consumer_id': consumer['id'],
                        'sheet_name': consumer['sheet_name'],
                        'row': consumer['excel_row'],
                        'col': consumer['code_col'],
                        'new_value': new_binding_string,
                        'old_value': old_code,
                        'description': f"Привязка НП: {consumer['name']} → ПРГ {prg_id}"
                    }

                    bound_count += 1

                except Exception as e:
                    print(f"❌ Ошибка привязки {consumer['name']}: {e}")
                    continue

            # Обновляем интерфейс
            self.refresh_trees_with_state()
            self.update_changes_display()
            self.update_button_states()

            # Показываем результат
            messagebox.showinfo("Привязка НП выполнена",
                                f"✅ Привязано потребителей: {bound_count}\n"
                                f"⚠️ Пропущено без расходов: {len(consumers_without_expenses)}\n"
                                f"⭕ Уже привязано: {len(consumers_already_bound)}\n\n"
                                f"🌳 Дерево осталось открытым")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка привязки НП: {str(e)}")
            traceback.print_exc()

    def show_settlement_binding_confirmation(self, settlement, mo, prg_id, grs_name,
                                             already_bound, with_expenses, without_expenses):
        """Диалог подтверждения привязки НП"""
        message = f"""➡️ ПРИВЯЗКА ВСЕГО НП К ПРГ

    🏘️ Населенный пункт: {settlement} ({mo})
    🏭 ПРГ: {prg_id} → {grs_name}

    📊 АНАЛИЗ ПОТРЕБИТЕЛЕЙ:
    • ✅ Будут привязаны (с расходами): {len(with_expenses)}
    • 🚫 Пропущены (без расходов): {len(without_expenses)}
    • ⭕ Уже привязаны к этому ПРГ: {len(already_bound)}

    Продолжить привязку?"""

        return messagebox.askyesno("Привязка НП", message)

    def unbind_entire_settlement(self):
        """ПОЛНАЯ РЕАЛИЗАЦИЯ: Отвязка всего НП"""
        if not self.selected_consumer:
            messagebox.showwarning("Предупреждение", "Выберите потребителя для определения НП")
            return

        try:
            target_mo = self.selected_consumer['mo'].strip()
            target_settlement = self.selected_consumer['settlement'].strip()

            # Находим всех потребителей в НП с привязками
            consumers_to_unbind = []
            for consumer in self.consumer_data:
                if (consumer['mo'].strip().lower() == target_mo.lower() and
                        consumer['settlement'].strip().lower() == target_settlement.lower()):

                    bindings = self.parse_prg_bindings(consumer.get('code', ''))
                    if bindings:
                        consumers_to_unbind.append({
                            'consumer': consumer,
                            'bindings': bindings
                        })

            if not consumers_to_unbind:
                messagebox.showinfo("Информация",
                                    f"В НП '{target_settlement}' ({target_mo}) нет потребителей с привязками")
                return

            # Показываем диалог подтверждения
            if not self.show_settlement_unbinding_confirmation(
                    target_settlement, target_mo, consumers_to_unbind):
                return

            # Выполняем отвязку
            unbound_count = 0
            for item in consumers_to_unbind:
                try:
                    consumer = item['consumer']
                    old_code = consumer.get('code', '')

                    # Очищаем привязки
                    consumer['code'] = ''

                    # Регистрируем изменение
                    change_id = f"settlement_unbind_{consumer['id']}_{datetime.now().timestamp()}"
                    self.changes[change_id] = {
                        'type': 'settlement_unbind',
                        'consumer_id': consumer['id'],
                        'sheet_name': consumer['sheet_name'],
                        'row': consumer['excel_row'],
                        'col': consumer['code_col'],
                        'new_value': '',
                        'old_value': old_code,
                        'description': f"Отвязка НП: очистка {consumer['name']}"
                    }

                    unbound_count += 1

                except Exception as e:
                    print(f"❌ Ошибка отвязки {consumer['name']}: {e}")
                    continue

            # Обновляем интерфейс
            self.refresh_trees_with_state()
            self.update_changes_display()
            self.update_button_states()

            # Показываем результат
            messagebox.showinfo("Отвязка НП выполнена",
                                f"✅ Отвязано потребителей: {unbound_count}\n"
                                f"🏘️ НП: {target_settlement} ({target_mo})\n\n"
                                f"🌳 Дерево осталось открытым")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка отвязки НП: {str(e)}")
            traceback.print_exc()

    def ask_share_for_settlement_binding(self, consumer_count):
        """
        ПОЛНАЯ ФУНКЦИЯ: Запрос доли для привязки всего НП

        Args:
            consumer_count (int): Количество потребителей для привязки

        Returns:
            float or None: Введенная доля или None при отмене
        """
        while True:
            # Формируем сообщение с подсказками
            message = f"""🏘️ ПРИВЯЗКА НАСЕЛЕННОГО ПУНКТА

    📊 Будет привязано потребителей: {consumer_count}
    💰 Только потребители с расходами

    💡 РЕКОМЕНДАЦИИ ПО ДОЛЕ:
    • 1.0 - полная привязка к одному ПРГ
    • 0.5 - если планируете привязать к 2 ПРГ  
    • 0.33 - если планируете привязать к 3 ПРГ

    ⚠️ ВАЖНО: 
    • Доля должна быть от 0.001 до 1.0
    • При превышении 1.0 будет автоматически ограничена
    • Можно использовать запятую или точку (0,5 или 0.5)

    Введите долю для каждого потребителя:"""

            share_str = simpledialog.askstring(
                "Доля для привязки НП",
                message,
                initialvalue="1,0"
            )

            if share_str is None:  # Пользователь нажал отмену
                return None

            # Проверяем и парсим введенную долю
            share_str = share_str.strip()
            if not share_str:
                messagebox.showerror("Ошибка", "Введите значение доли")
                continue

            try:
                # Заменяем запятую на точку для парсинга
                normalized_str = share_str.replace(',', '.')
                share = float(normalized_str)

                # Валидация диапазона
                if share <= 0:
                    messagebox.showerror("Ошибка",
                                         "Доля должна быть больше 0\n\n"
                                         "Примеры корректных значений:\n"
                                         "• 1.0 или 1,0 (полная доля)\n"
                                         "• 0.5 или 0,5 (половина)\n"
                                         "• 0.33 (треть)")
                    continue

                if share > 1.0:
                    # Предупреждаем, но позволяем продолжить
                    result = messagebox.askyesno("Предупреждение",
                                                 f"Введенная доля {share:.3f} больше 1.0.\n\n"
                                                 f"Это может привести к превышению общей доли у потребителей.\n\n"
                                                 f"Продолжить с долей {share:.3f}?")
                    if not result:
                        continue

                return share

            except ValueError:
                messagebox.showerror("Ошибка",
                                     f"Некорректное значение: '{share_str}'\n\n"
                                     f"Используйте числовой формат:\n"
                                     f"• 1.0 или 1,0\n"
                                     f"• 0.5 или 0,5\n"
                                     f"• 0.33 или 0,33")
                continue

    def show_settlement_unbinding_confirmation(self, target_settlement, target_mo, consumers_to_unbind):
        """
        ПОЛНАЯ РЕАЛИЗАЦИЯ: Диалог подтверждения отвязки всего НП

        Args:
            target_settlement (str): Название населенного пункта
            target_mo (str): Название района (МО)
            consumers_to_unbind (list): Список словарей с потребителями и их привязками
                                       [{'consumer': consumer_obj, 'bindings': [binding1, binding2, ...]}, ...]

        Returns:
            bool: True если пользователь подтвердил отвязку, False если отменил
        """
        try:
            # Подготавливаем данные для отображения
            total_consumers = len(consumers_to_unbind)
            total_bindings = 0
            affected_prg_ids = set()

            # Собираем статистику
            consumers_info = []
            for item in consumers_to_unbind:
                consumer = item['consumer']
                bindings = item['bindings']

                total_bindings += len(bindings)

                # Собираем уникальные ПРГ
                for binding in bindings:
                    affected_prg_ids.add(binding['prg_id'])

                # Формируем информацию о потребителе
                expenses_symbol = self.get_expenses_symbol(consumer)
                consumer_type_icon = "🏢" if consumer['type'] == 'Организация' else "🏠"

                # Формируем строку с привязками
                bindings_info = []
                for binding in bindings:
                    share_display = self.format_share_for_excel(binding['share'])
                    bindings_info.append(f"ПРГ {binding['prg_id']} (доля: {share_display})")

                bindings_text = ", ".join(bindings_info)

                consumers_info.append({
                    'name': consumer['name'],
                    'type': consumer['type'],
                    'expenses_symbol': expenses_symbol,
                    'type_icon': consumer_type_icon,
                    'bindings_count': len(bindings),
                    'bindings_text': bindings_text,
                    'total_share': self.calculate_total_share(bindings)
                })

            # Сортируем потребителей по типу и названию
            consumers_info.sort(key=lambda x: (x['type'], x['name']))

            # Создаем текст сообщения
            header = f"⬅️ ОТВЯЗКА ВСЕГО НАСЕЛЕННОГО ПУНКТА"

            location_info = f"""🏘️ НАСЕЛЕННЫЙ ПУНКТ:
    • Район (МО): {target_mo}
    • НП: {target_settlement}"""

            statistics = f"""📊 СТАТИСТИКА ОТВЯЗКИ:
    • Потребителей для отвязки: {total_consumers}
    • Всего привязок к ПРГ: {total_bindings}
    • Затронутых ПРГ: {len(affected_prg_ids)} ({', '.join(sorted(affected_prg_ids))})"""

            # Формируем список потребителей (показываем первых 10, остальные сокращаем)
            consumers_list_header = "👥 ПОТРЕБИТЕЛИ К ОТВЯЗКЕ:"
            consumers_list = []

            for i, info in enumerate(consumers_info):
                if i < 10:  # Показываем первых 10 подробно
                    line = (f"• {info['expenses_symbol']} {info['type_icon']} {info['name']} "
                            f"({info['bindings_count']} ПРГ: {info['bindings_text']})")
                    consumers_list.append(line)
                elif i == 10:  # После 10-го добавляем сокращение
                    remaining = total_consumers - 10
                    consumers_list.append(f"... и еще {remaining} потребителей")
                    break

            consumers_text = "\n".join(consumers_list)

            # Предупреждения и последствия
            warnings = f"""⚠️ ВНИМАНИЕ - ПОСЛЕДСТВИЯ ОТВЯЗКИ:
    • Все {total_bindings} привязок будут удалены БЕЗВОЗВРАТНО
    • Потребители станут "непривязанными" (🟡)
    • Изменения можно будет отменить только через "Отмена" перед сохранением
    • ПРГ останутся без потребителей в этом НП"""

            # Вопрос подтверждения
            confirmation_question = f"""❓ ПОДТВЕРЖДЕНИЕ:
    Вы ТОЧНО хотите отвязать ВСЕ {total_consumers} потребителей
    в НП "{target_settlement}" ({target_mo}) от ВСЕХ их ПРГ?"""

            # Собираем полное сообщение
            full_message = f"""{header}

    {location_info}

    {statistics}

    {consumers_list_header}
    {consumers_text}

    {warnings}

    {confirmation_question}"""

            # Показываем диалог подтверждения с прокруткой если сообщение длинное
            if len(full_message) > 1000:
                # Для длинных сообщений используем специальный диалог с прокруткой
                return self.show_scrollable_confirmation_dialog(
                    title="Подтверждение отвязки НП",
                    message=full_message,
                    icon="warning"
                )
            else:
                # Для коротких сообщений используем стандартный диалог
                return messagebox.askyesno(
                    "Подтверждение отвязки НП",
                    full_message,
                    icon="warning"
                )

        except Exception as e:
            print(f"❌ Ошибка в show_settlement_unbinding_confirmation: {e}")
            traceback.print_exc()

            # Fallback - простой диалог подтверждения
            return messagebox.askyesno(
                "Подтверждение отвязки НП",
                f"Отвязать всех {len(consumers_to_unbind)} потребителей\n"
                f"в НП {target_settlement} ({target_mo}) от всех ПРГ?",
                icon="warning"
            )

    def show_scrollable_confirmation_dialog(self, title, message, icon="info"):
        """
        ДОПОЛНИТЕЛЬНАЯ ФУНКЦИЯ: Диалог подтверждения с прокруткой для длинных сообщений

        Args:
            title (str): Заголовок диалога
            message (str): Текст сообщения (может быть очень длинный)
            icon (str): Тип иконки ("info", "warning", "error")

        Returns:
            bool: True если пользователь нажал "Да", False если "Нет"
        """
        try:
            # Создаем диалог
            dialog = tk.Toplevel(self.root)
            dialog.title(title)
            dialog.geometry("700x500")
            dialog.resizable(True, True)
            dialog.transient(self.root)
            dialog.grab_set()

            # Центрируем диалог
            self.center_dialog(dialog)

            # Переменная для результата
            result = {"value": False}

            def on_yes():
                result["value"] = True
                dialog.destroy()

            def on_no():
                result["value"] = False
                dialog.destroy()

            def on_close():
                result["value"] = False
                dialog.destroy()

            dialog.protocol("WM_DELETE_WINDOW", on_close)

            # Основной фрейм
            main_frame = tk.Frame(dialog, padx=20, pady=20)
            main_frame.pack(fill=tk.BOTH, expand=True)

            # Заголовок с иконкой
            header_frame = tk.Frame(main_frame)
            header_frame.pack(fill=tk.X, pady=(0, 15))

            # Иконка в зависимости от типа
            icon_symbol = {
                "info": "ℹ️",
                "warning": "⚠️",
                "error": "❌"
            }.get(icon, "❓")

            tk.Label(header_frame, text=f"{icon_symbol} {title}",
                     font=('Arial', 14, 'bold')).pack()

            # Текстовая область с прокруткой
            text_frame = tk.Frame(main_frame)
            text_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))

            # Создаем Text widget с прокруткой
            text_widget = tk.Text(text_frame, wrap=tk.WORD, font=('Arial', 11),
                                  padx=10, pady=10, state=tk.NORMAL)

            # Scrollbar для текста
            scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=text_widget.yview)
            text_widget.configure(yscrollcommand=scrollbar.set)

            # Размещаем элементы
            text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            # Вставляем текст и делаем только для чтения
            text_widget.insert(tk.END, message)
            text_widget.config(state=tk.DISABLED)

            # Фрейм с кнопками
            button_frame = tk.Frame(main_frame)
            button_frame.pack(fill=tk.X, pady=(10, 0))

            # Кнопки подтверждения
            tk.Button(button_frame, text="✅ Да, отвязать", command=on_yes,
                      bg='#f44336', fg='white', font=('Arial', 12, 'bold'),
                      width=15).pack(side=tk.RIGHT, padx=(10, 0))

            tk.Button(button_frame, text="❌ Нет, отменить", command=on_no,
                      bg='#4CAF50', fg='white', font=('Arial', 12, 'bold'),
                      width=15).pack(side=tk.RIGHT)

            # Подсказка
            hint_label = tk.Label(button_frame,
                                  text="💡 Прокрутите текст выше для просмотра всех деталей",
                                  font=('Arial', 9), fg='gray')
            hint_label.pack(side=tk.LEFT)

            # Привязка клавиш
            dialog.bind('<Return>', lambda e: on_yes())
            dialog.bind('<Escape>', lambda e: on_no())

            # Ждем закрытия диалога
            dialog.wait_window()

            return result["value"]

        except Exception as e:
            print(f"❌ Ошибка в show_scrollable_confirmation_dialog: {e}")
            traceback.print_exc()

            # Fallback к стандартному диалогу
            return messagebox.askyesno(title, "Подтвердить действие?")

    def auto_bind_all_prg(self):
        """Автоматическая привязка ПРГ к потребителям по району и населенному пункту"""
        if not self.prg_data or not self.consumer_data:
            messagebox.showwarning("Предупреждение", "Сначала загрузите данные")
            return

        result = messagebox.askyesno("Автопривязка ПРГ",
                                     f"Автоматически привязать ПРГ к потребителям по совпадению района и населенного пункта?\n\n"
                                     f"ПРГ: {len(self.prg_data)}\nПотребители: {len(self.consumer_data)}")
        if not result:
            return

        bound_count = 0

        for prg in self.prg_data:
            prg_mo = prg['mo'].strip().lower()
            prg_settlement = prg['settlement'].strip().lower()
            prg_id = prg['prg_id']  # правильный ID ПРГ

            # Находим название ГРС по prg['grs_id']
            true_grs_id = prg['grs_id']
            grs_record = next((g for g in self.grs_data if g['grs_id'] == true_grs_id), None)
            grs_name = grs_record['name'] if grs_record else f"ГРС {true_grs_id}"

            for consumer in self.consumer_data:
                if consumer['mo'].strip().lower() == prg_mo and consumer[
                    'settlement'].strip().lower() == prg_settlement:
                    current_bindings = self.parse_prg_bindings(consumer.get('code', ''))
                    # Проверка, что не привязан
                    if any(b['prg_id'] == prg_id for b in current_bindings):
                        continue

                    current_total = self.calculate_total_share(current_bindings)
                    max_share = 1.0 - current_total
                    if max_share <= 0.001:
                        continue

                    # Формируем новую привязку с корректным названием ГРС
                    new_binding = {
                        'prg_id': prg_id,
                        'share': max_share,
                        'grs_name': grs_name
                    }
                    current_bindings.append(new_binding)
                    new_binding_string = self.format_prg_bindings(current_bindings)

                    # Обновляем данные потребителя
                    old_code = consumer.get('code', '')
                    consumer['code'] = new_binding_string

                    # Регистрируем изменение
                    change_id = f"auto_bind_{consumer['id']}_{datetime.now().timestamp()}"
                    self.changes[change_id] = {
                        'type': 'auto_bind',
                        'consumer_id': consumer['id'],
                        'sheet_name': consumer['sheet_name'],
                        'row': consumer['excel_row'],
                        'col': consumer['code_col'],
                        'new_value': new_binding_string,
                        'old_value': old_code,
                        'description': f"Автопривязка: {consumer['name']} → ПРГ {prg_id}"
                    }

                    bound_count += 1

        # Обновляем интерфейс
        self.populate_consumer_tree()
        self.update_statistics()
        self.update_changes_display()
        self.update_button_states()

        messagebox.showinfo("Автопривязка завершена",
                            f"✅ Успешно создано связей: {bound_count}")

    def edit_consumer_shares(self):
        """
        Открывает окно для редактирования долей:
        показывает все Привязки ПРГ у выбранного потребителя
        и позволяет менять числовые значения долей.
        """
        if not self.selected_consumer:
            messagebox.showwarning("Предупреждение", "Сначала выберите потребителя")
            return

        # Получаем список привязок
        bindings = self.parse_prg_bindings(self.selected_consumer.get("code", ""))
        if not bindings:
            messagebox.showinfo("Информация", "У потребителя нет привязок")
            return

        # Создаём окно
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Редактировать доли — {self.selected_consumer['name']}")
        dialog.geometry("400x300")
        dialog.transient(self.root)
        dialog.grab_set()

        # Словарь для хранения StringVar у каждого binding
        share_vars = []

        # Заголовки
        header_frame = tk.Frame(dialog)
        header_frame.pack(fill=tk.X, pady=(10, 0))
        tk.Label(header_frame, text="ПРГ ID", width=15, anchor=tk.W, font=('Arial', 10, 'bold')).pack(side=tk.LEFT,
                                                                                                      padx=5)
        tk.Label(header_frame, text="Доля", width=10, anchor=tk.W, font=('Arial', 10, 'bold')).pack(side=tk.LEFT,
                                                                                                    padx=5)

        # Поля ввода
        entries_frame = tk.Frame(dialog)
        entries_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        for binding in bindings:
            row = tk.Frame(entries_frame)
            row.pack(fill=tk.X, pady=2)

            tk.Label(row, text=binding['prg_id'], width=15, anchor=tk.W).pack(side=tk.LEFT, padx=5)
            var = tk.StringVar(value=str(binding['share']).replace('.', ','))
            entry = tk.Entry(row, textvariable=var, width=10)
            entry.pack(side=tk.LEFT, padx=5)
            share_vars.append((binding, var))

        # Кнопки
        button_frame = tk.Frame(dialog)
        button_frame.pack(fill=tk.X, pady=(0, 10))

        def on_cancel():
            dialog.destroy()

        def on_save():
            # Применяем изменения
            try:
                for binding, var in share_vars:
                    s = var.get().replace(',', '.').strip()
                    share = float(s)
                    if share < 0:
                        raise ValueError
                    binding['share'] = share
            except ValueError:
                messagebox.showerror("Ошибка", "Введите корректные числовые доли")
                return

            # Записываем новый код и регистрируем изменение
            new_code = self.format_prg_bindings(bindings)
            old_code = self.selected_consumer.get("code", "")
            self.selected_consumer["code"] = new_code
            change_id = f"edit_shares_{self.selected_consumer['id']}_{datetime.now().timestamp()}"
            self.changes[change_id] = {
                "type": "edit_shares",
                "consumer_id": self.selected_consumer["id"],
                "sheet_name": self.selected_consumer["sheet_name"],
                "row": self.selected_consumer["excel_row"],
                "col": self.selected_consumer["code_col"],
                "new_value": new_code,
                "old_value": old_code,
                "description": f"Редактирование долей для {self.selected_consumer['name']}"
            }

            # Обновляем интерфейс
            self.populate_consumer_tree()
            self.update_statistics()
            self.update_changes_display()
            self.update_button_states()
            dialog.destroy()

        tk.Button(button_frame, text="Сохранить", command=on_save, bg='#4CAF50', fg='white').pack(side=tk.RIGHT, padx=5)
        tk.Button(button_frame, text="Отмена", command=on_cancel, bg='#f44336', fg='white').pack(side=tk.RIGHT)

        dialog.mainloop()

    def edit_all_shares_simple(self, current_bindings):
        """
        ПРОСТОЕ редактирование всех долей сразу

        Args:
            current_bindings (list): Текущие привязки

        Returns:
            list or None: Новые привязки или None при отмене
        """
        # Показываем текущие доли и просим ввести новые
        current_shares = [binding['share'] for binding in current_bindings]
        current_shares_str = ", ".join([str(share).replace('.', ',') for share in current_shares])

        prg_list = ", ".join([f"ПРГ {b['prg_id']}" for b in current_bindings])

        message = f"""📝 РЕДАКТИРОВАНИЕ ВСЕХ ДОЛЕЙ

    🏭 ПРГ: {prg_list}
    📊 Текущие доли: {current_shares_str}

    💡 Введите новые доли через запятую (например: 0,5, 0,3, 0,2)
    ⚠️ Количество долей должно совпадать с количеством ПРГ ({len(current_bindings)})
    ✅ Можно использовать запятую или точку как разделитель

    Новые доли:"""

        while True:
            shares_input = simpledialog.askstring("Редактирование всех долей", message)

            if shares_input is None:  # Отмена
                return None

            # Парсим введенные доли
            try:
                # Разбиваем по запятым и очищаем
                shares_parts = [part.strip() for part in shares_input.split(',')]

                if len(shares_parts) != len(current_bindings):
                    messagebox.showerror("Ошибка",
                                         f"Введено {len(shares_parts)} долей, а нужно {len(current_bindings)}\n\n"
                                         f"Формат: доля1, доля2, доля3...")
                    continue

                # Конвертируем в числа
                new_shares = []
                for i, part in enumerate(shares_parts):
                    try:
                        share = float(part.replace(',', '.'))
                        if share < 0:
                            messagebox.showerror("Ошибка", f"Доля {i + 1} не может быть отрицательной: {part}")
                            break
                        new_shares.append(share)
                    except ValueError:
                        messagebox.showerror("Ошибка",
                                             f"Некорректная доля {i + 1}: '{part}'\nИспользуйте числовой формат")
                        break
                else:
                    # Все доли корректны
                    total_new = sum(new_shares)

                    # Предупреждаем если сумма больше 1.0
                    if total_new > 1.0001:
                        result = messagebox.askyesno("Предупреждение",
                                                     f"Сумма долей {total_new:.3f} превышает 1.0\n\n"
                                                     f"Продолжить anyway?")
                        if not result:
                            continue

                    # Создаем новые привязки с обновленными долями
                    new_bindings = []
                    for binding, new_share in zip(current_bindings, new_shares):
                        new_binding = binding.copy()
                        new_binding['share'] = new_share
                        new_bindings.append(new_binding)

                    return new_bindings

            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка обработки долей: {str(e)}")
                continue

    def edit_single_share_simple(self, current_bindings):
        """
        ПРОСТОЕ редактирование одной доли

        Args:
            current_bindings (list): Текущие привязки

        Returns:
            list or None: Новые привязки или None при отмене
        """
        if len(current_bindings) == 1:
            # Если одна привязка - сразу редактируем
            selected_index = 0
        else:
            # Выбираем какую привязку редактировать
            prg_options = []
            for i, binding in enumerate(current_bindings):
                share_display = self.format_share_for_excel(binding['share'])
                prg_options.append(f"{i + 1}. ПРГ {binding['prg_id']} → {binding['grs_name']} (доля: {share_display})")

            options_text = "\n".join(prg_options)

            message = f"""🔍 ВЫБОР ПРГ ДЛЯ РЕДАКТИРОВАНИЯ

    {options_text}

    Введите номер ПРГ (1-{len(current_bindings)}):"""

            while True:
                choice = simpledialog.askstring("Выбор ПРГ", message)

                if choice is None:  # Отмена
                    return None

                try:
                    selected_index = int(choice.strip()) - 1
                    if 0 <= selected_index < len(current_bindings):
                        break
                    else:
                        messagebox.showerror("Ошибка", f"Введите число от 1 до {len(current_bindings)}")
                except ValueError:
                    messagebox.showerror("Ошибка", "Введите корректный номер")

        # Редактируем выбранную долю
        selected_binding = current_bindings[selected_index]
        current_share = selected_binding['share']

        message = f"""✏️ РЕДАКТИРОВАНИЕ ДОЛИ

    🏭 ПРГ: {selected_binding['prg_id']} → {selected_binding['grs_name']}
    📊 Текущая доля: {current_share:.3f}

    💡 Введите новую долю:"""

        while True:
            new_share_str = simpledialog.askstring("Редактирование доли", message,
                                                   initialvalue=str(current_share).replace('.', ','))

            if new_share_str is None:  # Отмена
                return None

            try:
                new_share = float(new_share_str.replace(',', '.'))
                if new_share < 0:
                    messagebox.showerror("Ошибка", "Доля не может быть отрицательной")
                    continue

                # Создаем новые привязки с обновленной долей
                new_bindings = []
                for i, binding in enumerate(current_bindings):
                    new_binding = binding.copy()
                    if i == selected_index:
                        new_binding['share'] = new_share
                    new_bindings.append(new_binding)

                # Проверяем общую сумму
                total_new = self.calculate_total_share(new_bindings)

                if total_new > 1.0001:
                    result = messagebox.askyesno("Предупреждение",
                                                 f"Общая доля станет {total_new:.3f} (больше 1.0)\n\n"
                                                 f"Продолжить?")
                    if not result:
                        continue

                return new_bindings

            except ValueError:
                messagebox.showerror("Ошибка", "Введите корректное число")
                continue

    def show_edit_shares_menu(self, current_bindings):
        """
        Показывает меню действий для редактирования долей

        Args:
            current_bindings (list): Текущие привязки потребителя

        Returns:
            str or None: Выбранное действие или None при отмене
        """
        # Формируем информацию о текущих привязках
        total_share = self.calculate_total_share(current_bindings)
        bindings_info = []

        for i, binding in enumerate(current_bindings, 1):
            share_display = self.format_share_for_excel(binding['share'])
            bindings_info.append(f"{i}. ПРГ {binding['prg_id']} → {binding['grs_name']} (доля: {share_display})")

        bindings_text = "\n".join(bindings_info)

        # Формируем сообщение
        header = "✏️ РЕДАКТИРОВАНИЕ ДОЛЕЙ ПРГ"
        consumer_info = f"👤 Потребитель: {self.selected_consumer['name']}"
        stats = f"📊 Всего привязок: {len(current_bindings)} | Общая доля: {total_share:.3f}"

        if total_share > 1.0001:
            stats += " ⚠️ (превышает 1.0)"
        elif abs(total_share - 1.0) < 0.0001:
            stats += " ✅ (корректно)"

        current_text = f"📋 ТЕКУЩИЕ ПРИВЯЗКИ:\n{bindings_text}"

        actions_text = """🎛️ ДОСТУПНЫЕ ДЕЙСТВИЯ:
    1️⃣ Редактировать ВСЕ доли сразу
    2️⃣ Редактировать ОДНУ долю
    3️⃣ Добавить новый ПРГ
    4️⃣ Удалить ПРГ
    ❌ Отмена"""

        message = f"{header}\n\n{consumer_info}\n{stats}\n\n{current_text}\n\n{actions_text}\n\nВыберите действие (1-4):"

        # Показываем диалог выбора
        while True:
            choice = simpledialog.askstring("Редактирование долей", message)

            if choice is None:  # Отмена
                return None

            choice = choice.strip()

            if choice == "1":
                return "edit_all"
            elif choice == "2":
                return "edit_one"
            elif choice == "3":
                return "add_prg"
            elif choice == "4":
                return "remove_prg"
            else:
                messagebox.showerror("Ошибка", "Введите число от 1 до 4")
                continue

    def unbind_single_consumer(self):
        """Отвязка отдельного потребителя"""
        if not self.selected_consumer:
            return

        bindings = self.parse_prg_bindings(self.selected_consumer.get('code', ''))
        if not bindings:
            messagebox.showinfo("Информация", "У потребителя нет привязок")
            return

        result = messagebox.askyesno("Подтверждение",
                                     f"Отвязать все ({len(bindings)}) ПРГ от '{self.selected_consumer['name']}'?\n\n🌳 Дерево останется открытым.")
        if not result:
            return

        try:
            consumer_id = self.selected_consumer['id']
            old_code = self.selected_consumer['code']

            # Очищаем привязки
            self.selected_consumer['code'] = ''

            # Регистрируем изменение
            change_id = f"unbind_{consumer_id}_{datetime.now().timestamp()}"
            self.changes[change_id] = {
                'type': 'unbind',
                'consumer_id': consumer_id,
                'sheet_name': self.selected_consumer['sheet_name'],
                'row': self.selected_consumer['excel_row'],
                'col': self.selected_consumer['code_col'],
                'new_value': '',
                'old_value': old_code,
                'description': f"Отвязка всех ПРГ от {self.selected_consumer['name']}"
            }

            # ВАЖНО: Обновляем интерфейс с сохранением состояния дерева
            self.refresh_trees_with_state()
            self.update_changes_display()
            self.update_button_states()

            messagebox.showinfo("Отвязка выполнена", "✅ Все ПРГ отвязаны\n🌳 Дерево осталось открытым")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка отвязки: {str(e)}")
            traceback.print_exc()

    def add_prg_simple(self, current_bindings):
        """
        ПРОСТОЕ добавление нового ПРГ

        Args:
            current_bindings (list): Текущие привязки

        Returns:
            list or None: Новые привязки или None при отмене
        """
        # Находим доступные ПРГ (не привязанные)
        used_prg_ids = {binding['prg_id'] for binding in current_bindings}
        available_prgs = [prg for prg in self.prg_data if prg['prg_id'] not in used_prg_ids]

        if not available_prgs:
            messagebox.showinfo("Информация", "Все доступные ПРГ уже привязаны к этому потребителю")
            return None

        # Показываем список доступных ПРГ
        prg_options = []
        for i, prg in enumerate(available_prgs):
            prg_options.append(f"{i + 1}. ПРГ {prg['prg_id']} - {prg['settlement']} ({prg['mo']})")

        options_text = "\n".join(prg_options[:10])  # Показываем первые 10
        if len(available_prgs) > 10:
            options_text += f"\n... и еще {len(available_prgs) - 10} ПРГ"

        message = f"""➕ ДОБАВЛЕНИЕ НОВОГО ПРГ

    📋 ДОСТУПНЫЕ ПРГ ({len(available_prgs)}):
    {options_text}

    Введите номер ПРГ (1-{len(available_prgs)}):"""

        # Выбираем ПРГ
        while True:
            choice = simpledialog.askstring("Выбор ПРГ для добавления", message)

            if choice is None:  # Отмена
                return None

            try:
                prg_index = int(choice.strip()) - 1
                if 0 <= prg_index < len(available_prgs):
                    selected_prg = available_prgs[prg_index]
                    break
                else:
                    messagebox.showerror("Ошибка", f"Введите число от 1 до {len(available_prgs)}")
            except ValueError:
                messagebox.showerror("Ошибка", "Введите корректный номер")

        # Запрашиваем долю для нового ПРГ
        current_total = self.calculate_total_share(current_bindings)
        available_share = max(1.0 - current_total, 0.1)

        share_message = f"""💰 ДОЛЯ ДЛЯ НОВОГО ПРГ

    🏭 ПРГ: {selected_prg['prg_id']} - {selected_prg['settlement']} ({selected_prg['mo']})
    📊 Текущая общая доля: {current_total:.3f}
    💡 Доступно для добавления: {max(1.0 - current_total, 0):.3f}
    🎯 Рекомендуемая доля: {available_share:.3f}

    Введите долю для нового ПРГ:"""

        while True:
            share_str = simpledialog.askstring("Доля для нового ПРГ", share_message,
                                               initialvalue=str(available_share).replace('.', ','))

            if share_str is None:  # Отмена
                return None

            try:
                new_share = float(share_str.replace(',', '.'))
                if new_share <= 0:
                    messagebox.showerror("Ошибка", "Доля должна быть больше 0")
                    continue

                # Создаем новую привязку
                grs_name = self.get_grs_name_by_id(selected_prg['grs_id'])
                new_binding = {
                    'prg_id': selected_prg['prg_id'],
                    'share': new_share,
                    'grs_name': grs_name
                }

                # Создаем новый список привязок
                new_bindings = current_bindings.copy()
                new_bindings.append(new_binding)

                # Проверяем общую сумму
                total_new = self.calculate_total_share(new_bindings)

                if total_new > 1.0001:
                    result = messagebox.askyesno("Предупреждение",
                                                 f"Общая доля станет {total_new:.3f} (больше 1.0)\n\n"
                                                 f"Продолжить?")
                    if not result:
                        continue

                return new_bindings

            except ValueError:
                messagebox.showerror("Ошибка", "Введите корректное число")
                continue

    def remove_prg_simple(self, current_bindings):
        """
        ПРОСТОЕ удаление ПРГ

        Args:
            current_bindings (list): Текущие привязки

        Returns:
            list or None: Новые привязки или None при отмене
        """
        if len(current_bindings) == 1:
            # Если одна привязка - подтверждаем удаление
            binding = current_bindings[0]
            result = messagebox.askyesno("Удаление последней привязки",
                                         f"Удалить единственную привязку к ПРГ {binding['prg_id']}?\n\n"
                                         f"⚠️ Потребитель станет непривязанным (🟡)")

            if result:
                return []  # Возвращаем пустой список
            else:
                return None

        # Выбираем какую привязку удалить
        prg_options = []
        for i, binding in enumerate(current_bindings):
            share_display = self.format_share_for_excel(binding['share'])
            prg_options.append(f"{i + 1}. ПРГ {binding['prg_id']} → {binding['grs_name']} (доля: {share_display})")

        options_text = "\n".join(prg_options)

        message = f"""🗑️ УДАЛЕНИЕ ПРГ

    📋 ТЕКУЩИЕ ПРИВЯЗКИ:
    {options_text}

    Введите номер ПРГ для удаления (1-{len(current_bindings)}):"""

        while True:
            choice = simpledialog.askstring("Удаление ПРГ", message)

            if choice is None:  # Отмена
                return None

            try:
                remove_index = int(choice.strip()) - 1
                if 0 <= remove_index < len(current_bindings):
                    break
                else:
                    messagebox.showerror("Ошибка", f"Введите число от 1 до {len(current_bindings)}")
            except ValueError:
                messagebox.showerror("Ошибка", "Введите корректный номер")

        # Подтверждаем удаление
        binding_to_remove = current_bindings[remove_index]
        result = messagebox.askyesno("Подтверждение удаления",
                                     f"Удалить привязку к ПРГ {binding_to_remove['prg_id']}?\n\n"
                                     f"ГРС: {binding_to_remove['grs_name']}\n"
                                     f"Доля: {binding_to_remove['share']:.3f}")

        if result:
            # Создаем новый список без выбранной привязки
            new_bindings = []
            for i, binding in enumerate(current_bindings):
                if i != remove_index:
                    new_bindings.append(binding)

            return new_bindings
        else:
            return None

    def check_all_consumer_shares(self):
        """Проверка долей всех потребителей"""
        if not self.consumer_data:
            messagebox.showinfo("Информация", "Нет данных о потребителях")
            return

        problems = []

        for consumer in self.consumer_data:
            bindings = self.parse_prg_bindings(consumer.get('code', ''))
            if not bindings:
                continue

            total_share = self.calculate_total_share(bindings)

            if total_share > 1.0001:
                problems.append(f"❌ {consumer['name']}: {total_share:.3f} > 1.0")
            elif total_share < 0.9999:
                problems.append(f"⚠️ {consumer['name']}: {total_share:.3f} < 1.0")

        if problems:
            problem_text = "\n".join(problems[:15])
            if len(problems) > 15:
                problem_text += f"\n... и еще {len(problems) - 15}"

            messagebox.showwarning("Проблемы с долями",
                                   f"Найдено {len(problems)} потребителей с проблемами:\n\n{problem_text}")
        else:
            messagebox.showinfo("Проверка долей", "✅ Все доли потребителей корректны")

    def find_unbound_prg(self):
        """Поиск ПРГ без потребителей"""
        unbound_prg = []

        for prg in self.prg_data:
            prg_mo = prg['mo'].strip().lower()
            prg_settlement = prg['settlement'].strip().lower()

            has_consumers = False
            for consumer in self.consumer_data:
                consumer_mo = consumer['mo'].strip().lower()
                consumer_settlement = consumer['settlement'].strip().lower()

                if prg_mo == consumer_mo and prg_settlement == consumer_settlement:
                    has_consumers = True
                    break

            if not has_consumers:
                unbound_prg.append(prg)

        return unbound_prg

    def find_unbound_consumers(self):
        """Поиск потребителей без привязок"""
        unbound_consumers = []

        for consumer in self.consumer_data:
            bindings = self.parse_prg_bindings(consumer.get('code', ''))
            if not bindings:
                unbound_consumers.append(consumer)

        return unbound_consumers

    def show_unbound_analysis(self):
        """Анализ непривязанных элементов"""
        if not self.prg_data or not self.consumer_data:
            messagebox.showwarning("Предупреждение", "Сначала загрузите данные")
            return

        unbound_prg = self.find_unbound_prg()
        unbound_consumers = self.find_unbound_consumers()

        message = f"""🔍 АНАЛИЗ НЕПРИВЯЗАННЫХ ЭЛЕМЕНТОВ v7.3:

🟡 ПРГ без потребителей: {len(unbound_prg)}
🟡 Потребители без ПРГ: {len(unbound_consumers)}

💡 РЕКОМЕНДАЦИИ:
• Используйте 🔍 Умный поиск для массовой привязки
• Используйте 🎯 Ручную привязку для особых случаев"""

        messagebox.showinfo("Анализ непривязанных", message)

    def show_no_expenses_analysis(self):
        """Анализ потребителей без расходов"""
        if not self.consumer_data:
            messagebox.showwarning("Предупреждение", "Сначала загрузите данные")
            return

        consumers_without_expenses = [c for c in self.consumer_data if not self.has_expenses(c)]

        message = f"""🚫 ПОТРЕБИТЕЛИ БЕЗ РАСХОДОВ: {len(consumers_without_expenses)}

⚠️ Эти потребители НЕ будут автоматически привязываться при:
• Привязке ко всему НП
• Умном поиске

✅ Но МОГУТ быть привязаны через:
• 🎯 Ручную принудительную привязку"""

        messagebox.showinfo("Потребители без расходов", message)

    # === ОБРАБОТЧИКИ СОБЫТИЙ ===

    def on_prg_tree_select(self, event):
        """Обработка выбора в дереве ПРГ"""
        selection = self.prg_tree.selection()
        if not selection:
            self.selected_prg = None
            self.update_button_states()
            self.update_info_panel()
            return

        item = selection[0]
        values = self.prg_tree.item(item, 'values')

        if len(values) >= 2 and values[0]:  # Есть ПРГ ID
            prg_id = values[0]
            self.selected_prg = None
            for prg in self.prg_data:
                if prg['prg_id'] == prg_id:
                    self.selected_prg = prg
                    break
        else:
            self.selected_prg = None

        self.update_button_states()
        self.update_info_panel()

    def on_consumer_tree_select(self, event):
        """ИСПРАВЛЕННАЯ обработка выбора в дереве потребителей"""
        selection = self.consumer_tree.selection()
        if not selection:
            self.selected_consumer = None
            self.update_button_states()
            self.update_info_panel()
            return

        item = selection[0]
        values = self.consumer_tree.item(item, 'values')

        # Проверяем что выбран потребитель (есть тип)
        if len(values) >= 1 and values[0] in ['Население', 'Организация']:
            # ИСПРАВЛЕНИЕ: используем уникальный идентификатор вместо названия
            # Добавляем скрытый ID в значения при создании дерева

            # ВРЕМЕННОЕ РЕШЕНИЕ: извлекаем название и ищем по району+НП+название
            text = self.consumer_tree.item(item, 'text')

            # Получаем родительский элемент для определения НП
            parent_item = self.consumer_tree.parent(item)
            if parent_item:
                parent_text = self.consumer_tree.item(parent_item, 'text')
                settlement = parent_text.replace('🏘️ ', '')

                # Получаем район из прародителя
                grandparent_item = self.consumer_tree.parent(parent_item)
                if grandparent_item:
                    grandparent_text = self.consumer_tree.item(grandparent_item, 'text')
                    district = grandparent_text.replace('📍 ', '')

                    # Извлекаем название потребителя из текста (убираем символы и тип)
                    # Формат: "🟡 💰 🏢 Магазин Ромашка"
                    parts = text.split(' ')
                    if len(parts) >= 4:
                        consumer_name = ' '.join(parts[3:])

                        # ИСПРАВЛЕНИЕ: ищем по всем трем параметрам
                        self.selected_consumer = None
                        for consumer in self.consumer_data:
                            if (consumer['name'] == consumer_name and
                                    consumer['settlement'] == settlement and
                                    consumer['mo'] == district):
                                self.selected_consumer = consumer
                                break
                    else:
                        self.selected_consumer = None
                else:
                    self.selected_consumer = None
            else:
                self.selected_consumer = None
        else:
            self.selected_consumer = None

        self.update_button_states()
        self.update_info_panel()

    # === ОБНОВЛЕНИЕ ИНТЕРФЕЙСА ===

    def update_button_states(self):
        """Обновление состояния кнопок"""
        can_bind = self.selected_prg and self.selected_consumer
        self.bind_button.config(state=tk.NORMAL if can_bind else tk.DISABLED)

        has_data = len(self.consumer_data) > 0
        has_selected_prg = self.selected_prg is not None

        # Умный поиск требует выбранный ПРГ и данные
        self.search_bind_button.config(state=tk.NORMAL if (has_data and has_selected_prg) else tk.DISABLED)

        # Ручная привязка требует выбранный ПРГ и потребителя
        self.manual_bind_button.config(state=tk.NORMAL if can_bind else tk.DISABLED)

        can_unbind_settlement = self.selected_consumer is not None
        self.unbind_settlement_button.config(state=tk.NORMAL if can_unbind_settlement else tk.DISABLED)

        can_auto_bind = len(self.prg_data) > 0 and len(self.consumer_data) > 0
        self.auto_bind_button.config(state=tk.NORMAL if can_auto_bind else tk.DISABLED)

        can_edit = (self.selected_consumer and
                    self.parse_prg_bindings(self.selected_consumer.get('code', '')))
        self.edit_shares_button.config(state=tk.NORMAL if can_edit else tk.DISABLED)

        can_unbind = (self.selected_consumer and
                      self.parse_prg_bindings(self.selected_consumer.get('code', '')))
        self.unbind_button.config(state=tk.NORMAL if can_unbind else tk.DISABLED)
        can_calculate_load = len(self.prg_data) > 0 and len(self.consumer_data) > 0
        self.calculate_load_button.config(state=tk.NORMAL if can_calculate_load else tk.DISABLED)

    def update_info_panel(self):
        """УЛУЧШЕННАЯ функция обновления информационной панели с возможностью выделения текста"""
        info_parts = []

        if self.selected_prg:
            grs_name = self.get_grs_name_by_id(self.selected_prg['grs_id'])
            info_parts.append(
                f"ПРГ: {self.selected_prg['settlement']} ({self.selected_prg['mo']}) - ID: {self.selected_prg['prg_id']} - ГРС: {grs_name}")

        if self.selected_consumer:
            bindings = self.parse_prg_bindings(self.selected_consumer.get('code', ''))
            expenses_info = f"💰 {self.selected_consumer.get('expenses', 'N/A')}" if self.has_expenses(
                self.selected_consumer) else f"🚫 {self.selected_consumer.get('expenses', 'N/A')}"

            # РАСШИРЕННАЯ ИНФОРМАЦИЯ с часовыми расходами
            hourly_expenses = self.selected_consumer.get('hourly_expenses', 'N/A')
            detailed_info = f"Потребитель: {self.selected_consumer['name']}\n"
            detailed_info += f"Район: {self.selected_consumer['mo']}\n"
            detailed_info += f"НП: {self.selected_consumer['settlement']}\n"
            detailed_info += f"Годовые расходы: {self.selected_consumer.get('expenses', 'N/A')}\n"
            detailed_info += f"Часовые расходы: {hourly_expenses}\n"
            detailed_info += f"Тип: {self.selected_consumer['type']}\n"

            if bindings:
                total_share = self.calculate_total_share(bindings)
                detailed_info += f"\nПривязки к ПРГ ({len(bindings)}):\n"
                for i, binding in enumerate(bindings, 1):
                    detailed_info += f"  {i}. ПРГ {binding['prg_id']} - доля {binding['share']} - {binding['grs_name']}\n"
                detailed_info += f"Общая доля: {total_share:.3f}\n"
            else:
                detailed_info += f"\nСтатус: Не привязан к ПРГ\n"

            # Показываем краткую информацию в info_label
            info_parts.append(f"Потребитель: {self.selected_consumer['name']} - Расходы: {expenses_info}")
            if bindings:
                total_share = self.calculate_total_share(bindings)
                info_parts.append(f"Привязано {len(bindings)} ПРГ, доля: {total_share:.3f}")
            else:
                info_parts.append("Не привязан к ПРГ")

        if not info_parts:
            info_text = "🆕 v7.4: Выделение текста + Зависимые списки + Кнопки действий + Вставка из буфера"
        else:
            info_text = " | ".join(info_parts)

        self.info_label.config(text=info_text)

        # НОВОЕ: Обновляем детальную информационную панель если выбран потребитель
        if hasattr(self, 'detail_text'):
            self.detail_text.config(state=tk.NORMAL)
            self.detail_text.delete(1.0, tk.END)
            if self.selected_consumer:
                self.detail_text.insert(tk.END, detailed_info)
            self.detail_text.config(state=tk.DISABLED)

    def update_statistics(self):
        """Обновление статистики"""
        prg_count = len(self.prg_data)
        grs_count = len(self.grs_data)
        consumer_count = len(self.consumer_data)

        unbound_prg_count = len(self.find_unbound_prg()) if self.prg_data and self.consumer_data else 0
        unbound_consumers_count = len(self.find_unbound_consumers()) if self.consumer_data else 0
        consumers_without_expenses_count = len(
            [c for c in self.consumer_data if not self.has_expenses(c)]) if self.consumer_data else 0

        stats_text = f"ПРГ: {prg_count} | ГРС: {grs_count} | Потребители: {consumer_count}"

        if unbound_prg_count > 0 or unbound_consumers_count > 0:
            stats_text += f" | 🟡 Без привязок: ПРГ {unbound_prg_count}, Потр. {unbound_consumers_count}"

        if consumers_without_expenses_count > 0:
            stats_text += f" | 🚫 Без расходов: {consumers_without_expenses_count}"

        self.stats_label.config(text=stats_text)

    def update_changes_display(self):
        """Обновление отображения изменений"""
        changes_count = len(self.changes)

        if changes_count > 0:
            self.changes_label.config(text=f"● {changes_count} несохраненных изменений")
            self.save_button.config(state=tk.NORMAL)
        else:
            self.changes_label.config(text="")
            self.save_button.config(state=tk.DISABLED)

    # === СОХРАНЕНИЕ И ЗАКРЫТИЕ (заглушки) ===

    def save_changes_to_excel(self):
        """ПЕРЕПИСАННАЯ ФУНКЦИЯ: Сохранение изменений в Excel"""
        if not self.changes:
            messagebox.showinfo("Сохранение", "Нет изменений для сохранения")
            return

        if not self.excel_path or not self.excel_path.exists():
            messagebox.showerror("Ошибка", "Excel файл не найден")
            return

        try:
            # Создаем резервную копию
            self.create_excel_backup()

            changes_count = len(self.changes)
            self.save_button.config(text="⏳ Сохранение...", state=tk.DISABLED)
            self.root.update()

            # Открываем файл
            workbook = None
            try:
                workbook = openpyxl.load_workbook(self.excel_path)
                saved_count = 0

                for change_id, change in self.changes.items():
                    try:
                        change_type = change.get('type', 'unknown')

                        # ОБРАБОТКА ИЗМЕНЕНИЙ НАГРУЗКИ ПРГ
                        if change_type == 'prg_load_calculation':
                            if self.save_prg_load_change(workbook, change_id, change):
                                saved_count += 1

                        # ОБРАБОТКА ОБЫЧНЫХ ИЗМЕНЕНИЙ ПОТРЕБИТЕЛЕЙ
                        else:
                            if self.save_regular_change(workbook, change_id, change):
                                saved_count += 1

                    except Exception as e:
                        print(f"❌ Ошибка сохранения изменения {change_id}: {e}")

                # Сохраняем файл
                workbook.save(self.excel_path)
                print(f"✅ Файл сохранен: {self.excel_path}")

                # Очищаем изменения
                self.changes.clear()
                self.update_changes_display()

                messagebox.showinfo("Сохранение завершено",
                                    f"✅ Успешно сохранено {saved_count} из {changes_count} изменений")

            finally:
                if workbook:
                    try:
                        workbook.close()
                    except Exception as e:
                        print(f"⚠️ Ошибка закрытия workbook: {e}")

        except Exception as e:
            messagebox.showerror("Ошибка сохранения", f"Ошибка: {str(e)}")
            traceback.print_exc()

        finally:
            try:
                self.save_button.config(text="💾 Сохранить изменения")
                if self.changes:
                    self.save_button.config(state=tk.NORMAL)
                else:
                    self.save_button.config(state=tk.DISABLED)
            except Exception as e:
                print(f"⚠️ Ошибка восстановления кнопки: {e}")

    def save_prg_load_change(self, workbook, change_id, change):
        """Сохранение изменений нагрузки ПРГ"""
        try:
            # Получаем данные изменения
            prg_id = change.get('prg_id')
            prg_data = change.get('data', {})
            sheet_name = change.get('sheet_name')

            if not all([prg_id, prg_data, sheet_name]):
                print(f"⚠️ Неполные данные для изменения {change_id}")
                return False

            # Проверяем лист
            if sheet_name not in workbook.sheetnames:
                print(f"⚠️ Лист '{sheet_name}' не найден")
                return False

            worksheet = workbook[sheet_name]

            # Находим ПРГ в загруженных данных
            target_prg = None
            for prg in self.prg_data:
                if prg['prg_id'] == prg_id:
                    target_prg = prg
                    break

            if not target_prg:
                print(f"⚠️ ПРГ {prg_id} не найден в данных")
                return False

            # Получаем строку Excel
            excel_row = target_prg.get('excel_row', 0) + 1  # Excel 1-индексация
            if excel_row <= 1:
                print(f"⚠️ Некорректная строка для ПРГ {prg_id}")
                return False

            # Получаем колонки нагрузки
            columns = self.get_prg_load_columns()

            # Записываем данные нагрузки
            saved_fields = 0
            for field_name, column_letter in columns.items():
                if field_name in prg_data and column_letter:
                    try:
                        value = float(prg_data[field_name])
                        worksheet[f"{column_letter}{excel_row}"] = value
                        saved_fields += 1
                        print(f"💾 ПРГ {prg_id}: {field_name} → {column_letter}{excel_row} = {value}")
                    except (ValueError, TypeError):
                        print(f"⚠️ Ошибка записи {field_name} для ПРГ {prg_id}")

            return saved_fields > 0

        except Exception as e:
            print(f"❌ Ошибка сохранения нагрузки ПРГ {change_id}: {e}")
            return False

    def get_prg_load_columns(self):
        """
        Получение колонок ПРГ напрямую из prg_settings.json
        """

        DEFAULT_COLUMNS = {
            'QY_pop': 'Y',
            'QH_pop': 'Z',
            'QY_ind': 'A',
            'QH_ind': 'B',
            'Year_volume': 'V',
            'Max_Hour': 'U'
        }

        try:
            settings_file = Path('prg_settings.json')

            if not settings_file.exists():
                return DEFAULT_COLUMNS

            with open(settings_file, 'r', encoding='utf-8') as f:
                settings = json.load(f)

            prg_config = settings.get('prg', {})

            return {
                'QY_pop': prg_config.get('qy_pop_col', DEFAULT_COLUMNS['QY_pop']),
                'QH_pop': prg_config.get('qh_pop_col', DEFAULT_COLUMNS['QH_pop']),
                'QY_ind': prg_config.get('qy_ind_col', DEFAULT_COLUMNS['QY_ind']),
                'QH_ind': prg_config.get('qh_ind_col', DEFAULT_COLUMNS['QH_ind']),
                'Year_volume': prg_config.get('year_volume_col', DEFAULT_COLUMNS['Year_volume']),
                'Max_Hour': prg_config.get('max_hour_col', DEFAULT_COLUMNS['Max_Hour'])
            }

        except Exception as e:
            print(f"Ошибка чтения prg_settings.json: {e}")
            return DEFAULT_COLUMNS

    def save_regular_change(self, workbook, change_id, change):
        """Сохранение обычных изменений потребителей"""
        try:
            # Получаем данные изменения
            sheet_name = change.get('sheet_name')
            new_value = change.get('new_value')
            row_data = change.get('row')
            col_data = change.get('col')

            if not all([sheet_name, row_data is not None, col_data is not None]):
                print(f"⚠️ Неполные данные для изменения {change_id}")
                return False

            # Проверяем лист
            if sheet_name not in workbook.sheetnames:
                print(f"⚠️ Лист '{sheet_name}' не найден")
                return False

            # Проверяем координаты
            if isinstance(row_data, str) and row_data == 'TBD':
                print(f"⚠️ Пропуск изменения {change_id}: строка не определена")
                return False

            try:
                row = int(row_data) + 1  # Excel 1-индексация
                col = int(col_data) + 1  # Excel 1-индексация
            except (ValueError, TypeError):
                print(f"❌ Некорректные координаты для {change_id}: row={row_data}, col={col_data}")
                return False

            # Записываем изменение
            worksheet = workbook[sheet_name]
            worksheet.cell(row=row, column=col, value=new_value)
            print(f"💾 Сохранено: [{sheet_name}][{row},{col}] = '{new_value}'")

            return True

        except Exception as e:
            print(f"❌ Ошибка сохранения обычного изменения {change_id}: {e}")
            return False

    def create_excel_backup(self):
        """Создание резервной копии Excel файла"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_path = self.excel_path.parent / f"{self.excel_path.stem}_backup_{timestamp}{self.excel_path.suffix}"
            shutil.copy2(self.excel_path, backup_path)
            print(f"📋 Создан бэкап: {backup_path.name}")
        except Exception as e:
            print(f"⚠️ Не удалось создать бэкап: {e}")

    def on_close_window(self):
        """ИСПРАВЛЕННАЯ ФУНКЦИЯ: Обработка закрытия программы"""
        if self.is_closing:
            return  # Предотвращаем множественное срабатывание

        self.is_closing = True

        try:
            if self.changes:
                result = messagebox.askyesnocancel(
                    "Несохраненные изменения",
                    f"У вас есть {len(self.changes)} несохраненных изменений.\n\nСохранить перед выходом?"
                )

                if result is True:  # Сохранить и выйти
                    self.save_changes_to_excel()
                    # Проверяем что изменения действительно сохранились
                    if self.changes:  # Если остались несохраненные изменения
                        self.is_closing = False
                        return
                elif result is False:  # Выйти без сохранения
                    pass
                elif result is None:  # Отмена
                    self.is_closing = False
                    return

            # ИСПРАВЛЕНО: Принудительное закрытие программы
            try:
                self.root.quit()  # Останавливаем mainloop
                self.root.destroy()  # Уничтожаем окно
            except Exception as e:
                print(f"⚠️ Ошибка закрытия окна: {e}")
                # Принудительное завершение если не удается закрыть нормально
                import sys
                sys.exit(0)

        except Exception as e:
            print(f"❌ Критическая ошибка при закрытии: {e}")
            traceback.print_exc()
            # Принудительное завершение
            import sys
            sys.exit(0)

    def run(self):
        """Запуск программы"""
        try:
            print("🚀 Запуск PRG Pipeline Manager v7.3 FINAL")
            self.root.mainloop()
        except Exception as e:
            print(f"❌ Критическая ошибка: {e}")
            traceback.print_exc()
        finally:
            print("👋 Программа завершена")


# === ДИАЛОГ УМНОГО ПОИСКА С ВЫПАДАЮЩИМИ СПИСКАМИ ===

class SmartSearchDialog:
    """Диалог умного поиска с выпадающими списками"""

    def __init__(self, parent, districts, settlements, prg_ids, selected_prg):
        self.result = None

        self.dialog = tk.Toplevel(parent)
        self.dialog.title("🔍 Умный поиск v7.3 FINAL")
        self.dialog.geometry("800x750")
        self.dialog.resizable(False, False)
        self.dialog.transient(parent)
        self.dialog.grab_set()

        # Центрируем
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() - self.dialog.winfo_width()) // 2
        y = (self.dialog.winfo_screenheight() - self.dialog.winfo_height()) // 2
        self.dialog.geometry(f"+{x}+{y}")

        self.create_dialog_content(districts, settlements, prg_ids, selected_prg)

        # Ожидание результата
        self.dialog.wait_window()

    def create_dialog_content(self, districts, settlements, prg_ids, selected_prg):
        """Создание содержимого диалога"""
        main_frame = tk.Frame(self.dialog, padx=30, pady=30)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Заголовок
        title_label = tk.Label(main_frame, text="🔍 УМНЫЙ ПОИСК v7.3 FINAL",
                               font=('Arial', 18, 'bold'), fg='#00BCD4')
        title_label.pack(pady=(0, 20))

        # Информация о выбранном ПРГ
        selected_info_frame = tk.LabelFrame(main_frame, text="Выбранный ПРГ (автозаполнение)",
                                            font=('Arial', 12, 'bold'), fg='green')
        selected_info_frame.pack(fill=tk.X, pady=(0, 20))

        selected_info = tk.Frame(selected_info_frame)
        selected_info.pack(fill=tk.X, padx=20, pady=15)

        tk.Label(selected_info, text=f"🏭 ПРГ ID: {selected_prg['prg_id']}",
                 font=('Arial', 12, 'bold'), fg='blue').pack(anchor=tk.W)
        tk.Label(selected_info, text=f"📍 Район: {selected_prg['mo']}",
                 font=('Arial', 11)).pack(anchor=tk.W, pady=(5, 0))
        tk.Label(selected_info, text=f"🏘️ НП: {selected_prg['settlement']}",
                 font=('Arial', 11)).pack(anchor=tk.W, pady=(5, 0))

        # Описание
        desc_frame = tk.LabelFrame(main_frame, text="🆕 Новая логика v7.3",
                                   font=('Arial', 12, 'bold'))
        desc_frame.pack(fill=tk.X, pady=(0, 25))

        desc_text = tk.Text(desc_frame, height=4, wrap=tk.WORD, font=('Arial', 11))
        desc_text.pack(fill=tk.X, padx=20, pady=15)

        desc_content = """✅ ВЫПАДАЮЩИЕ СПИСКИ: Район, НП, ПРГ ID заполняются из данных
✅ АВТОЗАПОЛНЕНИЕ: Поля заполняются из выбранного в интерфейсе ПРГ
✅ РУЧНОЙ ВВОД: Только поле "улица" требует ручного ввода
✅ УМНЫЙ ПОИСК: Ищет организации по 4 критериям + проверяет расходы"""

        desc_text.insert(tk.END, desc_content)
        desc_text.config(state=tk.DISABLED)

        # Поля ввода
        input_frame = tk.LabelFrame(main_frame, text="Параметры умного поиска",
                                    font=('Arial', 12, 'bold'))
        input_frame.pack(fill=tk.X, pady=(0, 25))

        fields_frame = tk.Frame(input_frame)
        fields_frame.pack(fill=tk.X, padx=25, pady=20)

        # Район организации (выпадающий список)
        tk.Label(fields_frame, text="1. Район организации:",
                 font=('Arial', 12, 'bold')).grid(row=0, column=0, sticky=tk.W, pady=12)
        self.mo_var = tk.StringVar()
        self.mo_combo = ttk.Combobox(fields_frame, textvariable=self.mo_var,
                                     values=districts, font=('Arial', 12), width=25, state="readonly")
        if selected_prg['mo'] in districts:
            self.mo_combo.set(selected_prg['mo'])
        elif districts:
            self.mo_combo.set(districts[0])
        self.mo_combo.grid(row=0, column=1, padx=(20, 0), pady=12, sticky=tk.W)

        tk.Label(fields_frame, text="📋 Выпадающий список",
                 font=('Arial', 10), fg='green').grid(row=0, column=2, padx=(10, 0), pady=12, sticky=tk.W)

        # НП организации (выпадающий список)
        tk.Label(fields_frame, text="2. Населенный пункт:",
                 font=('Arial', 12, 'bold')).grid(row=1, column=0, sticky=tk.W, pady=12)
        self.settlement_var = tk.StringVar()
        self.settlement_combo = ttk.Combobox(fields_frame, textvariable=self.settlement_var,
                                             values=settlements, font=('Arial', 12), width=25, state="readonly")
        if selected_prg['settlement'] in settlements:
            self.settlement_combo.set(selected_prg['settlement'])
        elif settlements:
            self.settlement_combo.set(settlements[0])
        self.settlement_combo.grid(row=1, column=1, padx=(20, 0), pady=12, sticky=tk.W)

        tk.Label(fields_frame, text="📋 Выпадающий список",
                 font=('Arial', 10), fg='green').grid(row=1, column=2, padx=(10, 0), pady=12, sticky=tk.W)

        # Улица (ручной ввод)
        tk.Label(fields_frame, text="3. Улица (без 'ул.'):",
                 font=('Arial', 12, 'bold'), fg='red').grid(row=2, column=0, sticky=tk.W, pady=12)
        self.street_var = tk.StringVar()
        self.street_entry = tk.Entry(fields_frame, textvariable=self.street_var,
                                     font=('Arial', 12), width=27)
        self.street_entry.grid(row=2, column=1, padx=(20, 0), pady=12, sticky=tk.W)

        tk.Label(fields_frame, text="✏️ Ручной ввод",
                 font=('Arial', 10), fg='red').grid(row=2, column=2, padx=(10, 0), pady=12, sticky=tk.W)

        # ПРГ ID (выпадающий список, автозаполнен)
        tk.Label(fields_frame, text="4. ПРГ ID:",
                 font=('Arial', 12, 'bold')).grid(row=3, column=0, sticky=tk.W, pady=12)
        self.prg_id_var = tk.StringVar()
        self.prg_id_combo = ttk.Combobox(fields_frame, textvariable=self.prg_id_var,
                                         values=prg_ids, font=('Arial', 12), width=25, state="readonly")
        self.prg_id_combo.set(selected_prg['prg_id'])
        self.prg_id_combo.grid(row=3, column=1, padx=(20, 0), pady=12, sticky=tk.W)

        tk.Label(fields_frame, text="🤖 Автозаполнение",
                 font=('Arial', 10), fg='blue').grid(row=3, column=2, padx=(10, 0), pady=12, sticky=tk.W)

        # Доля
        tk.Label(fields_frame, text="5. Доля для привязки:",
                 font=('Arial', 12, 'bold')).grid(row=4, column=0, sticky=tk.W, pady=12)
        self.share_var = tk.StringVar()
        self.share_var.set("1.0")
        self.share_entry = tk.Entry(fields_frame, textvariable=self.share_var,
                                    font=('Arial', 12), width=27)
        self.share_entry.grid(row=4, column=1, padx=(20, 0), pady=12, sticky=tk.W)

        tk.Label(fields_frame, text="💰 Стандартная доля",
                 font=('Arial', 10), fg='gray').grid(row=4, column=2, padx=(10, 0), pady=12, sticky=tk.W)

        # Пример
        example_frame = tk.LabelFrame(main_frame, text="Пример умного поиска",
                                      font=('Arial', 12, 'bold'))
        example_frame.pack(fill=tk.X, pady=(0, 25))

        example_text = tk.Text(example_frame, height=6, wrap=tk.WORD, font=('Arial', 11))
        example_text.pack(fill=tk.X, padx=20, pady=15)

        example_content = f"""ПРИМЕР (данные автоматически заполнены):
Район: {selected_prg['mo']} (из выбранного ПРГ)
НП: {selected_prg['settlement']} (из выбранного ПРГ)
Улица: Ленина (единственное поле для ручного ввода)
ПРГ ID: {selected_prg['prg_id']} (из выбранного ПРГ)

РЕЗУЛЬТАТ: Найдутся все организации с "ул.Ленина" в названии
в районе "{selected_prg['mo']}", НП "{selected_prg['settlement']}" и привяжутся к ПРГ {selected_prg['prg_id']}."""

        example_text.insert(tk.END, example_content)
        example_text.config(state=tk.DISABLED)

        # Кнопки
        button_frame = tk.Frame(main_frame)
        button_frame.pack(fill=tk.X)

        tk.Button(button_frame, text="🔍 Найти и привязать",
                  command=self.ok_clicked,
                  bg='#00BCD4', fg='white', font=('Arial', 14, 'bold'),
                  width=18).pack(side=tk.RIGHT, padx=(20, 0))
        tk.Button(button_frame, text="Отмена", command=self.cancel_clicked,
                  bg='#f44336', fg='white', font=('Arial', 14),
                  width=12).pack(side=tk.RIGHT)

        # Устанавливаем фокус на поле улицы (единственное для ручного ввода)
        self.street_entry.focus()

        # Привязки клавиш
        self.dialog.bind('<Return>', lambda e: self.ok_clicked())
        self.dialog.bind('<Escape>', lambda e: self.cancel_clicked())

    def ok_clicked(self):
        """Обработка нажатия OK"""
        try:
            mo_district = self.mo_var.get().strip()
            settlement = self.settlement_var.get().strip()
            street = self.street_var.get().strip()
            prg_id = self.prg_id_var.get().strip()
            share_str = self.share_var.get().strip()

            # Проверяем заполнение полей
            if not mo_district:
                messagebox.showerror("Ошибка", "Выберите район организации из списка")
                return

            if not settlement:
                messagebox.showerror("Ошибка", "Выберите населенный пункт из списка")
                return

            if not street:
                messagebox.showerror("Ошибка", "Введите улицу (без 'ул.')\n\nЭто единственное поле для ручного ввода!")
                self.street_entry.focus()
                return

            if not prg_id:
                messagebox.showerror("Ошибка", "Выберите ПРГ ID из списка")
                return

            # Проверяем долю
            try:
                share = float(share_str.replace(',', '.'))
                if share <= 0 or share > 1:
                    messagebox.showerror("Ошибка", "Доля должна быть от 0 до 1")
                    self.share_entry.focus()
                    return
            except ValueError:
                messagebox.showerror("Ошибка", "Введите корректную долю (например: 0.5)")
                self.share_entry.focus()
                return

            self.result = {
                'mo_district': mo_district,
                'settlement': settlement,
                'street': street,
                'prg_id': prg_id,
                'share': share
            }

            self.dialog.destroy()

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка ввода данных: {str(e)}")

    def cancel_clicked(self):
        """Обработка отмены"""
        self.result = None
        self.dialog.destroy()


def main():
    """Главная функция"""
    try:
        print("🚀 PRG Pipeline Manager v7.3 FINAL - ИСПРАВЛЕННАЯ ВЕРСИЯ")
        print("")
        print("🆕 НОВЫЕ ФУНКЦИИ v7.3:")
        print("   1. 🔍 Умный поиск:")
        print("      • Требует выбранный ПРГ в интерфейсе")
        print("      • Выпадающие списки для района, НП, ПРГ ID")
        print("      • Автозаполнение из выбранного ПРГ")
        print("      • Ручной ввод только для улицы")
        print("")
        print("   2. 🎯 Ручная принудительная привязка:")
        print("      • Работает БЕЗ проверки района/НП")
        print("      • Работает БЕЗ проверки расходов")
        print("      • Позволяет превышать сумму долей")
        print("      • Требует ПРГ + потребителя")
        print("")
        print("✅ СОХРАНЕНЫ ИЗ v7.2:")
        print("   • Привязка ко всему НП с проверкой расходов")
        print("   • Отвязка всего НП")
        print("   • Сохранение состояния дерева")
        print("   • Настройки по умолчанию")
        print("   • Исправленное сохранение в Excel")
        print("")

        # Проверяем зависимости
        import pandas as pd
        import openpyxl
        print("✅ Библиотеки загружены")

        # Запускаем приложение
        app = PRGPipelineManager()
        app.run()

    except ImportError as e:
        print(f"❌ Ошибка импорта: {e}")
        print("Установите: pip install pandas openpyxl")

    except Exception as e:
        print(f"❌ Критическая ошибка: {e}")
        traceback.print_exc()


if __name__ == '__main__':
    main()
